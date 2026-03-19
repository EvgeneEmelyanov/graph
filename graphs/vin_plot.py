import os
import re
from dataclasses import dataclass
from typing import List, Optional, Tuple

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
from matplotlib.colors import to_rgb

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# =========================
# DEFAULTS (EDIT THESE ONCE)
# =========================

DEFAULT_EXCEL_PATH = r"D:\comb2.xlsx"
DEFAULT_SHEET = "SWEEP_2"

DEFAULT_BLOCK_RANGE = "A2:Y14"
DEFAULT_GAP_ROWS = 3

DEFAULT_SCENARIO_LABELS: Optional[List[str]] = ["SS", "D"]  # legend labels
CELL_LABELS_SHORT: List[str] = ["SS", "D"]  # short names shown in cells

USE_DEFAULT_LABELS = True

# =========================
# WINNER MAP SETTINGS
# =========================

TITLE_TEMPLATE = "Выигрышный сценарий (min {param})"

# Если разница между лучшим и вторым лучшим < 1% -> серый (и НИЧЕГО не пишем в ячейке)
TIE_THRESHOLD_PCT = 1.0
TIE_GRAY_RGB = np.array([0.92, 0.92, 0.92], dtype=float)

# Layout
RIGHT_MARGIN = 0.90
FIGSIZE = (14, 6)

# Grid
SHOW_CELL_GRID = True
CELL_GRID_LW = 0.35
CELL_GRID_ALPHA = 0.35

# Ticks
X_LABEL_ROTATION = 90
X_LABEL_FONTSIZE = 8
Y_LABEL_FONTSIZE = 8
BOTTOM_MARGIN = 0.32

# Cell text overlay
DRAW_CELL_LABELS = True
CELL_LABEL_FONTSIZE = 7
CELL_LABEL_ALPHA = 0.95
CELL_LABEL_EVERY_N = 1

# Robust parsing
ALLOW_ROW_PADDING_WITH_NAN = True
ALLOW_ROW_TRUNCATION = False
SKIP_FULLY_EMPTY_ROWS = True

# =========================
# Data model
# =========================

@dataclass
class Table2D:
    x_labels: List[str]
    y_labels: List[str]
    values: np.ndarray


# =========================
# Parsing helpers (tabs)
# =========================

def _split_blocks_by_tabs(line: str) -> List[str]:
    return [b for b in re.split(r"\t{2,}", line.rstrip("\n")) if b != ""]

def _split_cells_by_tabs_keep_empty(block: str) -> List[str]:
    return block.rstrip("\n").split("\t")

def _to_float_ru_or_nan(s: object) -> float:
    if s is None:
        return np.nan
    s = str(s).replace("\xa0", "").replace(" ", "").strip()
    if s == "":
        return np.nan
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return np.nan

def parse_tables_from_paste(paste_text: str) -> List[Table2D]:
    raw_lines = paste_text.splitlines()
    lines = [ln for ln in raw_lines if ln.strip()]
    if not lines:
        raise ValueError("Empty input.")

    header_blocks = _split_blocks_by_tabs(lines[0])
    if not header_blocks:
        raise ValueError("Header line could not be split into table blocks. Keep tabs in the input.")

    x_labels_per_table: List[List[str]] = []
    for hb in header_blocks:
        cells = _split_cells_by_tabs_keep_empty(hb)
        while cells and str(cells[0]).strip() == "":
            cells.pop(0)
        while cells and str(cells[-1]).strip() == "":
            cells.pop()
        x_labels_per_table.append([str(c).strip() for c in cells if str(c).strip() != ""])

    n_tables = len(x_labels_per_table)
    if n_tables == 0:
        raise ValueError("No tables detected in header.")
    n_cols = len(x_labels_per_table[0])
    if n_cols == 0:
        raise ValueError("Header has no X labels (columns).")
    if any(len(x) != n_cols for x in x_labels_per_table):
        raise ValueError("Not all tables have the same number of X columns (header mismatch).")

    y_labels: List[str] = []
    values_per_table: List[List[List[float]]] = [[] for _ in range(n_tables)]

    for ln in lines[1:]:
        blocks = _split_blocks_by_tabs(ln)
        if len(blocks) != n_tables:
            raise ValueError(
                f"Line has {len(blocks)} blocks, expected {n_tables}. "
                f"Make sure table gaps are preserved as multiple TABs."
            )

        row_y: Optional[str] = None
        row_all_empty = True

        for ti, blk in enumerate(blocks):
            cells = _split_cells_by_tabs_keep_empty(blk)

            while cells and str(cells[-1]).strip() == "":
                cells.pop()

            if len(cells) == 0:
                yv = ""
                vals = [np.nan] * n_cols
            else:
                yv = (cells[0] if cells else "")
                vals_raw = cells[1:] if len(cells) > 1 else []

                expected_vals = n_cols
                if len(vals_raw) < expected_vals:
                    if not ALLOW_ROW_PADDING_WITH_NAN:
                        raise ValueError(
                            f"Table {ti+1}: row has {1 + len(vals_raw)} cells, expected {1 + expected_vals}."
                        )
                    vals_raw = vals_raw + [""] * (expected_vals - len(vals_raw))
                elif len(vals_raw) > expected_vals:
                    if not ALLOW_ROW_TRUNCATION:
                        raise ValueError(
                            f"Table {ti+1}: row has {1 + len(vals_raw)} cells, expected {1 + expected_vals}."
                        )
                    vals_raw = vals_raw[:expected_vals]

                vals = [_to_float_ru_or_nan(v) for v in vals_raw]

            yv_str = str(yv).strip()
            if yv_str != "" or any(np.isfinite(v) for v in vals):
                row_all_empty = False

            if row_y is None:
                row_y = yv_str
            elif row_y != yv_str:
                raise ValueError("Y labels differ across tables on the same row; input misaligned.")

            values_per_table[ti].append(vals)

        if SKIP_FULLY_EMPTY_ROWS and row_all_empty:
            for ti in range(n_tables):
                values_per_table[ti].pop()
            continue

        y_labels.append(str(row_y if row_y is not None else "").strip())

    tables: List[Table2D] = []
    for ti in range(n_tables):
        arr = np.array(values_per_table[ti], dtype=float)
        tables.append(Table2D(x_labels=x_labels_per_table[ti], y_labels=y_labels, values=arr))
    return tables


# =========================
# Excel reading
# =========================

def _block_range_for_graph_index(base_range: str, graph_index: int, gap_rows: int) -> Tuple[str, int, int, int, int]:
    if graph_index < 1:
        raise ValueError("graph_index must be >= 1")
    min_col, min_row, max_col, max_row = range_boundaries(base_range)
    height = (max_row - min_row + 1)
    shift = (graph_index - 1) * (height + gap_rows)
    new_min_row = min_row + shift
    new_max_row = max_row + shift
    a1 = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
    return a1, min_col, new_min_row, max_col, new_max_row

def _cells_to_tabbed_text(ws: Worksheet, min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    lines = []
    for r in range(min_row, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append("" if v is None else str(v))
        lines.append("\t".join(row_vals))
    return "\n".join(lines)

def parse_tables_from_excel(xlsx_path: str, sheet_name: str, a1_range: str) -> Tuple[List[Table2D], str]:
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")
        ws = wb[sheet_name]

        min_col, min_row, max_col, max_row = range_boundaries(a1_range)

        title_row = min_row - 1
        title_col = min_col
        param_name = ""
        if title_row >= 1:
            v = ws.cell(row=title_row, column=title_col).value
            param_name = "" if v is None else str(v).strip()

        block_text = _cells_to_tabbed_text(ws, min_col, min_row, max_col, max_row)
        tables = parse_tables_from_paste(block_text)
        return tables, param_name
    finally:
        wb.close()


# =========================
# Winner-map computation (min objective)
# =========================

def compute_winner_second_and_pct(tables: List[Table2D]) -> Tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    """
    winner_idx[y,x] = argmin scenario, or -1 if no data
    best[y,x]       = best value
    second[y,x]     = second best value (NaN if cannot)
    pct_better[y,x] = (second-best)/|second| * 100 (for min)  -> percent that winner is better than runner-up
    valid_mask[y,x] = at least one finite value exists
    """
    if not tables:
        raise ValueError("No tables provided.")

    ny = len(tables[0].y_labels)
    nx = len(tables[0].x_labels)
    for t in tables:
        if t.values.shape != (ny, nx):
            raise ValueError("All tables must have identical shape and grids.")

    stack = np.stack([t.values for t in tables], axis=0)  # [k, ny, nx]
    finite = np.isfinite(stack)
    valid_mask = np.any(finite, axis=0)

    stack2 = np.where(finite, stack, np.inf)              # for argmin
    winner_idx = np.argmin(stack2, axis=0).astype(int)
    winner_idx[~valid_mask] = -1

    if stack2.shape[0] < 2:
        best = np.full((ny, nx), np.nan, dtype=float)
        second = np.full((ny, nx), np.nan, dtype=float)
        pct = np.full((ny, nx), np.nan, dtype=float)
        best[valid_mask] = np.min(stack2[:, valid_mask], axis=0)
        return winner_idx, best, second, pct, valid_mask

    part = np.partition(stack2, kth=1, axis=0)
    best = part[0, :, :]
    second = part[1, :, :]

    best = np.where(np.isfinite(best), best, np.nan)
    second = np.where(np.isfinite(best) & np.isfinite(second), second, np.nan)

    eps = 1e-12
    # percent that best is better than second for MIN objective:
    # (second - best) / max(|second|, eps) * 100
    pct = (second - best) / np.maximum(np.abs(second), eps) * 100.0

    best[~valid_mask] = np.nan
    second[~valid_mask] = np.nan
    pct[~valid_mask] = np.nan

    return winner_idx, best, second, pct, valid_mask


# =========================
# Plotting
# =========================

def _base_rgb_colors(k: int) -> List[np.ndarray]:
    base = plt.rcParams["axes.prop_cycle"].by_key().get("color", ["C0", "C1", "C2", "C3"])
    return [np.array(to_rgb(base[i % len(base)]), dtype=float) for i in range(k)]

def _ideal_text_color(rgb: np.ndarray) -> str:
    r, g, b = float(rgb[0]), float(rgb[1]), float(rgb[2])
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return "black" if lum > 0.6 else "white"

def plot_winner_map_with_ties(
    tables: List[Table2D],
    scenario_labels: List[str],
    title_main: str,
    param_name: str,
):
    ref = tables[0]
    k = len(tables)

    winner_idx, best, second, pct_better, valid_mask = compute_winner_second_and_pct(tables)

    ny, nx = winner_idx.shape
    img = np.ones((ny, nx, 3), dtype=float)  # default white

    colors = _base_rgb_colors(k)

    # Tie mask: pct < 1% (and valid)
    tie_mask = valid_mask & np.isfinite(pct_better) & (pct_better < TIE_THRESHOLD_PCT)

    # Paint winners (solid color), ties as gray, invalid as white
    for i in range(k):
        mask = (winner_idx == i) & valid_mask & (~tie_mask)
        if np.any(mask):
            img[mask] = colors[i]

    img[tie_mask] = TIE_GRAY_RGB
    img[~valid_mask] = 1.0

    fig, ax = plt.subplots(figsize=FIGSIZE)
    ax.imshow(img, origin="lower", aspect="auto", interpolation="nearest")

    # Grid
    if SHOW_CELL_GRID:
        ax.set_xticks(np.arange(-0.5, nx, 1), minor=True)
        ax.set_yticks(np.arange(-0.5, ny, 1), minor=True)
        ax.grid(which="minor", color="black", linewidth=CELL_GRID_LW, alpha=CELL_GRID_ALPHA)

    ax.set_xlabel("Доля II")
    ax.set_ylabel("Доля I")
    ax.set_title(title_main)

    ax.set_xticks(np.arange(len(ref.x_labels)))
    ax.set_yticks(np.arange(len(ref.y_labels)))
    ax.set_xticklabels(ref.x_labels, rotation=X_LABEL_ROTATION, ha="center", va="top", fontsize=X_LABEL_FONTSIZE)
    ax.set_yticklabels(ref.y_labels, fontsize=Y_LABEL_FONTSIZE)

    ax.tick_params(axis="x", which="both", bottom=True, top=True, labelbottom=True, length=4)
    ax.tick_params(axis="y", which="both", left=True, right=True, labelleft=True, length=4)

    # Text: first line short label, second line percent better
    if DRAW_CELL_LABELS:
        if len(CELL_LABELS_SHORT) != k:
            raise ValueError(f"CELL_LABELS_SHORT has {len(CELL_LABELS_SHORT)} items, but there are {k} scenarios.")

        for y in range(0, ny, CELL_LABEL_EVERY_N):
            for x in range(0, nx, CELL_LABEL_EVERY_N):
                if not valid_mask[y, x]:
                    continue
                if tie_mask[y, x]:
                    continue  # <- "серую ячейку" оставляем без надписей

                idx = int(winner_idx[y, x])
                if idx < 0 or idx >= k:
                    continue

                pct = float(pct_better[y, x])
                txt = f"{CELL_LABELS_SHORT[idx]}\n{pct:.1f}%"

                cell_rgb = img[y, x, :]
                color_txt = _ideal_text_color(cell_rgb)

                ax.text(
                    x, y, txt,
                    ha="center", va="center",
                    fontsize=CELL_LABEL_FONTSIZE,
                    color=color_txt,
                    alpha=CELL_LABEL_ALPHA,
                    clip_on=True,
                    linespacing=0.95,
                )

    # Legend
    legend_title = f"Сценарий с min {param_name}" if param_name else "Сценарий с min показателем"
    handles = [Patch(facecolor=colors[i], edgecolor="black", label=str(scenario_labels[i])) for i in range(k)]
    handles.append(Patch(facecolor=TIE_GRAY_RGB, edgecolor="black", label=f"Почти равно (< {TIE_THRESHOLD_PCT:.0f}%)"))

    ax.legend(
        handles=handles,
        title=legend_title,
        loc="upper left",
        bbox_to_anchor=(1.02, 1.0),
        borderaxespad=0.0,
    )

    plt.tight_layout(rect=(0, 0, RIGHT_MARGIN, 1))
    plt.subplots_adjust(bottom=BOTTOM_MARGIN)

    return fig


# =========================
# Main
# =========================

if __name__ == "__main__":
    xlsx_path = DEFAULT_EXCEL_PATH
    sheet_name = DEFAULT_SHEET

    graph_idx_s = input("Graph index (1..N) (default 1): ").strip()
    graph_idx = int(graph_idx_s) if graph_idx_s else 1
    if graph_idx < 1:
        raise ValueError("Graph index must be >= 1")

    block_range, _, _, _, _ = _block_range_for_graph_index(
        DEFAULT_BLOCK_RANGE, graph_idx, DEFAULT_GAP_ROWS
    )

    print(f"Using block range: {block_range}")

    tables, param_name = parse_tables_from_excel(xlsx_path, sheet_name, block_range)

    labels = DEFAULT_SCENARIO_LABELS if USE_DEFAULT_LABELS else None
    if labels is None:
        labels = [f"S{i+1}" for i in range(len(tables))]
    if len(labels) != len(tables):
        raise ValueError(
            f"DEFAULT_SCENARIO_LABELS has {len(labels)} items, but block contains {len(tables)} tables."
        )

    pname = param_name if param_name else "показатель"
    title = TITLE_TEMPLATE.format(param=pname)

    fig = plot_winner_map_with_ties(
        tables=tables,
        scenario_labels=labels,
        title_main=title,
        param_name=pname,
    )

    out_dir = r"D:\10_results\plots"
    os.makedirs(out_dir, exist_ok=True)
    fig_path = os.path.join(out_dir, f"winner_map_{graph_idx}.png")
    fig.savefig(fig_path, dpi=300)

    plt.show()