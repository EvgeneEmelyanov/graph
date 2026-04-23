import os
import re

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.ticker import FuncFormatter, ScalarFormatter
from openpyxl import load_workbook


# =========================================================
# НАСТРОЙКИ ФАЙЛА
# =========================================================
EXCEL_PATH = r"D:\1comb_results.xlsx"
SHEET_NAME = "SWEEP_2"
OUTPUT_DIR = r"D:\results"


# =========================================================
# ПОДПИСИ СТОЛБЦОВ
# =========================================================
HEADER_METRIC = "ПОКАЗАТЕЛЬ"
HEADER_SCHEME_1 = "Секционированная (С)\nсистема шин"
HEADER_SCHEME_2 = "Двойная (Д)\nсистема шин"
HEADER_DELTA = "Δ = (С − Д)/С ·100%"


# =========================================================
# ОБЩИЕ ПАРАМЕТРЫ ДАННЫХ
# =========================================================
PEAK_LOAD_KW = 1346.0
WT_COUNT = 2

# Обрезка сетки:
# X_TRIM = (убрать_слева, убрать_справа)
# Y_TRIM = (убрать_сверху, убрать_снизу)
X_TRIM = (0, 0)
Y_TRIM = (0, 0)


# =========================================================
# СПИСОК МЕТРИК
# =========================================================
# Формат:
# (ключ_в_excel, подпись_слева, единица_на_colorbar)
METRICS = [
    ("LCOE", "LCOE", "руб/кВт·ч"),
    ("LPSP", "LPSP", ""),
    ("LOLP", "LOLP", ""),
    ("ENS_evtN", "Деф. событий", "шт"),
    ("ENS_evtMaxH", "Макс. длительность деф.", "ч"),
]

# Для каких метрик дополнительно делать отдельные картинки
INDIVIDUAL_METRICS = [
    "LCOE",
    "LPSP",
    "LOLP",
]


# =========================================================
# АЛИАСЫ ЗАГОЛОВКОВ В EXCEL
# =========================================================
HEADER_ALIASES = {
    "LCOE": [
        "LCOE",
        "LCOE,руб/кВт∙ч",
        "LCOE, руб/кВт∙ч",
        "LCOE,руб/кВт·ч",
        "LCOE, руб/кВт·ч",
        "LCOE,руб/кВт*ч",
        "LCOE, руб/кВт*ч",
    ],
    "LPSP": [
        "LPSP",
    ],
    "LOLP": [
        "LOLP",
    ],
    "ENS_evtN": [
        "ENS_evtN",
    ],
    "ENS_evtMaxH": [
        "ENS_evtMaxH",
        "ENS_evt_MaxH",
        "ENS_evtMaxH ",
        "ENS_evt_MaxH ",
    ],
}


# =========================================================
# ГЕОМЕТРИЯ ОБЩЕЙ КАРТИНКИ
# =========================================================
# Размер всей фигуры
SUMMARY_FIG_WIDTH = 22
SUMMARY_ROW_HEIGHT = 5.2

# Базовые ширины столбцов таблицы:
# 1 столбец = подпись показателя
# 2 = график С
# 3 = график Д
# 4 = график Δ
SUMMARY_WIDTH_RATIOS = [1.1, 2.15, 2.15, 2.15]

# Высота строки заголовков относительно строк с графиками
SUMMARY_HEADER_HEIGHT_RATIO = 0.16

# Базовый зазор между столбцами таблицы
SUMMARY_WSPACE = 0.24

# Базовый вертикальный зазор между строками
SUMMARY_HSPACE = 0.16


# =========================================================
# ГЕОМЕТРИЯ ОТДЕЛЬНОЙ КАРТИНКИ
# =========================================================
INDIVIDUAL_FIG_WIDTH = 18
INDIVIDUAL_FIG_HEIGHT = 6.2
INDIVIDUAL_WIDTH_RATIOS = [1.1, 2.15, 2.15, 2.15]
INDIVIDUAL_HEADER_HEIGHT_RATIO = 0.16
INDIVIDUAL_WSPACE = 0.24
INDIVIDUAL_HSPACE = 0.08


# =========================================================
# РУЧНЫЕ СДВИГИ СТОЛБЦОВ
# =========================================================
# Это ключевой блок для ручной настройки.
#
# Как работает:
# - COL2_SHIFT_X управляет столбцом "Секционированная"
# - COL3_SHIFT_X управляет столбцом "Двойная"
# - COL4_SHIFT_X управляет столбцом "Δ"
#
# Положительное значение = сдвиг вправо
# Отрицательное значение = сдвиг влево
#
# Чтобы уменьшить зазор между С и Д:
#   COL3_SHIFT_X делай отрицательнее
#
# Чтобы увеличить зазор между Д и Δ:
#   COL4_SHIFT_X делай положительнее
#
COL1_SHIFT_X = 0.0
COL2_SHIFT_X = 0.0
COL3_SHIFT_X = 0.0
COL4_SHIFT_X = 0.03

# Если хочешь двигать colorbar вместе со столбцом:
# True  = двигать вместе
# False = оставить на месте
MOVE_CB_WITH_COL3 = True
MOVE_CB_WITH_COL4 = True


# =========================================================
# ВНЕШНИЙ ВИД HEATMAP
# =========================================================
# Для summary лучше обычно auto
SUMMARY_IMAGE_ASPECT = "auto"

# Для одиночных картинок equal даёт более квадратные клетки
INDIVIDUAL_IMAGE_ASPECT = "equal"

# Параметры colorbar
CB_FRACTION = 0.046
CB_PAD = 0.04

# Поворот подписей X
X_TICK_ROTATION = 90


# =========================================================
# РАЗМЕРЫ ШРИФТОВ
# =========================================================
AXIS_LABEL_SIZE = 14
TICK_SIZE = 12
CB_LABEL_SIZE = 13
CB_TICK_SIZE = 12
HEADER_SIZE = 16
METRIC_LABEL_SIZE = 15


# =========================================================
# ПАЛИТРЫ
# =========================================================
MATTE_METRIC_CMAP = LinearSegmentedColormap.from_list(
    "metric",
    ["#f2f2f2", "#cfd8dc", "#8ea9b8", "#5f86a2", "#426f91"]
)

MATTE_DIFF_CMAP = LinearSegmentedColormap.from_list(
    "diff",
    ["#4d7a99", "#8fa8b7", "#e9e9e9", "#ddb48c", "#c98d5b"]
)


# =========================================================
# ФОРМАТТЕРЫ
# =========================================================
def comma_formatter(decimals=2):
    def _fmt(x, pos):
        if not np.isfinite(x):
            return ""
        s = f"{x:.{decimals}f}".rstrip("0").rstrip(".")
        return s.replace(".", ",")
    return FuncFormatter(_fmt)


class CommaScalarFormatter(ScalarFormatter):
    def __call__(self, x, pos=None):
        s = super().__call__(x, pos)
        return s.replace(".", ",")


# =========================================================
# ВСПОМОГАТЕЛЬНЫЕ
# =========================================================


def normalize_text(x):
    if x is None:
        return ""
    return (
        str(x)
        .strip()
        .replace(" ", "")
        .replace("·", "∙")
        .replace("*", "∙")
    )


def to_float(v):
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return np.nan


def safe_filename(name):
    s = str(name).strip()
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    s = s.replace("\n", "_").replace("\r", "_")
    return s


def shift_axis(ax, dx=0.0, dy=0.0):
    pos = ax.get_position()
    ax.set_position([pos.x0 + dx, pos.y0 + dy, pos.width, pos.height])


def apply_trim(x, y, z):
    xl, xr = X_TRIM
    yt, yb = Y_TRIM

    xe = len(x) - xr if xr > 0 else len(x)
    ye = len(y) - yb if yb > 0 else len(y)

    return x[xl:xe], y[yt:ye], z[yt:ye, xl:xe]


def rel_diff_percent(sec, dbl):
    out = np.full_like(sec, np.nan, dtype=float)
    mask = np.abs(sec) > 1e-12
    out[mask] = (sec[mask] - dbl[mask]) / sec[mask] * 100.0
    return out


def xlabels(x):
    return [str(int(round(v))) if np.isfinite(v) else "" for v in x]


def ylabels(y):
    labels = []
    for v in y:
        if np.isfinite(v):
            labels.append(str(int(round(v * WT_COUNT / PEAK_LOAD_KW * 100))))
        else:
            labels.append("")
    return labels


def style_axis(ax, x, y, show_x=False, show_y=False):
    ax.set_xticks(np.arange(len(x)))
    ax.set_xticklabels(xlabels(x), rotation=X_TICK_ROTATION, fontsize=TICK_SIZE)

    ax.set_yticks(np.arange(len(y)))
    ax.set_yticklabels(ylabels(y), fontsize=TICK_SIZE)

    ax.tick_params(axis="x", pad=2)
    ax.tick_params(axis="y", pad=2)

    if show_x:
        ax.set_xlabel("Мощность ДГУ, кВт", fontsize=AXIS_LABEL_SIZE)

    if show_y:
        ax.set_ylabel("Мощность ВЭУ, %", fontsize=AXIS_LABEL_SIZE)


def setup_metric_colorbar(cb, key, label, unit):
    if key in ["LOLP", "LPSP"]:
        fmt = CommaScalarFormatter(useMathText=True)
        fmt.set_powerlimits((0, 0))
        cb.formatter = fmt
    else:
        cb.formatter = comma_formatter(2)

    cb.update_ticks()
    cb.ax.tick_params(labelsize=CB_TICK_SIZE)

    cbar_label = f"{label}, {unit}" if unit else label
    cb.set_label(cbar_label, fontsize=CB_LABEL_SIZE)


def setup_delta_colorbar(cb):
    cb.formatter = comma_formatter(2)
    cb.update_ticks()
    cb.ax.tick_params(labelsize=CB_TICK_SIZE)
    cb.set_label("Δ, %", fontsize=CB_LABEL_SIZE)


# =========================================================
# ПОИСК И ЧТЕНИЕ ДАННЫХ ИЗ EXCEL
# =========================================================
def find_metric(ws, key):
    aliases = HEADER_ALIASES.get(key, [key])
    aliases = [normalize_text(a) for a in aliases]

    found = []

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = normalize_text(ws.cell(r, c).value)
            if val in aliases:
                found.append((r, c))

    found = sorted(found, key=lambda rc: (rc[0], rc[1]))

    if len(found) < 2:
        raise ValueError(
            f"Для метрики '{key}' найдено блоков: {len(found)}. "
            f"Проверь подписи в Excel или добавь вариант в HEADER_ALIASES."
        )

    return found[:2]


def read_block(ws, r, c):
    xrow = r + 1
    yrow = r + 2
    xcol = c + 1

    xcols = []
    cc = xcol
    while cc <= ws.max_column and ws.cell(xrow, cc).value not in [None, ""]:
        xcols.append(cc)
        cc += 1

    yrows = []
    rr = yrow
    while rr <= ws.max_row and ws.cell(rr, c).value not in [None, ""]:
        yrows.append(rr)
        rr += 1

    if not xcols:
        raise ValueError(f"Не удалось прочитать ось X для блока в ячейке ({r}, {c})")
    if not yrows:
        raise ValueError(f"Не удалось прочитать ось Y для блока в ячейке ({r}, {c})")

    x = np.array([to_float(ws.cell(xrow, k).value) for k in xcols], dtype=float)
    y = np.array([to_float(ws.cell(k, c).value) for k in yrows], dtype=float)
    z = np.array(
        [[to_float(ws.cell(rr, cc).value) for cc in xcols] for rr in yrows],
        dtype=float
    )

    return x, y, z


def prepare_metric_data(ws):
    prepared = []

    for key, label, unit in METRICS:
        (r1, c1), (r2, c2) = find_metric(ws, key)

        x1, y1, z1 = apply_trim(*read_block(ws, r1, c1))
        x2, y2, z2 = apply_trim(*read_block(ws, r2, c2))

        d = rel_diff_percent(z1, z2)

        prepared.append(
            {
                "key": key,
                "label": label,
                "unit": unit,
                "x1": x1,
                "y1": y1,
                "z1": z1,
                "x2": x2,
                "y2": y2,
                "z2": z2,
                "diff": d,
            }
        )

    return prepared


# =========================================================
# СЛУЖЕБНАЯ ОТРИСОВКА ОДНОЙ СТРОКИ ТАБЛИЦЫ
# =========================================================
def draw_table_row(fig, gs, row_index, item, is_last_row, aspect_mode):
    key = item["key"]
    label = item["label"]
    unit = item["unit"]
    x1 = item["x1"]
    y1 = item["y1"]
    z1 = item["z1"]
    x2 = item["x2"]
    y2 = item["y2"]
    z2 = item["z2"]
    d = item["diff"]

    vmin = np.nanmin([z1, z2])
    vmax = np.nanmax([z1, z2])

    # -----------------------------------------------------
    # 1-й столбец: подпись показателя
    # -----------------------------------------------------
    ax0 = fig.add_subplot(gs[row_index, 0])
    ax0.axis("off")
    ax0.text(
        0.5,
        0.5,
        label,
        ha="center",
        va="center",
        fontsize=METRIC_LABEL_SIZE,
        fontweight="bold",
        rotation=90,
    )
    if COL1_SHIFT_X != 0:
        shift_axis(ax0, dx=COL1_SHIFT_X)

    # -----------------------------------------------------
    # 2-й столбец: секционированная
    # -----------------------------------------------------
    ax1 = fig.add_subplot(gs[row_index, 1])
    ax1.imshow(z1, aspect=aspect_mode, cmap=MATTE_METRIC_CMAP, vmin=vmin, vmax=vmax)
    style_axis(ax1, x1, y1, show_x=is_last_row, show_y=True)
    if COL2_SHIFT_X != 0:
        shift_axis(ax1, dx=COL2_SHIFT_X)

    # -----------------------------------------------------
    # 3-й столбец: двойная
    # -----------------------------------------------------
    ax2 = fig.add_subplot(gs[row_index, 2])
    im2 = ax2.imshow(z2, aspect=aspect_mode, cmap=MATTE_METRIC_CMAP, vmin=vmin, vmax=vmax)
    style_axis(ax2, x2, y2, show_x=is_last_row, show_y=False)

    cb2 = fig.colorbar(im2, ax=ax2, fraction=CB_FRACTION, pad=CB_PAD)
    setup_metric_colorbar(cb2, key, label, unit)

    # СДВИГАТЬ ТОЛЬКО ПОСЛЕ СОЗДАНИЯ COLORBAR
    if COL3_SHIFT_X != 0:
        shift_axis(ax2, dx=COL3_SHIFT_X)
        if MOVE_CB_WITH_COL3:
            shift_axis(cb2.ax, dx=COL3_SHIFT_X)

    # -----------------------------------------------------
    # 4-й столбец: delta
    # -----------------------------------------------------
    ax3 = fig.add_subplot(gs[row_index, 3])
    lim = np.nanmax(np.abs(d))
    if not np.isfinite(lim) or lim == 0:
        lim = 1.0

    im3 = ax3.imshow(d, aspect=aspect_mode, cmap=MATTE_DIFF_CMAP, vmin=-lim, vmax=lim)
    style_axis(ax3, x1, y1, show_x=is_last_row, show_y=False)

    cb3 = fig.colorbar(im3, ax=ax3, fraction=CB_FRACTION, pad=CB_PAD)
    setup_delta_colorbar(cb3)

    # СДВИГАТЬ ТОЛЬКО ПОСЛЕ СОЗДАНИЯ COLORBAR
    if COL4_SHIFT_X != 0:
        shift_axis(ax3, dx=COL4_SHIFT_X)
        if MOVE_CB_WITH_COL4:
            shift_axis(cb3.ax, dx=COL4_SHIFT_X)


# =========================================================
# ОТДЕЛЬНАЯ КАРТИНКА ДЛЯ ОДНОЙ МЕТРИКИ
# =========================================================
def save_individual_figure(item, out_path):
    fig = plt.figure(figsize=(INDIVIDUAL_FIG_WIDTH, INDIVIDUAL_FIG_HEIGHT))
    gs = fig.add_gridspec(
        2,
        4,
        height_ratios=[INDIVIDUAL_HEADER_HEIGHT_RATIO, 1.0],
        width_ratios=INDIVIDUAL_WIDTH_RATIOS,
        hspace=INDIVIDUAL_HSPACE,
        wspace=INDIVIDUAL_WSPACE,
    )

    headers = [
        HEADER_METRIC,
        HEADER_SCHEME_1,
        HEADER_SCHEME_2,
        HEADER_DELTA,
    ]

    header_shifts = [COL1_SHIFT_X, COL2_SHIFT_X, COL3_SHIFT_X, COL4_SHIFT_X]

    for j in range(4):
        ax = fig.add_subplot(gs[0, j])
        ax.axis("off")
        ax.text(
            0.5,
            0.08,
            headers[j],
            ha="center",
            va="center",
            fontsize=HEADER_SIZE,
            fontweight="bold",
        )
        if header_shifts[j] != 0:
            shift_axis(ax, dx=header_shifts[j])

    draw_table_row(
        fig=fig,
        gs=gs,
        row_index=1,
        item=item,
        is_last_row=True,
        aspect_mode=INDIVIDUAL_IMAGE_ASPECT,
    )

    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


# =========================================================
# ОБЩАЯ КАРТИНКА
# =========================================================
def save_summary_figure(prepared, out_path):
    rows = len(prepared)

    fig = plt.figure(figsize=(SUMMARY_FIG_WIDTH, SUMMARY_ROW_HEIGHT * rows))
    gs = fig.add_gridspec(
        rows + 1,
        4,
        height_ratios=[SUMMARY_HEADER_HEIGHT_RATIO] + [1.0] * rows,
        width_ratios=SUMMARY_WIDTH_RATIOS,
        hspace=SUMMARY_HSPACE,
        wspace=SUMMARY_WSPACE,
    )

    headers = [
        HEADER_METRIC,
        HEADER_SCHEME_1,
        HEADER_SCHEME_2,
        HEADER_DELTA,
    ]

    header_shifts = [COL1_SHIFT_X, COL2_SHIFT_X, COL3_SHIFT_X, COL4_SHIFT_X]

    for j in range(4):
        ax = fig.add_subplot(gs[0, j])
        ax.axis("off")
        ax.text(
            0.5,
            0.10,
            headers[j],
            ha="center",
            va="center",
            fontsize=HEADER_SIZE,
            fontweight="bold",
        )
        if header_shifts[j] != 0:
            shift_axis(ax, dx=header_shifts[j])

    for i, item in enumerate(prepared):
        draw_table_row(
            fig=fig,
            gs=gs,
            row_index=i + 1,
            item=item,
            is_last_row=(i == rows - 1),
            aspect_mode=SUMMARY_IMAGE_ASPECT,
        )

    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


# =========================================================
# ОСНОВНОЙ КОД
# =========================================================
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    prepared = prepare_metric_data(ws)

    # Общая картинка
    summary_path = os.path.join(OUTPUT_DIR, "SS_vs_D_table_compare.png")
    save_summary_figure(prepared, summary_path)

    # Отдельные картинки
    prepared_map = {item["key"]: item for item in prepared}

    for metric_key in INDIVIDUAL_METRICS:
        if metric_key not in prepared_map:
            raise ValueError(f"Метрика '{metric_key}' отсутствует в METRICS")

        item = prepared_map[metric_key]
        out_name = safe_filename(f"{metric_key}_compare") + ".png"
        out_path = os.path.join(OUTPUT_DIR, out_name)
        save_individual_figure(item, out_path)

    print("Готово.")
    print(f"Общая фигура: {summary_path}")
    print(f"Отдельные фигуры сохранены в папку: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()