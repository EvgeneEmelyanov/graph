import os
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap, BoundaryNorm
from matplotlib.patches import Patch
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter


# =========================
# НАСТРОЙКИ
# =========================

EXCEL_ITEMS = [
    {"path": r"D:\SS.xlsx", "label": "Секционированная система", "out_name": "sectioned_system"},
    {"path": r"D:\D.xlsx", "label": "Двойная система", "out_name": "double_system"},
]

SHEET_NAME = "SWEEP_2"
OUTPUT_DIR = r"D:\results"

BASE_RANGE = "A2:AU6"

X_AXIS_LABEL = "Мощность ДГУ, кВт"
Y_AXIS_LABEL = "Количество ДГУ"

SCENARIO_AGGREGATION = "mean"

TOP_THRESHOLD_LCOE = 0.99
TOP_THRESHOLD_RELIABILITY = 0.999

GRID_COLOR = "#8a8a8a"

INTERSECTION_COLORS = [
    "#ffffff",  # 0 none
    "#f4a261",  # 1 LCOE
    "#457b9d",  # 2 reliability
    "#222222",  # 3 both
]

INTERSECTION_LABELS = {
    0: "Не входит в топ",
    1: "Только LCOE",
    2: "Только надежность",
    3: "LCOE + надежность",
}


# =========================
# DATA MODEL
# =========================

@dataclass
class Table2D:
    x_labels: List[str]
    y_labels: List[str]
    values: np.ndarray


# =========================
# HELPERS
# =========================

def _to_float(v):
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", ".").strip()
    try:
        return float(s)
    except:
        return np.nan


def build_range(base_range, offset):
    c1, r1, c2, r2 = range_boundaries(base_range)
    return f"{get_column_letter(c1)}{r1+offset}:{get_column_letter(c2)}{r2+offset}"


def read_table(ws, rng):
    min_c, min_r, max_c, max_r = range_boundaries(rng)

    x = [str(ws.cell(min_r, c).value) for c in range(min_c+1, max_c+1)]
    y = []
    vals = []

    for r in range(min_r+1, max_r+1):
        y.append(str(ws.cell(r, min_c).value))
        row = [_to_float(ws.cell(r, c).value) for c in range(min_c+1, max_c+1)]
        vals.append(row)

    return Table2D(x, y, np.array(vals, dtype=float))


def normalize_min(arr):
    vmin, vmax = np.nanmin(arr), np.nanmax(arr)
    if np.isclose(vmin, vmax):
        return np.ones_like(arr)
    return (vmax - arr) / (vmax - vmin)


def aggregate(mats):
    return np.nanmean(np.stack(mats), axis=0)


def top_mask(score, threshold):
    best = np.nanmax(score)
    return score >= best * threshold


def build_code(lcoe_mask, rel_mask):
    code = np.zeros_like(lcoe_mask, dtype=int)
    for i in range(code.shape[0]):
        for j in range(code.shape[1]):
            if lcoe_mask[i, j] and rel_mask[i, j]:
                code[i, j] = 3
            elif lcoe_mask[i, j]:
                code[i, j] = 1
            elif rel_mask[i, j]:
                code[i, j] = 2
    return code


def setup_axis(ax, x, y, title):
    ax.set_title(title)
    ax.set_xticks(range(len(x)))
    ax.set_yticks(range(len(y)))
    ax.set_xticklabels(x, rotation=90, fontsize=8)
    ax.set_yticklabels(y, fontsize=8)
    ax.set_xlabel(X_AXIS_LABEL)
    ax.set_ylabel(Y_AXIS_LABEL)

    ax.set_xticks(np.arange(-0.5, len(x), 1), minor=True)
    ax.set_yticks(np.arange(-0.5, len(y), 1), minor=True)
    ax.grid(which="minor", color=GRID_COLOR, linewidth=0.4)


# =========================
# ОСНОВНАЯ ЛОГИКА
# =========================

def process_file(item):
    wb = load_workbook(item["path"], data_only=True)
    ws = wb[SHEET_NAME]

    # Чтение всех метрик
    LCOE = read_table(ws, build_range(BASE_RANGE, 0))
    ENS = read_table(ws, build_range(BASE_RANGE, 24))
    LOLH = read_table(ws, build_range(BASE_RANGE, 32))
    EVT_N = read_table(ws, build_range(BASE_RANGE, 40))
    EVT_MAX = read_table(ws, build_range(BASE_RANGE, 48))

    # LOLH -> в год
    LOLH.values = LOLH.values / 20.0

    # Нормировка
    LCOE_s = normalize_min(LCOE.values)
    ENS_s = normalize_min(ENS.values)
    LOLH_s = normalize_min(LOLH.values)
    EVT_N_s = normalize_min(EVT_N.values)
    EVT_MAX_s = normalize_min(EVT_MAX.values)

    # Надежность (взвешенная)
    REL = (
        0.25 * ENS_s +
        0.25 * LOLH_s +
        0.25 * EVT_N_s +
        0.25 * EVT_MAX_s
    )

    # Маски top 99%
    mask_lcoe = top_mask(LCOE_s, TOP_THRESHOLD_LCOE)
    mask_rel = top_mask(REL, TOP_THRESHOLD_RELIABILITY)

    code = build_code(mask_lcoe, mask_rel)

    # ===== ГРАФИК =====
    fig, ax = plt.subplots(figsize=(14, 8))
    cmap = ListedColormap(INTERSECTION_COLORS)
    norm = BoundaryNorm(np.arange(-0.5, 4.5, 1), cmap.N)

    ax.imshow(code, cmap=cmap, norm=norm, origin="lower", aspect="auto")

    setup_axis(ax, LCOE.x_labels, LCOE.y_labels, item["label"])

    legend = [
        Patch(color=INTERSECTION_COLORS[i], label=INTERSECTION_LABELS[i])
        for i in range(4)
    ]
    ax.legend(handles=legend, bbox_to_anchor=(1.02, 1), loc="upper left")

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out = os.path.join(OUTPUT_DIR, item["out_name"] + ".png")

    plt.tight_layout()
    plt.savefig(out, dpi=300)
    plt.close()

    print(f"Сохранено: {out}")


# =========================
# MAIN
# =========================

if __name__ == "__main__":
    for item in EXCEL_ITEMS:
        process_file(item)