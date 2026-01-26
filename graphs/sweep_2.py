import numpy as np
from openpyxl import load_workbook

# =============================
# НАСТРОЙКИ
# =============================
file = r"D:\results.xlsx"

sheet = "SWEEP_2"

out_html = r"D:\tri0.html"
out_png  = r"D:\tri0.png"

# Диапазоны:
# X: B2:V2
# Y: A3:A23
# Z: B{Z_START_ROW}:V{Z_END_ROW}
X_ROW = 2
Y_START_ROW, Y_END_ROW = 3, 9
X_START_COL, X_END_COL = 2, 12
Z_START_ROW, Z_END_ROW = 3,9

# =============================
# ФОРМА ОБЛАСТИ ДАННЫХ
# =============================
# "triangle" -> обрезаем по X+Y<=1
# "rect"     -> используем весь прямоугольник таблицы Z
SHAPE_MODE = "rect"  # <-- меняйте тут
TRI_LIMIT = 1.0000001

# =============================
# ТИКИ ОСЕЙ
# =============================
# "auto" -> тики берём по фактическим X/Y (нужно для rect с любыми осями)
# "01"   -> тики 0..1 шагом TICK_STEP (старый режим)
TICK_MODE = "auto"   # <-- меняйте тут
TICK_STEP = 0.10     # используется только при TICK_MODE="01"
MAX_TICKS = 8        # ограничение числа подписей в режиме auto

# --- Matplotlib (PNG) угол (менять здесь при необходимости)
ELEV = 22
AZIM = 35

# --- Plotly (HTML) стартовый "угол камеры"
CAMERA_EYE = dict(x=1.6, y=1.6, z=0.9)

# =============================
# ВСПОМОГАТЕЛЬНОЕ
# =============================
def to_float(v):
    """Excel -> float. Поддерживает числа, строки с запятой, игнорирует пустые/###."""
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s == "###":
            return np.nan
        s = s.replace(" ", "").replace("\u00A0", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return np.nan
    return np.nan


def _cell_to_label(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def get_z_label(ws, z_start_row, z_end_row):
    """
    Правила:
    - Если Z начинается с Y_START_ROW (обычно 3), подпись из A1.
    - Иначе подпись из (row=z_start_row-2, col=X_START_COL-1).
    """
    if z_start_row == Y_START_ROW:
        return _cell_to_label(ws.cell(row=1, column=1).value)

    label_row = z_start_row - 2
    label_col = X_START_COL - 1  # столбец слева от X (обычно A)
    return _cell_to_label(ws.cell(row=label_row, column=label_col).value)


def nice_ticks(vals, max_ticks=8):
    """Разреживает тики (уникальные, отсортированные), чтобы не было каши."""
    u = np.unique(vals[np.isfinite(vals)])
    u = np.sort(u)
    if u.size == 0:
        return np.array([])
    if u.size <= max_ticks:
        return u
    idx = np.linspace(0, u.size - 1, max_ticks).round().astype(int)
    idx = np.unique(idx)
    return u[idx]


def format_ticks(vals):
    """
    Форматирование подписей:
    - большие значения без дроби
    - средние 1 знак
    - маленькие 2 знака
    + запятая
    """
    out = []
    for v in vals:
        av = abs(float(v))
        if av >= 100:
            s = f"{v:.0f}"
        elif av >= 10:
            s = f"{v:.1f}"
        else:
            s = f"{v:.2f}"
        out.append(s.replace(".", ","))
    return out


def get_axis_ticks(x_vals, y_vals):
    """Возвращает (x_ticks, x_ticktext, y_ticks, y_ticktext)"""
    if TICK_MODE == "01":
        t = np.round(np.arange(0.0, 1.0 + 1e-9, TICK_STEP), 2)
        tt = [f"{v:.2f}".replace(".", ",") for v in t]
        return t, tt, t, tt

    xt = nice_ticks(x_vals, MAX_TICKS)
    yt = nice_ticks(y_vals, MAX_TICKS)
    return xt, format_ticks(xt), yt, format_ticks(yt)


def build_mesh_from_grid(X, Y, Z):
    """
    X,Y,Z: (ny,nx)
    Возвращает vx,vy,vz и треугольники I,J,K.
    Работает и для rect, и для triangle (при наличии NaN после маски).
    """
    valid_mask = np.isfinite(Z)
    verts_r, verts_c = np.where(valid_mask)

    vx = X[verts_r, verts_c].astype(float)
    vy = Y[verts_r, verts_c].astype(float)
    vz = Z[verts_r, verts_c].astype(float)

    idx = -np.ones(Z.shape, dtype=int)
    idx[verts_r, verts_c] = np.arange(vx.size)

    I, J, K = [], [], []
    rows, cols = Z.shape

    for r in range(rows - 1):
        for c in range(cols - 1):
            a = idx[r, c]
            b = idx[r, c + 1]
            d = idx[r + 1, c]
            e = idx[r + 1, c + 1]

            present = [t for t in (a, b, e, d) if t != -1]

            if len(present) == 4:
                I += [a, a]
                J += [b, e]
                K += [e, d]
            elif len(present) == 3:
                I.append(present[0])
                J.append(present[1])
                K.append(present[2])

    return vx, vy, vz, I, J, K


# =============================
# ЧТЕНИЕ EXCEL (data_only=True)
# =============================
wb = load_workbook(file, data_only=True)
ws = wb[sheet]

z_axis_title = get_z_label(ws, Z_START_ROW, Z_END_ROW) or "Z"

x = np.array(
    [to_float(ws.cell(row=X_ROW, column=col).value) for col in range(X_START_COL, X_END_COL + 1)],
    dtype=float
)
y = np.array(
    [to_float(ws.cell(row=row, column=1).value) for row in range(Y_START_ROW, Y_END_ROW + 1)],
    dtype=float
)

expected_rows = len(y)
actual_rows = Z_END_ROW - Z_START_ROW + 1
if actual_rows != expected_rows:
    raise RuntimeError(
        f"Диапазон Z по строкам не совпадает с Y: Y={expected_rows} строк, Z={actual_rows} строк.\n"
        f"Исправьте Z_START_ROW/Z_END_ROW или Y_START_ROW/Y_END_ROW."
    )

z = np.full((len(y), len(x)), np.nan, dtype=float)
for i, row in enumerate(range(Z_START_ROW, Z_END_ROW + 1)):
    for j, col in enumerate(range(X_START_COL, X_END_COL + 1)):
        z[i, j] = to_float(ws.cell(row=row, column=col).value)

if np.all(np.isnan(x)) or np.all(np.isnan(y)):
    raise RuntimeError("X или Y не считались (всё NaN). Проверьте лист/диапазоны и сохранённые значения.")

# =============================
# МАСКА ОБЛАСТИ (triangle/rect)
# =============================
Xg, Yg = np.meshgrid(x, y)

mask = np.isnan(z)
if SHAPE_MODE == "triangle":
    mask = mask | ((Xg + Yg) > TRI_LIMIT)

Z = z.copy()
Z[mask] = np.nan

valid = Z[np.isfinite(Z)]
if valid.size == 0:
    raise RuntimeError(
        "В Z нет валидных значений.\n"
        "Чаще всего: в xlsx нет сохранённого кэша формул (data_only=True), или читается не тот блок."
    )

z_min = float(np.min(valid))
z_max = float(np.max(valid))

# тики по X/Y (для rect это обязательно)
x_ticks, x_ticktext, y_ticks, y_ticktext = get_axis_ticks(x, y)

# =====================================================================
# 1) WEB-ВЕРСИЯ (HTML) — Plotly: Mesh3d + точки
# =====================================================================
import plotly.graph_objects as go

colorscale = [
    [0.00, "limegreen"],
    [0.20, "yellowgreen"],
    [0.40, "yellow"],
    [0.60, "orange"],
    [0.80, "red"],
    [1.00, "maroon"],
]

vx, vy, vz, I, J, K = build_mesh_from_grid(Xg, Yg, Z)

mesh = go.Mesh3d(
    x=vx, y=vy, z=vz,
    i=I, j=J, k=K,
    intensity=vz,
    colorscale=colorscale,
    cmin=z_min, cmax=z_max,
    showscale=True,
    colorbar=dict(
        title=dict(text=z_axis_title),
        thickness=18,
        len=0.75
    ),
    flatshading=False,
    hovertemplate="x=%{x}<br>y=%{y}<br>z=%{z:.2f}<extra></extra>",
    showlegend=False
)

points = go.Scatter3d(
    x=vx, y=vy, z=vz,
    mode="markers",
    marker=dict(
        size=4,
        opacity=1.0,
        color=vz,
        colorscale=colorscale,
        cmin=z_min,
        cmax=z_max,
        showscale=False
    ),
    hovertemplate="x=%{x}<br>y=%{y}<br>z=%{z:.2f}<extra></extra>",
    showlegend=False
)

fig_html = go.Figure(data=[mesh, points])

fig_html.update_layout(
    template="plotly_white",
    width=1400,
    height=850,
    margin=dict(l=40, r=40, t=40, b=40),
    scene=dict(
        xaxis=dict(
            title="X",
            tickmode="array",
            tickvals=x_ticks.tolist(),
            ticktext=x_ticktext,
        ),
        yaxis=dict(
            title="Y",
            tickmode="array",
            tickvals=y_ticks.tolist(),
            ticktext=y_ticktext,
        ),
        zaxis=dict(title=z_axis_title),
        camera=dict(eye=CAMERA_EYE),
        aspectmode="manual",
        aspectratio=dict(x=1.2, y=1.2, z=0.7),
    )
)

fig_html.write_html(out_html, include_plotlyjs="cdn", auto_open=True)
print(f"✅ HTML сохранён и открыт: {out_html}")

# =====================================================================
# 2) КАРТИНКА (PNG) — Matplotlib: trisurf + точки
# =====================================================================
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, Normalize
import matplotlib.tri as mtri

colors = ["limegreen", "yellowgreen", "yellow", "orange", "red", "maroon"]
cmap = LinearSegmentedColormap.from_list("custom_scale", colors, N=256)
norm = Normalize(vmin=z_min, vmax=z_max)

triang = mtri.Triangulation(vx, vy, triangles=np.column_stack([I, J, K]))

mm_to_in = 1 / 25.4
fig_width_in = 360 * mm_to_in
fig_height_in = 220 * mm_to_in

fig = plt.figure(figsize=(fig_width_in, fig_height_in), dpi=300)
ax = fig.add_axes([0.08, 0.10, 0.72, 0.80], projection="3d")

ax.plot_trisurf(
    triang, vz,
    cmap=cmap,
    norm=norm,
    linewidth=0.3,
    edgecolor="black",
    antialiased=True
)

ax.scatter(
    vx, vy, vz,
    s=10,
    c=vz,
    cmap=cmap,
    norm=norm,
    depthshade=False
)

ax.set_xlabel("X", labelpad=12, fontsize=12)
ax.set_ylabel("Y", labelpad=12, fontsize=12)
ax.set_zlabel(z_axis_title, labelpad=30, fontsize=12, rotation=180)

ax.tick_params(axis="x", pad=4, labelsize=10)
ax.tick_params(axis="y", pad=4, labelsize=10)
ax.tick_params(axis="z", pad=6, labelsize=10)

ax.set_xticks(x_ticks)
ax.set_yticks(y_ticks)
ax.set_xticklabels(x_ticktext, fontsize=10)
ax.set_yticklabels(y_ticktext, fontsize=10)

ax.view_init(elev=ELEV, azim=AZIM)

fig.patch.set_facecolor("white")
ax.set_facecolor("white")
plt.subplots_adjust(left=0.05, right=0.92, top=0.95, bottom=0.08)

plt.savefig(out_png, dpi=300)
plt.close(fig)
print(f"✅ PNG сохранён: {out_png}")

print(f"Shape mode: {SHAPE_MODE} | Tick mode: {TICK_MODE}")
print(f"Z range: [{z_min:.3g}, {z_max:.3g}]  (valid points: {valid.size})")
