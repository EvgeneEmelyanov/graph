import re
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# =========================
# НАСТРОЙКИ ПОЛЬЗОВАТЕЛЯ
# =========================

EXCEL_PATH = r"D:\comb2.xlsx"
SHEET_NAME: Optional[str] = "SWEEP_2"  # None = первый лист

# --- 1) РЕЖИМ ОЦЕНКИ ---
# "metrics" -> оцениваем по метрикам
# "super"   -> оцениваем по супер-критериям
SCORE_MODE = "super"  # "metrics" | "super"

# --- 2) МЕТРИКИ, которые читаем из Excel ---
METRICS = [
    # economy
    "LCOE, руб/кВт∙ч",

    # reliability
    "ENS,кВт∙ч",
    "LOLE_h",
    "ENS1_mean",
    "ENS2_mean",
    "ENS_evtN",
    "ENS_evtMaxH",

    # operations
    "Расход топлива, тыс.тонн",
    "Моточасы, тыс.мч",
]

# Направления: 'min' (меньше лучше) или 'max' (больше лучше)
DIRECTIONS: Dict[str, str] = {m: "min" for m in METRICS}

# Веса метрик (актуально для SCORE_MODE="metrics" и для супер-критерия если method="wsum")
METRIC_WEIGHTS: Dict[str, float] = {m: 1.0 for m in METRICS}

# Как оценивать каждую группу по каждой метрике (схлопываем 2D-таблицу в 1 число):
#  - "best"   : min/max по таблице
#  - "mean"   : среднее по таблице
#  - "median" : медиана по таблице
AGGREGATOR = "best"  # "best" | "mean" | "median"

# Как объединять (сумма нормированных или ранги):
COMBINE_MODE = "sum"  # "sum" | "rank"

# Названия групп (сценариев) слева-направо (если None — G1..Gk)
GROUP_NAMES: Optional[List[str]] = None

# Скан справа, чтобы найти конец заголовка
MAX_COL_SCAN = 500

# --- 3) СУПЕР-КРИТЕРИИ ---
# Выбираешь, какие метрики входят в какой критерий.
# Важно: метрики, указанные здесь, должны присутствовать в METRICS.
SUPER_GROUPS: Dict[str, List[str]] = {
    "economy": ["LCOE, руб/кВт∙ч"],
    "reliability": ["ENS,кВт∙ч", "LOLE_h", "ENS_evtN", "ENS_evtMaxH"],
    # "operations": ["Расход топлива, тыс.тонн", "Моточасы, тыс.мч"],
    # "operations": ["Моточасы, тыс.мч"],
}

# Агрегирование внутри супер-критерия (по НОРМИРОВАННЫМ значениям 0..1):
#   "wsum"  : взвешенная сумма (веса берём из METRIC_WEIGHTS)
#   "mean"  : среднее
#   "max"   : worst-case (полезно для надежности)
#   "pnorm" : p-норма (p задаётся в SUPER_P)
SUPER_AGG_METHOD: Dict[str, str] = {
    "economy": "wsum",
    "reliability": "pnorm",
    "operations": "mean",
}

SUPER_P: Dict[str, float] = {"reliability": 4.0}

# Веса супер-критериев (между собой)
SUPER_WEIGHTS: Dict[str, float] = {"economy": 0.4, "reliability": 0.4, "operations": 0.2}

# (опционально) Пороговые ограничения на агрегированные метрики группы (до нормировки):
# Пример: CONSTRAINTS = {"LOLE_h": ("<=", 5.0), "ENS,кВт∙ч": ("<=", 1000.0)}
# Если нарушено -> группа получает +inf/NaN (хуже всех).
CONSTRAINTS: Dict[str, Tuple[str, float]] = {}

# =========================
# Data model
# =========================

@dataclass
class Table2D:
    x_labels: np.ndarray
    y_labels: np.ndarray
    values: np.ndarray  # [ny, nx]


# =========================
# Helpers
# =========================

def _to_float_ru_or_nan(v: object) -> float:
    if v is None:
        return np.nan
    if isinstance(v, (int, float, np.integer, np.floating)):
        return float(v)
    s = str(v).replace("\xa0", "").replace(" ", "").strip()
    if s == "":
        return np.nan
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return np.nan

def _is_text(v: object) -> bool:
    if v is None:
        return False
    return bool(re.search(r"[A-Za-zА-Яа-я]", str(v)))

def _find_metric_row(ws: Worksheet, metric: str) -> int:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is None:
            continue
        if str(v).strip() == metric:
            return r
    raise ValueError(f"Metric '{metric}' not found in column A.")

def _detect_table_height(ws: Worksheet, header_row: int) -> int:
    r = header_row + 1
    last = header_row
    while r <= ws.max_row:
        a = ws.cell(row=r, column=1).value
        if _is_text(a):
            break
        any_non_empty = False
        for c in range(1, 21):
            if ws.cell(row=r, column=c).value not in (None, ""):
                any_non_empty = True
                break
        if not any_non_empty:
            break
        last = r
        r += 1
    return last

def _find_blocks_in_header(ws: Worksheet, header_row: int) -> List[Tuple[int, int]]:
    last_non_empty = 1
    for c in range(1, min(ws.max_column, MAX_COL_SCAN) + 1):
        v = ws.cell(row=header_row, column=c).value
        if v not in (None, ""):
            last_non_empty = c

    blocks: List[Tuple[int, int]] = []
    c = 2
    while c <= last_non_empty:
        v = _to_float_ru_or_nan(ws.cell(row=header_row, column=c).value)
        if np.isfinite(v):
            start = c
            c += 1
            while c <= last_non_empty:
                vv = _to_float_ru_or_nan(ws.cell(row=header_row, column=c).value)
                if not np.isfinite(vv):
                    break
                c += 1
            end = c
            blocks.append((start, end))
        else:
            c += 1
    if not blocks:
        raise ValueError("Could not detect any numeric header blocks (groups) in the header row.")
    return blocks

def load_metric_tables_from_excel(xlsx_path: str, sheet_name: Optional[str], metric: str) -> List[Table2D]:
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
        mr = _find_metric_row(ws, metric)
        header_row = mr + 1
        if header_row > ws.max_row:
            raise ValueError(f"Metric '{metric}' has no header row below it.")

        max_row = _detect_table_height(ws, header_row)
        blocks = _find_blocks_in_header(ws, header_row)

        tables: List[Table2D] = []
        for (c0, c1) in blocks:
            x = np.array([_to_float_ru_or_nan(ws.cell(row=header_row, column=c).value) for c in range(c0, c1)], dtype=float)

            y_list: List[float] = []
            vals: List[List[float]] = []
            for r in range(header_row + 1, max_row + 1):
                yv = _to_float_ru_or_nan(ws.cell(row=r, column=1).value)
                row_vals = [_to_float_ru_or_nan(ws.cell(row=r, column=c).value) for c in range(c0, c1)]
                if (not np.isfinite(yv)) and all(not np.isfinite(v) for v in row_vals):
                    break
                y_list.append(yv)
                vals.append(row_vals)

            arr = np.array(vals, dtype=float)
            tables.append(Table2D(x_labels=x, y_labels=np.array(y_list, dtype=float), values=arr))

        return tables
    finally:
        wb.close()


# =========================
# Scoring
# =========================

def _finite(vals: np.ndarray) -> np.ndarray:
    return vals[np.isfinite(vals)]

def _aggregate_value(table: Table2D, direction: str, aggregator: str) -> float:
    v = _finite(table.values)
    if v.size == 0:
        return np.nan
    if aggregator == "best":
        return float(np.min(v) if direction == "min" else np.max(v))
    if aggregator == "mean":
        return float(np.mean(v))
    if aggregator == "median":
        return float(np.median(v))
    raise ValueError(f"Unknown AGGREGATOR={aggregator}")

def _normalize(values: np.ndarray, direction: str) -> np.ndarray:
    v = values.astype(float)
    if np.all(~np.isfinite(v)):
        return np.full_like(v, np.nan, dtype=float)

    finite = v[np.isfinite(v)]
    vmin = float(np.min(finite))
    vmax = float(np.max(finite))
    if abs(vmax - vmin) < 1e-15:
        out = np.zeros_like(v, dtype=float)
        out[~np.isfinite(v)] = np.nan
        return out

    if direction == "min":
        out = (v - vmin) / (vmax - vmin)
    elif direction == "max":
        out = (vmax - v) / (vmax - vmin)
    else:
        raise ValueError(f"Direction must be 'min' or 'max', got: {direction}")

    out[~np.isfinite(v)] = np.nan
    return out

def _fmt(v: float) -> str:
    if not np.isfinite(v):
        return "NaN"
    av = abs(v)
    if av != 0 and (av < 1e-3 or av >= 1e4):
        return f"{v:.6e}"
    return f"{v:.6f}".rstrip("0").rstrip(".")

def _check_constraints(raw_agg: Dict[str, float]) -> bool:
    for m, (op, thr) in CONSTRAINTS.items():
        v = raw_agg.get(m, np.nan)
        if not np.isfinite(v):
            return False
        if op == "<=" and not (v <= thr):
            return False
        if op == "<" and not (v < thr):
            return False
        if op == ">=" and not (v >= thr):
            return False
        if op == ">" and not (v > thr):
            return False
        if op == "==" and not (v == thr):
            return False
        if op == "!=" and not (v != thr):
            return False
        if op not in ("<=", "<", ">=", ">", "==", "!="):
            raise ValueError(f"Unknown constraint operator: {op} (metric {m})")
    return True

def _super_aggregate(group_name: str, norm_vals_by_metric: Dict[str, np.ndarray]) -> np.ndarray:
    mets = SUPER_GROUPS[group_name]
    method = SUPER_AGG_METHOD.get(group_name, "mean").lower()

    vals = np.stack([norm_vals_by_metric[m] for m in mets], axis=0)  # [nm, k]

    if method == "mean":
        return np.nanmean(vals, axis=0)

    if method == "max":
        return np.nanmax(vals, axis=0)

    if method == "pnorm":
        p = float(SUPER_P.get(group_name, 2.0))
        return (np.nanmean(np.power(vals, p), axis=0)) ** (1.0 / p)

    if method == "wsum":
        w = np.array([float(METRIC_WEIGHTS.get(m, 1.0)) for m in mets], dtype=float)
        if np.all(w == 0):
            w = np.ones_like(w)
        w = w / np.sum(w)
        return np.nansum(vals * w[:, None], axis=0)

    raise ValueError(f"Unknown SUPER_AGG_METHOD for {group_name}: {method}")


def compare_groups_from_excel(
    xlsx_path: str,
    sheet_name: Optional[str],
    metrics: List[str],
    directions: Dict[str, str],
    group_names: Optional[List[str]] = None,
) -> None:
    # 1) загрузить таблицы по каждой метрике
    metric_tables: Dict[str, List[Table2D]] = {}
    k_expected: Optional[int] = None

    for m in metrics:
        tabs = load_metric_tables_from_excel(xlsx_path, sheet_name, m)
        if k_expected is None:
            k_expected = len(tabs)
        elif len(tabs) != k_expected:
            raise ValueError(f"Metric '{m}' has {len(tabs)} groups, expected {k_expected}.")
        metric_tables[m] = tabs

    assert k_expected is not None
    k = k_expected

    if group_names is None or len(group_names) != k:
        group_names = [f"G{i+1}" for i in range(k)]

    # 2) агрегируем каждую 2D-таблицу метрики в 1 число на группу
    raw_agg: Dict[str, np.ndarray] = {}
    for m in metrics:
        dirn = directions.get(m, "min")
        raw_agg[m] = np.array([_aggregate_value(t, dirn, AGGREGATOR) for t in metric_tables[m]], dtype=float)

    # 3) constraints по агрегированным значениям (до нормировки)
    ok_mask = np.ones(k, dtype=bool)
    if CONSTRAINTS:
        for i in range(k):
            row = {m: float(raw_agg[m][i]) for m in metrics}
            ok_mask[i] = _check_constraints(row)

    # 4) нормировка по каждой метрике МЕЖДУ группами
    norm_by_metric: Dict[str, np.ndarray] = {}
    for m in metrics:
        dirn = directions.get(m, "min")
        norm_by_metric[m] = _normalize(raw_agg[m], dirn)

    # применяем constraints: запрещённые группы делаем NaN (хуже всех)
    for m in metrics:
        v = norm_by_metric[m].copy()
        v[~ok_mask] = np.nan
        norm_by_metric[m] = v

    # 5) собираем итог
    if SCORE_MODE == "metrics":
        if COMBINE_MODE == "sum":
            total = np.zeros(k, dtype=float)
            any_used = np.zeros(k, dtype=bool)
            for m in metrics:
                w = float(METRIC_WEIGHTS.get(m, 1.0))
                nm = norm_by_metric[m]
                mask = np.isfinite(nm)
                total[mask] += w * nm[mask]
                any_used[mask] = True
            total[~any_used] = np.nan

        elif COMBINE_MODE == "rank":
            # ранги по raw_agg (до нормировки)
            rank_sum = np.zeros(k, dtype=float)
            for m in metrics:
                w = float(METRIC_WEIGHTS.get(m, 1.0))
                dirn = directions.get(m, "min")
                v = raw_agg[m].copy()
                # запретные -> худшие
                v[~ok_mask] = np.nan
                nan_mask = ~np.isfinite(v)
                if np.all(nan_mask):
                    continue
                if dirn == "min":
                    sort_idx = np.argsort(np.where(nan_mask, np.inf, v))
                elif dirn == "max":
                    sort_idx = np.argsort(np.where(nan_mask, -np.inf, -v))
                else:
                    raise ValueError(f"Direction must be 'min' or 'max', got: {dirn}")
                ranks = np.empty(k, dtype=float)
                ranks[sort_idx] = np.arange(1, k + 1, dtype=float)
                rank_sum += w * ranks
            total = rank_sum
        else:
            raise ValueError("COMBINE_MODE must be 'sum' or 'rank'")

        detail_header = "=== Details (per metric, aggregated) ==="
        detail_rows = [(m, raw_agg[m]) for m in metrics]

    elif SCORE_MODE == "super":
        # считаем супер-критерии из нормированных метрик
        super_vals: Dict[str, np.ndarray] = {}
        for gname, mets in SUPER_GROUPS.items():
            if not mets:
                continue
            for m in mets:
                if m not in metrics:
                    raise ValueError(f"Metric '{m}' used in SUPER_GROUPS['{gname}'] but not in METRICS.")
            super_vals[gname] = _super_aggregate(gname, norm_by_metric)

        # итог по супер-критериям
        if COMBINE_MODE == "sum":
            total = np.zeros(k, dtype=float)
            any_used = np.zeros(k, dtype=bool)
            for gname, arr in super_vals.items():
                w = float(SUPER_WEIGHTS.get(gname, 1.0))
                mask = np.isfinite(arr)
                total[mask] += w * arr[mask]
                any_used[mask] = True
            total[~any_used] = np.nan

        elif COMBINE_MODE == "rank":
            # ранги по каждому супер-критерию (уже 0..1 где меньше лучше)
            rank_sum = np.zeros(k, dtype=float)
            for gname, arr in super_vals.items():
                w = float(SUPER_WEIGHTS.get(gname, 1.0))
                v = arr.copy()
                nan_mask = ~np.isfinite(v)
                if np.all(nan_mask):
                    continue
                sort_idx = np.argsort(np.where(nan_mask, np.inf, v))  # меньше лучше
                ranks = np.empty(k, dtype=float)
                ranks[sort_idx] = np.arange(1, k + 1, dtype=float)
                rank_sum += w * ranks
            total = rank_sum
        else:
            raise ValueError("COMBINE_MODE must be 'sum' or 'rank'")

        detail_header = "=== Details (super-criteria, from normalized metrics) ==="
        detail_rows = [(f"SC_{gname}", super_vals[gname]) for gname in super_vals.keys()]

    else:
        raise ValueError("SCORE_MODE must be 'metrics' or 'super'")

    # 6) печать итогов
    order = np.argsort(np.where(np.isfinite(total), total, np.inf))

    print(f"\n=== RESULT: SCORE_MODE={SCORE_MODE}, COMBINE_MODE={COMBINE_MODE}, AGGREGATOR={AGGREGATOR} ===")
    for idx in order:
        g = group_names[idx]
        sc = total[idx]
        flag = "" if ok_mask[idx] else " (CONSTRAINTS FAIL)"
        print(f"{g}: score={'NaN' if not np.isfinite(sc) else f'{sc:.6f}'}{flag}")

    # 7) детализация
    print(f"\n{detail_header}")
    for name, arr in detail_rows:
        print(f"\n[{name}]")
        for i in range(k):
            print(f"  {group_names[i]}: {_fmt(float(arr[i]))}")

    # 8) полезно: распечатать сырые агрегированные метрики (всегда)
    print("\n=== Raw aggregated metrics (per group) ===")
    for m in metrics:
        print(f"\n[{m}] dir={directions.get(m,'min')}, aggregator={AGGREGATOR}")
        for i in range(k):
            v = float(raw_agg[m][i])
            flag = "" if ok_mask[i] else " (CONSTRAINTS FAIL)"
            print(f"  {group_names[i]}: {_fmt(v)}{flag}")


if __name__ == "__main__":
    compare_groups_from_excel(
        xlsx_path=EXCEL_PATH,
        sheet_name=SHEET_NAME,
        metrics=METRICS,
        directions=DIRECTIONS,
        group_names=GROUP_NAMES,
    )