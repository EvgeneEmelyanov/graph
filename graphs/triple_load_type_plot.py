import numpy as np
from openpyxl import load_workbook

# НАСТРОЙКИ: ФАЙЛЫ
files = [
    r"D:\10_results\02_Scopus\4.1.xlsx",
    r"D:\10_results\02_Scopus\5.1.xlsx",
    r"D:\10_results\02_Scopus\6.1.xlsx",
]
sheet = "SWEEP_2"

out_html = r"D:\tri.html"
out_png  = r"D:\tri.png"

# =============================
# ВЫБОР НАБОРА ДАННЫХ (1/2/3)
# =============================
DATASET = 2  # <-- меняйте: 1, 2 или 3

X_ROW = 2
TRI_LIMIT = 1.0000001

if DATASET == 1:
    Y_START_ROW, Y_END_ROW = 3, 9
    X_START_COL, X_END_COL = 2, 12
    Z_BLOCKS = [(3, 9), (14, 20), (25, 31), (36, 42)]
    SHAPE_MODE = "rect"
elif DATASET == 2:
    Y_START_ROW, Y_END_ROW = 3, 12
    X_START_COL, X_END_COL = 2, 12
    Z_BLOCKS = [(3, 12), (17, 26), (31, 40), (45, 54)]
    SHAPE_MODE = "rect"
elif DATASET == 3:
    Y_START_ROW, Y_END_ROW = 3, 13
    X_START_COL, X_END_COL = 2, 12
    Z_BLOCKS = [(3, 13), (18, 28), (33, 43), (48, 58)]
    SHAPE_MODE = "triangle"
else:
    raise ValueError("DATASET должен быть 1, 2 или 3")

# ПОДПИСИ СТОЛБЦОВ (ТИП НАГРУЗКИ)
COLUMN_TITLES = [
    "Промышленная нагрузка",
    "Коммунально-бытовая нагрузка",
    "Сельскохозяйственная нагрузка",
]

# ТИКИ ОСЕЙ
TICK_MODE = "auto"   # "auto" или "01"
TICK_STEP = 0.10
MAX_TICKS = 8

# =============================
# МАСШТАБ ЦВЕТА
# =============================
COLOR_MODE = "each"              # "row" или "each"
EACH_MPL_COLORBAR = "right_one"  # "right_one" или "per_plot"

# =============================
# УГЛЫ ДЛЯ КАЖДОГО ГРАФИКА ОТДЕЛЬНО
# =============================
N_ROWS = len(Z_BLOCKS)
N_COLS = len(files)

ELEV_AZIM = [[(22, 35) for _ in range(N_COLS)] for __ in range(N_ROWS)]
CAMERA_EYE = [[dict(x=1.6, y=1.6, z=0.9) for _ in range(N_COLS)] for __ in range(N_ROWS)]

# =============================
# ШРИФТЫ
# =============================
HTML_COLTITLE_FS = 20
HTML_AXIS_TITLE_FS = 10
HTML_TICK_FS = 11
HTML_COLORBAR_TITLE_FS = 16
HTML_COLORBAR_TICK_FS = 12

PNG_COLTITLE_FS = 14
PNG_AXIS_LABEL_FS = 10
PNG_TICK_FS = 8
PNG_ZTICK_FS = 8
PNG_CBAR_LABEL_FS = 10
PNG_CBAR_TICK_FS = 8

# =============================
# ВСПОМОГАТЕЛЬНОЕ
# =============================
def to_float(v):
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s == "###":
            return np.nan
        s = s.replace(" ", "").replace("\u00A0", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return np.nan
    return np.nan


def _cell_to_label(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()


def nice_ticks(vals, max_ticks=8):
    u = np.unique(vals[np.isfinite(vals)])
    u = np.sort(u)
    if u.size == 0:
        return np.array([])
    if u.size <= max_ticks:
        return u
    idx = np.linspace(0, u.size - 1, max_ticks).round().astype(int)
    idx = np.unique(idx)
    return u[idx]


def format_ticks(vals):
    out = []
    for v in vals:
        av = abs(float(v))
        if av >= 100:
            s = f"{v:.0f}"
        elif av >= 10:
            s = f"{v:.1f}"
        else:
            s = f"{v:.2f}"
        out.append(s.replace(".", ","))
    return out


def get_axis_ticks(x_vals, y_vals):
    if TICK_MODE == "01":
        t = np.round(np.arange(0.0, 1.0 + 1e-9, TICK_STEP), 2)
        tt = [f"{v:.2f}".replace(".", ",") for v in t]
        return t, tt, t, tt

    xt = nice_ticks(x_vals, MAX_TICKS)
    yt = nice_ticks(y_vals, MAX_TICKS)
    return xt, format_ticks(xt), yt, format_ticks(yt)


def build_mesh_from_block(x, y, Z):
    Xg, Yg = np.meshgrid(x, y)

    valid_mask = np.isfinite(Z)
    rr, cc = np.where(valid_mask)

    vx = Xg[rr, cc].astype(float)
    vy = Yg[rr, cc].astype(float)
    vz = Z[rr, cc].astype(float)

    idx = -np.ones(Z.shape, dtype=int)
    idx[rr, cc] = np.arange(vx.size)

    I, J, K = [], [], []
    rows, cols = Z.shape

    for r in range(rows - 1):
        for c in range(cols - 1):
            a = idx[r, c]
            b = idx[r, c + 1]
            d = idx[r + 1, c]
            e = idx[r + 1, c + 1]

            present = [t for t in (a, b, e, d) if t != -1]

            if len(present) == 4:
                I += [a, a]
                J += [b, e]
                K += [e, d]
            elif len(present) == 3:
                I.append(present[0])
                J.append(present[1])
                K.append(present[2])

    return vx, vy, vz, I, J, K


def read_axis_labels_from_first_file():
    """
    X берём из A2, Y из A3 (одно значение для всех графиков).
    Берём из первого файла в списке.
    """
    wb = load_workbook(files[0], data_only=True)
    ws = wb[sheet]
    xlab = _cell_to_label(ws.cell(row=1, column=2).value)
    ylab = _cell_to_label(ws.cell(row=1, column=3).value)
    return (xlab or "X"), (ylab or "Y")


def get_z_label(ws, z_start_row):
    """
    Z для каждой строки свой:
    - если блок начинается с Y_START_ROW -> Z из A1
    - иначе Z из (row=z_start_row-2, col=X_START_COL-1)
    """
    if z_start_row == Y_START_ROW:
        return _cell_to_label(ws.cell(row=1, column=1).value)
    return _cell_to_label(ws.cell(row=z_start_row - 2, column=X_START_COL - 1).value)


def read_block(file_path, z_start_row, z_end_row):
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet]

    zlab = get_z_label(ws, z_start_row)

    x = np.array(
        [to_float(ws.cell(row=X_ROW, column=col).value) for col in range(X_START_COL, X_END_COL + 1)],
        dtype=float
    )
    y = np.array(
        [to_float(ws.cell(row=row, column=1).value) for row in range(Y_START_ROW, Y_END_ROW + 1)],
        dtype=float
    )

    expected_rows = len(y)
    actual_rows = z_end_row - z_start_row + 1
    if actual_rows != expected_rows:
        raise RuntimeError(
            f"[{file_path}] Диапазон Z по строкам не совпадает с Y: Y={expected_rows} строк, Z={actual_rows} строк.\n"
            f"Исправьте Z_START_ROW/Z_END_ROW или Y_START_ROW/Y_END_ROW."
        )

    z = np.full((len(y), len(x)), np.nan, dtype=float)
    for i, row in enumerate(range(z_start_row, z_end_row + 1)):
        for j, col in enumerate(range(X_START_COL, X_END_COL + 1)):
            z[i, j] = to_float(ws.cell(row=row, column=col).value)

    if np.all(np.isnan(x)) or np.all(np.isnan(y)):
        raise RuntimeError(f"[{file_path}] X или Y не считались (всё NaN). Проверьте лист/диапазоны.")

    Xg, Yg = np.meshgrid(x, y)

    mask = np.isnan(z)
    if SHAPE_MODE == "triangle":
        mask = mask | ((Xg + Yg) > TRI_LIMIT)

    Z = z.copy()
    Z[mask] = np.nan

    return x, y, Z, zlab


# =============================
# ЧТЕНИЕ ОСЕЙ X/Y (ГЛОБАЛЬНО)
# =============================
X_AXIS_TITLE, Y_AXIS_TITLE = read_axis_labels_from_first_file()

# =============================
# ЧТЕНИЕ ВСЕХ БЛОКОВ + ДИАПАЗОНЫ ДЛЯ ЦВЕТА
# =============================
blocks = [[None for _ in range(N_COLS)] for _ in range(N_ROWS)]
z_labels = [""] * N_ROWS

row_ranges = [(None, None) for _ in range(N_ROWS)]
cell_ranges = [[(None, None) for _ in range(N_COLS)] for _ in range(N_ROWS)]

x_ref = None
y_ref = None

for row_i, (z0, z1) in enumerate(Z_BLOCKS):
    row_valid = []
    for col_j, fp in enumerate(files):
        x, y, Z, zlab = read_block(fp, z0, z1)

        if x_ref is None:
            x_ref, y_ref = x, y

        if (not z_labels[row_i]) and zlab:
            z_labels[row_i] = zlab

        valid = Z[np.isfinite(Z)]
        if valid.size == 0:
            raise RuntimeError(
                f"[{fp}] Блок Z {z0}:{z1} не содержит валидных значений.\n"
                f"Чаще всего: в xlsx нет сохранённого кэша формул (data_only=True)."
            )

        cell_ranges[row_i][col_j] = (float(np.min(valid)), float(np.max(valid)))
        row_valid.append(valid)
        blocks[row_i][col_j] = (x, y, Z)

    row_all = np.concatenate(row_valid)
    row_ranges[row_i] = (float(np.min(row_all)), float(np.max(row_all)))

x_ticks, x_ticktext, y_ticks, y_ticktext = get_axis_ticks(x_ref, y_ref)

colorscale = [
    [0.00, "limegreen"],
    [0.20, "yellowgreen"],
    [0.40, "yellow"],
    [0.60, "orange"],
    [0.80, "red"],
    [1.00, "maroon"],
]

def get_range(row_i, col_j):
    if COLOR_MODE == "each":
        return cell_ranges[row_i][col_j]
    return row_ranges[row_i]


# =====================================================================
# 1) HTML (Plotly)
# =====================================================================
import plotly.graph_objects as go
from plotly.subplots import make_subplots

fig_html = make_subplots(
    rows=N_ROWS, cols=N_COLS,
    specs=[[{"type": "scene"}] * N_COLS] * N_ROWS,
    horizontal_spacing=0.02,
    vertical_spacing=0.03
)

XL, XR = 0.04, 0.90
YB, YT = 0.06, 0.96
COL_GAP = 0.03
ROW_GAP = 0.04

CB_GAP = 0.012
CB_XSHIFT = -0.025
CB_LEN_FRACTION = 0.80

col_w = (XR - XL - (N_COLS - 1) * COL_GAP) / float(N_COLS)
row_h = (YT - YB - (N_ROWS - 1) * ROW_GAP) / float(N_ROWS)

x_domains = []
for j in range(N_COLS):
    x0 = XL + j * (col_w + COL_GAP)
    x_domains.append([x0, x0 + col_w])

y_domains = []
for i in range(N_ROWS):
    top = YT - i * (row_h + ROW_GAP)
    y_domains.append([top - row_h, top])

fig_html.update_layout(
    template="plotly_white",
    width=1900,
    height=1450,
    margin=dict(l=40, r=170, t=90, b=90),
)

# заголовки столбцов
column_annotations = []
for j, title in enumerate(COLUMN_TITLES[:N_COLS]):
    x_center = 0.5 * (x_domains[j][0] + x_domains[j][1])
    y_top = y_domains[0][1] + 0.030
    column_annotations.append(
        dict(
            x=x_center, y=y_top,
            xref="paper", yref="paper",
            text=f"<b>{title}</b>",
            showarrow=False,
            xanchor="center",
            yanchor="bottom",
            font=dict(size=HTML_COLTITLE_FS),
        )
    )
fig_html.update_layout(annotations=column_annotations)

for row_i in range(N_ROWS):
    z_axis_title = z_labels[row_i] if z_labels[row_i] else "Z"

    row_dom_h = (y_domains[row_i][1] - y_domains[row_i][0])
    cb_len = row_dom_h * CB_LEN_FRACTION

    for col_j in range(N_COLS):
        z_min, z_max = get_range(row_i, col_j)
        x, y, Z = blocks[row_i][col_j]
        vx, vy, vz, I, J, K = build_mesh_from_block(x, y, Z)

        showscale = (col_j == (N_COLS - 1)) if (COLOR_MODE == "row") else True
        CB_THICKNESS = 15

        mesh = go.Mesh3d(
            x=vx, y=vy, z=vz,
            i=I, j=J, k=K,
            intensity=vz,
            colorscale=colorscale,
            cmin=z_min, cmax=z_max,
            showscale=False,
            flatshading=False,
            hovertemplate="x=%{x}<br>y=%{y}<br>z=%{z:.2f}<extra></extra>",
            showlegend=False,
        )


        points = go.Scatter3d(
            x=vx, y=vy, z=vz,
            mode="markers",
            marker=dict(
                size=3,
                opacity=1.0,
                color=vz,
                colorscale=colorscale,
                cmin=z_min,
                cmax=z_max,
                showscale=False
            ),
            hovertemplate="x=%{x}<br>y=%{y}<br>z=%{z:.2f}<extra></extra>",
            showlegend=False
        )

        fig_html.add_trace(mesh, row=row_i + 1, col=col_j + 1)
        fig_html.add_trace(points, row=row_i + 1, col=col_j + 1)

        scene_idx = row_i * N_COLS + col_j + 1
        scene_name = "scene" if scene_idx == 1 else f"scene{scene_idx}"

        cam_eye = CAMERA_EYE[row_i][col_j]

        fig_html.update_layout(**{
            scene_name: dict(
                domain=dict(x=x_domains[col_j], y=y_domains[row_i]),
                xaxis=dict(
                    title=dict(text=X_AXIS_TITLE, font=dict(size=HTML_AXIS_TITLE_FS)),
                    tickfont=dict(size=HTML_TICK_FS),
                    tickmode="array",
                    tickvals=x_ticks.tolist(),
                    ticktext=x_ticktext,
                ),
                yaxis=dict(
                    title=dict(text=Y_AXIS_TITLE, font=dict(size=HTML_AXIS_TITLE_FS)),
                    tickfont=dict(size=HTML_TICK_FS),
                    tickmode="array",
                    tickvals=y_ticks.tolist(),
                    ticktext=y_ticktext,
                ),
                zaxis=dict(
                    title=dict(text=z_axis_title, font=dict(size=HTML_AXIS_TITLE_FS)),
                    tickfont=dict(size=HTML_TICK_FS),
                ),
                camera=dict(eye=cam_eye),
                aspectmode="manual",
                aspectratio=dict(x=1.2, y=1.2, z=0.7),
            )
        })

fig_html.write_html(out_html, include_plotlyjs="cdn", auto_open=True)
print(f"✅ HTML сохранён и открыт: {out_html}")

# =====================================================================
# 2) PNG (Matplotlib)
# =====================================================================
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, Normalize
import matplotlib.tri as mtri

colors = ["limegreen", "yellowgreen", "yellow", "orange", "red", "maroon"]
cmap = LinearSegmentedColormap.from_list("custom_scale", colors, N=256)

mm_to_in = 1 / 25.4
fig_width_in = 520 * mm_to_in
fig_height_in = 420 * mm_to_in

fig = plt.figure(figsize=(fig_width_in, fig_height_in), dpi=300)

W = 0.285
H = 0.205
X0 = [0.045, 0.345, 0.645]
Y0 = [0.765, 0.535, 0.305, 0.075]
COLTITLE_Y = min(0.985, (Y0[0] + H + 0.020))

for j, title in enumerate(COLUMN_TITLES[:N_COLS]):
    x_center = X0[j] + 0.5 * W
    fig.text(
        x_center, COLTITLE_Y,
        title,
        ha="center", va="bottom",
        fontsize=PNG_COLTITLE_FS,
        fontweight="bold",
    )

row_norms = [Normalize(vmin=row_ranges[i][0], vmax=row_ranges[i][1]) for i in range(N_ROWS)]
cell_norms = [[Normalize(vmin=cell_ranges[i][j][0], vmax=cell_ranges[i][j][1]) for j in range(N_COLS)] for i in range(N_ROWS)]

for row_i in range(N_ROWS):
    z_axis_title = z_labels[row_i] if z_labels[row_i] else "Z"

    for col_j in range(N_COLS):
        norm = cell_norms[row_i][col_j] if (COLOR_MODE == "each") else row_norms[row_i]

        ax = fig.add_axes([X0[col_j], Y0[row_i], W, H], projection="3d")

        x, y, Z = blocks[row_i][col_j]
        vx, vy, vz, I, J, K = build_mesh_from_block(x, y, Z)

        triang = mtri.Triangulation(vx, vy, triangles=np.column_stack([I, J, K]))

        ax.plot_trisurf(
            triang, vz,
            cmap=cmap,
            norm=norm,
            linewidth=0.25,
            edgecolor="black",
            antialiased=True
        )

        ax.scatter(
            vx, vy, vz,
            s=6,
            c=vz,
            cmap=cmap,
            norm=norm,
            depthshade=False
        )

        elev, azim = ELEV_AZIM[row_i][col_j]
        ax.view_init(elev=elev, azim=azim)

        ax.set_xlabel(X_AXIS_TITLE, labelpad=6, fontsize=PNG_AXIS_LABEL_FS)
        ax.set_ylabel(Y_AXIS_TITLE, labelpad=6, fontsize=PNG_AXIS_LABEL_FS)
        ax.set_zlabel(z_axis_title, labelpad=8, fontsize=PNG_AXIS_LABEL_FS)

        ax.set_xticks(x_ticks)
        ax.set_yticks(y_ticks)
        ax.set_xticklabels(x_ticktext, fontsize=PNG_TICK_FS)
        ax.set_yticklabels(y_ticktext, fontsize=PNG_TICK_FS)
        ax.tick_params(axis="z", labelsize=PNG_ZTICK_FS)

        if COLOR_MODE == "each" and EACH_MPL_COLORBAR == "per_plot":
            pos = ax.get_position()
            cax = fig.add_axes([pos.x1 + 0.010, pos.y0 + 0.04, 0.010, pos.height * 0.62])
            mappable = plt.cm.ScalarMappable(norm=norm, cmap=cmap)
            cb = fig.colorbar(mappable, cax=cax)
            cb.ax.tick_params(labelsize=PNG_CBAR_TICK_FS)

if COLOR_MODE == "row":
    cbar_x = 0.93
    cbar_w = 0.016
    cbar_h = 0.18
    cbar_ys = [0.79, 0.56, 0.33, 0.10]
    for row_i in range(N_ROWS):
        z_axis_title = z_labels[row_i] if z_labels[row_i] else "Z"
        cax = fig.add_axes([cbar_x, cbar_ys[row_i], cbar_w, cbar_h])
        mappable = plt.cm.ScalarMappable(norm=row_norms[row_i], cmap=cmap)
        cb = fig.colorbar(mappable, cax=cax)
        cb.set_label(z_axis_title, fontsize=PNG_CBAR_LABEL_FS)
        cb.ax.tick_params(labelsize=PNG_CBAR_TICK_FS)

elif COLOR_MODE == "each" and EACH_MPL_COLORBAR == "right_one":
    all_valid = []
    for i in range(N_ROWS):
        for j in range(N_COLS):
            _, _, Z = blocks[i][j]
            v = Z[np.isfinite(Z)]
            if v.size:
                all_valid.append(v)
    all_valid = np.concatenate(all_valid)
    gmin, gmax = float(np.min(all_valid)), float(np.max(all_valid))
    gnorm = Normalize(vmin=gmin, vmax=gmax)

    cax = fig.add_axes([0.93, 0.10, 0.016, 0.84])
    mappable = plt.cm.ScalarMappable(norm=gnorm, cmap=cmap)
    cb = fig.colorbar(mappable, cax=cax)
    cb.set_label("Z", fontsize=PNG_CBAR_LABEL_FS)
    cb.ax.tick_params(labelsize=PNG_CBAR_TICK_FS)

fig.patch.set_facecolor("white")
plt.savefig(out_png, dpi=300)
plt.close(fig)
print(f"✅ PNG сохранён: {out_png}")

print(f"DATASET={DATASET} | Shape mode: {SHAPE_MODE} | Tick mode: {TICK_MODE}")
print(f"Axis titles: X='{X_AXIS_TITLE}' | Y='{Y_AXIS_TITLE}'")
