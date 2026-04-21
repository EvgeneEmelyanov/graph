import os
import re
from pathlib import Path

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.ticker import FuncFormatter, ScalarFormatter
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries


# ============================================================
# НАСТРОЙКИ
# ============================================================
# FILE = r"D:\1.xlsx"
# FILE = r"D:\discharge_current.xlsx"
FILE = r"D:\reserve_level.xlsx"
SHEET = "SWEEP_2"
OUTPUT_DIR = r"D:\results"

# Выбор критерия по номеру // 1, 2, 3, 4, 5, 8, 9, 10
TARGET_INDEX = 1

# Типовой размер блока
# MATRIX_RANGE = "A2:T13"
# MATRIX_RANGE = "A2:T43"
MATRIX_RANGE = "A2:AH43"

# Режим построения: "3D" или "2D"
PLOT_MODE = "3D"

# Подписи осей
X_AXIS_LABEL = "Мин. уровень заряда"
# X_AXIS_LABEL = "Номинальная мощность ДГУ, кВт"
# X_AXIS_LABEL = "Максимальный ток разряда, С"
Y_AXIS_LABEL = "Доля емкости СНЭ"
# Y_AXIS_LABEL = "Уровень загрузки ДГУ"

FIGSIZE = (12, 8)
DPI = 350

# Поворот для 3D
ELEV = 28
AZIM = 35

SURFACE_ALPHA = 0.98
EDGE_COLOR = "#6f6f6f"
EDGE_LINEWIDTH = 0.25
SHOW_COLORBAR = True

# Для 2D
CONTOUR_LEVELS = 120
CONTOUR_FILLED = True

# --------------------------------
# Обрезка данных
# x_trim = (убрать_слева, убрать_справа)
# y_trim = (убрать_сверху, убрать_снизу)
# --------------------------------
X_TRIM = (0, 0)
Y_TRIM = (0, 0)

# --------------------------------
# Прореживание подписей осей
# --------------------------------
TICK_STEP_X = 8
TICK_STEP_Y = 8
TICK_STEP_Z = 1

# Форматирование подписей
TICK_LABEL_DECIMALS_X = 3
TICK_LABEL_DECIMALS_Y = 3
TICK_LABEL_DECIMALS_Z = 3
COLORBAR_DECIMALS = 3

# Размеры шрифтов
TITLE_FONTSIZE = 18
AXIS_LABEL_FONTSIZE = 16
TICK_LABEL_FONTSIZE = 13
COLORBAR_LABEL_FONTSIZE = 15
COLORBAR_TICK_FONTSIZE = 12
Z_AXIS_SCALE_FONTSIZE = 12

# Отступы подписей
X_LABELPAD_3D = 10
Y_LABELPAD_3D = 12
X_LABELPAD_2D = 10
Y_LABELPAD_2D = 10

# Поля фигуры
LEFT_MARGIN = 0.10
RIGHT_MARGIN = 0.88
BOTTOM_MARGIN = 0.12
TOP_MARGIN = 0.96

# Положение подписи масштаба оси Z (в координатах axes fraction)
# Можно подвинуть при необходимости
Z_SCALE_TEXT_X = 0.02
Z_SCALE_TEXT_Y = 0.67

MATTE_DIFF_CMAP = LinearSegmentedColormap.from_list(
    "matte_diff",
    [
        "#4d7a99",
        "#8fa8b7",
        "#e9e9e9",
        "#ddb48c",
        "#c98d5b",
    ]
)

# ============================================================
# ПОДПИСИ КРИТЕРИЕВ
# ============================================================
METRICS = [
    ("LCOE, руб/кВт∙ч", "LCOE, руб/кВт·ч"),
    ("Расход топлива, тыс.тонн", "Расход топлива, тыс. т"),
    ("Моточасы, тыс.мч", "Моточасы, тыс. мч"),
    ("ENS,кВт∙ч", "ENS, кВт·ч"),
    ("LOLH", "LOLH, ч"),
    ("ENS_evtN", "Кол-во событий ENS"),
    ("ENS_evtAvgH", "ENS_evtAvgH"),
    ("ENS_evtMaxH", "Макс. длит. ENS, ч"),
    ("LOLP", "LOLP"),
    ("LPSP", "LPSP"),
    ("ENS1_mean", "ENS1_mean, кВт·ч"),
    ("ENS2_mean", "ENS2_mean, кВт·ч"),
    ("FailDg", "Отказы ДГУ, кол-во"),
    ("FailBt", "Отказы АКБ, кол-во"),
    ("BtRepl", "Замены АКБ, кол-во"),
]

METRIC_LABELS = dict(METRICS)


# ============================================================
# ФОРМАТТЕРЫ
# ============================================================
class CommaScalarFormatter(ScalarFormatter):
    def __call__(self, x, pos=None):
        s = super().__call__(x, pos)
        return s.replace(".", ",")


# ============================================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ============================================================
def ensure_dir(path: str):
    Path(path).mkdir(parents=True, exist_ok=True)


def safe_filename(name: str):
    name = str(name).strip()
    name = re.sub(r'[\\/*?:"<>|]', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    name = name.replace("∙", "_")
    name = name.replace("·", "_")
    name = name.replace(",", "_")
    return name


def is_numeric_like(value):
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    s = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        float(s)
        return True
    except Exception:
        return False


def to_float(value):
    if value is None or str(value).strip() == "":
        return np.nan
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "").replace(",", ".")
    return float(s)


def read_matrix_shape(matrix_range: str):
    min_col, min_row, max_col, max_row = range_boundaries(matrix_range)

    total_rows = max_row - min_row + 1
    total_cols = max_col - min_col + 1

    if total_rows < 3 or total_cols < 3:
        raise ValueError("MATRIX_RANGE слишком мал. Нужен минимум 3x3.")

    return {
        "min_col": min_col,
        "min_row": min_row,
        "max_col": max_col,
        "max_row": max_row,
        "total_rows": total_rows,
        "total_cols": total_cols,
    }


def validate_trim(trim, axis_name):
    if not isinstance(trim, (tuple, list)) or len(trim) != 2:
        raise ValueError(f"{axis_name}_TRIM должен быть парой из 2 чисел, например (1, 2)")
    start_trim, end_trim = trim
    if not isinstance(start_trim, int) or not isinstance(end_trim, int):
        raise ValueError(f"{axis_name}_TRIM должен содержать целые числа")
    if start_trim < 0 or end_trim < 0:
        raise ValueError(f"{axis_name}_TRIM не может содержать отрицательные значения")
    return start_trim, end_trim


def apply_axis_trim(x_raw, y_raw, z, x_trim=(0, 0), y_trim=(0, 0)):
    x_left, x_right = validate_trim(x_trim, "X")
    y_top, y_bottom = validate_trim(y_trim, "Y")

    x_len = len(x_raw)
    y_len = len(y_raw)

    if x_left + x_right >= x_len:
        raise ValueError(f"Слишком большая обрезка по X: {x_trim}, доступно значений: {x_len}")

    if y_top + y_bottom >= y_len:
        raise ValueError(f"Слишком большая обрезка по Y: {y_trim}, доступно значений: {y_len}")

    x_end = x_len - x_right if x_right > 0 else x_len
    y_end = y_len - y_bottom if y_bottom > 0 else y_len

    x_cut = x_raw[x_left:x_end]
    y_cut = y_raw[y_top:y_end]
    z_cut = z[y_top:y_end, x_left:x_end]

    return x_cut, y_cut, z_cut


def make_tick_values(values, step):
    if step is None or step < 1:
        raise ValueError("Шаг прореживания подписей должен быть целым числом >= 1")
    arr = np.asarray(values, dtype=float)
    return arr[::step]


def format_number_with_comma(value, decimals=3):
    if np.isnan(value):
        return ""
    s = f"{value:.{decimals}f}".rstrip("0").rstrip(".")
    return s.replace(".", ",")


def format_tick_labels(values, decimals=3):
    return [format_number_with_comma(v, decimals=decimals) for v in values]


def comma_formatter(decimals=3):
    def _formatter(x, pos):
        if not np.isfinite(x):
            return ""
        s = f"{x:.{decimals}f}".rstrip("0").rstrip(".")
        return s.replace(".", ",")
    return FuncFormatter(_formatter)


def detect_criteria(ws, matrix_range: str):
    shape = read_matrix_shape(matrix_range)

    min_col = shape["min_col"]
    criteria = []

    for r in range(1, ws.max_row + 1):
        title = ws.cell(r, min_col).value
        if title is None:
            continue

        title_str = str(title).strip()
        if not title_str:
            continue

        if is_numeric_like(title_str):
            continue

        matrix_start_row = r + 1
        matrix_end_row = matrix_start_row + shape["total_rows"] - 1

        if matrix_end_row > ws.max_row:
            continue

        ok = True

        for c in range(min_col + 1, shape["max_col"] + 1):
            if not is_numeric_like(ws.cell(matrix_start_row, c).value):
                ok = False
                break

        if ok:
            for rr in range(matrix_start_row + 1, matrix_end_row + 1):
                if not is_numeric_like(ws.cell(rr, min_col).value):
                    ok = False
                    break

        if ok:
            for rr in range(matrix_start_row + 1, matrix_end_row + 1):
                for c in range(min_col + 1, shape["max_col"] + 1):
                    if not is_numeric_like(ws.cell(rr, c).value):
                        ok = False
                        break
                if not ok:
                    break

        if ok:
            criteria.append(
                {
                    "excel_title": title_str,
                    "display_title": METRIC_LABELS.get(title_str, title_str),
                    "matrix_start_row": matrix_start_row,
                }
            )

    return criteria


def extract_matrix(ws, matrix_start_row: int, matrix_range: str):
    shape = read_matrix_shape(matrix_range)

    min_col = shape["min_col"]
    max_col = shape["max_col"]
    total_rows = shape["total_rows"]

    matrix_end_row = matrix_start_row + total_rows - 1

    x = [to_float(ws.cell(matrix_start_row, c).value)
         for c in range(min_col + 1, max_col + 1)]

    y = [to_float(ws.cell(r, min_col).value)
         for r in range(matrix_start_row + 1, matrix_end_row + 1)]

    z = []
    for r in range(matrix_start_row + 1, matrix_end_row + 1):
        row_vals = []
        for c in range(min_col + 1, max_col + 1):
            row_vals.append(to_float(ws.cell(r, c).value))
        z.append(row_vals)

    x = np.array(x, dtype=float)
    y = np.array(y, dtype=float)
    z = np.array(z, dtype=float)

    return x, y, z


def compute_axis_scale_exponent(values):
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]

    if arr.size == 0:
        return 0

    vmax = np.nanmax(np.abs(arr))
    if vmax == 0 or not np.isfinite(vmax):
        return 0

    exp = int(np.floor(np.log10(vmax)))

    if -3 < exp < 3:
        return 0

    return exp


def build_scaled_comma_formatter(scale_exp, decimals=3):
    scale = 10 ** scale_exp

    def _formatter(x, pos):
        if not np.isfinite(x):
            return ""
        v = x / scale
        s = f"{v:.{decimals}f}".rstrip("0").rstrip(".")
        return s.replace(".", ",")

    return FuncFormatter(_formatter)


def apply_sparse_z_ticks_and_scale(ax, z_values, step, decimals):
    all_zticks = np.asarray(ax.get_zticks(), dtype=float)
    all_zticks = all_zticks[np.isfinite(all_zticks)]

    if len(all_zticks) == 0:
        return 0

    sparse_zticks = all_zticks[::step]
    if len(sparse_zticks) == 0:
        sparse_zticks = all_zticks

    if all_zticks[-1] not in sparse_zticks:
        sparse_zticks = np.append(sparse_zticks, all_zticks[-1])

    ax.set_zticks(sparse_zticks)

    z_exp = compute_axis_scale_exponent(z_values)
    z_formatter = build_scaled_comma_formatter(z_exp, decimals=decimals)
    ax.zaxis.set_major_formatter(z_formatter)

    ax.tick_params(axis="z", labelsize=TICK_LABEL_FONTSIZE, pad=6)

    return z_exp


def add_manual_z_scale_text(ax, exponent):
    if exponent == 0:
        return

    scale_text = rf"$\times 10^{{{exponent}}}$"

    ax.text2D(
        Z_SCALE_TEXT_X,
        Z_SCALE_TEXT_Y,
        scale_text,
        transform=ax.transAxes,
        fontsize=Z_AXIS_SCALE_FONTSIZE,
        ha="left",
        va="bottom"
    )


def setup_common_2d_axis(ax, x, y):
    xticks = make_tick_values(x, TICK_STEP_X)
    yticks = make_tick_values(y, TICK_STEP_Y)

    ax.set_xticks(xticks)
    ax.set_yticks(yticks)

    ax.set_xticklabels(
        format_tick_labels(xticks, decimals=TICK_LABEL_DECIMALS_X),
        fontsize=TICK_LABEL_FONTSIZE
    )
    ax.set_yticklabels(
        format_tick_labels(yticks, decimals=TICK_LABEL_DECIMALS_Y),
        fontsize=TICK_LABEL_FONTSIZE
    )

    ax.set_xlabel(X_AXIS_LABEL, fontsize=AXIS_LABEL_FONTSIZE, labelpad=X_LABELPAD_2D)
    ax.set_ylabel(Y_AXIS_LABEL, fontsize=AXIS_LABEL_FONTSIZE, labelpad=Y_LABELPAD_2D)

    ax.tick_params(axis="x", labelsize=TICK_LABEL_FONTSIZE, pad=6)
    ax.tick_params(axis="y", labelsize=TICK_LABEL_FONTSIZE, pad=6)


def save_figure(fig, display_title, mode_suffix):
    ensure_dir(OUTPUT_DIR)
    out_name = safe_filename(f"{display_title}_{mode_suffix}") + ".png"
    out_path = os.path.join(OUTPUT_DIR, out_name)
    fig.savefig(out_path, dpi=DPI, bbox_inches="tight", pad_inches=0.08)
    plt.close(fig)
    print(f"Готово: {out_path}")


def create_colorbar(fig, ax, mappable, display_title, shrink, pad, aspect, labelpad):
    cbar = fig.colorbar(
        mappable,
        ax=ax,
        shrink=shrink,
        pad=pad,
        aspect=aspect
    )

    formatter = CommaScalarFormatter(useMathText=True)
    formatter.set_powerlimits((-3, 3))
    formatter.set_scientific(True)

    cbar.formatter = formatter
    cbar.update_ticks()

    cbar.ax.tick_params(labelsize=COLORBAR_TICK_FONTSIZE)

    cbar.set_label(
        display_title,
        fontsize=COLORBAR_LABEL_FONTSIZE,
        rotation=90,
        labelpad=labelpad
    )

    offset_text = cbar.ax.yaxis.get_offset_text()
    offset_text.set_size(COLORBAR_TICK_FONTSIZE)

    return cbar


def plot_3d_surface(display_title, x, y, Z):
    X, Y = np.meshgrid(x, y)

    fig = plt.figure(figsize=FIGSIZE, dpi=DPI)
    ax = fig.add_subplot(111, projection="3d")

    surf = ax.plot_surface(
        X, Y, Z,
        cmap=MATTE_DIFF_CMAP,
        linewidth=EDGE_LINEWIDTH,
        edgecolor=EDGE_COLOR,
        alpha=SURFACE_ALPHA,
        antialiased=True,
    )

    xticks = make_tick_values(x, TICK_STEP_X)
    yticks = make_tick_values(y, TICK_STEP_Y)

    ax.set_xticks(xticks)
    ax.set_yticks(yticks)

    ax.set_xticklabels(
        format_tick_labels(xticks, decimals=TICK_LABEL_DECIMALS_X),
        fontsize=TICK_LABEL_FONTSIZE
    )
    ax.set_yticklabels(
        format_tick_labels(yticks, decimals=TICK_LABEL_DECIMALS_Y),
        fontsize=TICK_LABEL_FONTSIZE
    )

    ax.set_xlabel(X_AXIS_LABEL, fontsize=AXIS_LABEL_FONTSIZE, labelpad=X_LABELPAD_3D)
    ax.set_ylabel(Y_AXIS_LABEL, fontsize=AXIS_LABEL_FONTSIZE, labelpad=Y_LABELPAD_3D)
    ax.set_zlabel("")

    ax.view_init(elev=ELEV, azim=AZIM)

    ax.tick_params(axis="x", labelsize=TICK_LABEL_FONTSIZE, pad=4)
    ax.tick_params(axis="y", labelsize=TICK_LABEL_FONTSIZE, pad=4)
    ax.tick_params(axis="z", labelsize=TICK_LABEL_FONTSIZE, pad=6)

    z_exp = apply_sparse_z_ticks_and_scale(
        ax=ax,
        z_values=Z,
        step=TICK_STEP_Z,
        decimals=TICK_LABEL_DECIMALS_Z
    )

    ax.set_box_aspect((1.25, 1.0, 0.75))

    fig.subplots_adjust(
        left=LEFT_MARGIN,
        right=RIGHT_MARGIN,
        bottom=BOTTOM_MARGIN,
        top=TOP_MARGIN
    )

    if SHOW_COLORBAR:
        create_colorbar(
            fig=fig,
            ax=ax,
            mappable=surf,
            display_title=display_title,
            shrink=0.82,
            pad=0.04,
            aspect=22,
            labelpad=16
        )

    fig.canvas.draw()

    # Скрываем стандартный offset у оси Z, чтобы не мешал
    ax.zaxis.get_offset_text().set_visible(False)

    # Добавляем свою подпись масштаба левее и над осью Z
    add_manual_z_scale_text(ax, z_exp)

    save_figure(fig, display_title, "3D")


def plot_2d_contour(display_title, x, y, Z):
    X, Y = np.meshgrid(x, y)

    fig, ax = plt.subplots(figsize=FIGSIZE, dpi=DPI)

    if CONTOUR_FILLED:
        mappable = ax.contourf(
            X, Y, Z,
            levels=CONTOUR_LEVELS,
            cmap=MATTE_DIFF_CMAP
        )
    else:
        mappable = ax.pcolormesh(
            X, Y, Z,
            cmap=MATTE_DIFF_CMAP,
            shading="auto"
        )

    setup_common_2d_axis(ax, x, y)

    fig.subplots_adjust(
        left=0.11,
        right=0.86,
        bottom=0.12,
        top=0.96
    )

    if SHOW_COLORBAR:
        create_colorbar(
            fig=fig,
            ax=ax,
            mappable=mappable,
            display_title=display_title,
            shrink=0.96,
            pad=0.03,
            aspect=25,
            labelpad=14
        )

    save_figure(fig, display_title, "2D")


def plot_metric(excel_title, display_title, x, y, Z):
    x, y, Z = apply_axis_trim(x, y, Z, x_trim=X_TRIM, y_trim=Y_TRIM)

    mode = str(PLOT_MODE).strip().upper()

    if mode == "3D":
        plot_3d_surface(display_title, x, y, Z)
    elif mode == "2D":
        plot_2d_contour(display_title, x, y, Z)
    else:
        raise ValueError('PLOT_MODE должен быть "3D" или "2D"')


# ============================================================
# ОСНОВНОЙ КОД
# ============================================================
def main():
    wb = load_workbook(FILE, data_only=True)
    ws = wb[SHEET]

    criteria = detect_criteria(ws, MATRIX_RANGE)

    if not criteria:
        raise RuntimeError("Критерии не найдены")

    print("\nСписок критериев:")
    for i, c in enumerate(criteria, start=1):
        print(f"{i}. {c['excel_title']}  -->  {c['display_title']}")

    if TARGET_INDEX < 1 or TARGET_INDEX > len(criteria):
        raise ValueError(f"Неверный TARGET_INDEX: {TARGET_INDEX}")

    selected = criteria[TARGET_INDEX - 1]

    print("\nВыбран критерий:")
    print(f"Excel:   {selected['excel_title']}")
    print(f"Подпись: {selected['display_title']}")
    print(f"Режим:   {PLOT_MODE}")

    x, y, Z = extract_matrix(
        ws,
        selected["matrix_start_row"],
        MATRIX_RANGE
    )

    plot_metric(
        excel_title=selected["excel_title"],
        display_title=selected["display_title"],
        x=x,
        y=y,
        Z=Z
    )


if __name__ == "__main__":
    main()