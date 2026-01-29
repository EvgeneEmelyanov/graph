import re
from pathlib import Path
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet


SRC_SHEET = "SWEEP_2"
DST_SHEET = "SWEEP_2"

DEFAULT_SRC_DIR = r"D:\\"
DEFAULT_OUT_PATH = r"D:\comb.xlsx"

ORDER = ["SN1", "SN2", "SS", "D"]
GAP_COLS = 1


def used_range(ws: Worksheet):
    return 1, 1, ws.max_row, ws.max_column


# ---------------- values vs styles cell copy ----------------

def copy_cell_value_only(src_value_cell, dst):
    # только значение (без формул)
    dst.value = src_value_cell.value


def copy_cell_style_only(src_style_cell, dst):
    # только стиль/форматы
    if src_style_cell.has_style:
        dst._style = copy(src_style_cell._style)
    dst.number_format = src_style_cell.number_format
    dst.font = copy(src_style_cell.font)
    dst.fill = copy(src_style_cell.fill)
    dst.border = copy(src_style_cell.border)
    dst.alignment = copy(src_style_cell.alignment)
    dst.protection = copy(src_style_cell.protection)
    # комментарий опционально (можно убрать)
    dst.comment = copy(src_style_cell.comment) if src_style_cell.comment else None


def copy_dimensions(src_ws: Worksheet, dst_ws: Worksheet, col_offset: int):
    for col_letter, dim in src_ws.column_dimensions.items():
        if not dim.width:
            continue
        src_col_idx = openpyxl_col_to_int(col_letter)
        dst_col_idx = src_col_idx + col_offset
        dst_letter = get_column_letter(dst_col_idx)
        dst_ws.column_dimensions[dst_letter].width = dim.width

    for row_idx, dim in src_ws.row_dimensions.items():
        if dim.height:
            dst_ws.row_dimensions[row_idx].height = dim.height


def openpyxl_col_to_int(col_letter: str) -> int:
    col_letter = col_letter.upper()
    n = 0
    for ch in col_letter:
        n = n * 26 + (ord(ch) - ord('A') + 1)
    return n


def copy_merged_cells(src_ws: Worksheet, dst_ws: Worksheet, col_offset: int):
    for mr in src_ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = mr.bounds
        dst_ws.merge_cells(
            start_row=min_row,
            start_column=min_col + col_offset,
            end_row=max_row,
            end_column=max_col + col_offset,
        )


# ---------------- conditional formatting copy with shift ----------------

_CELL_RE = re.compile(r'^(\$?)([A-Z]{1,3})(\$?)(\d+)$')


def _shift_cell_a1(cell: str, col_offset: int) -> str:
    m = _CELL_RE.match(cell.upper())
    if not m:
        raise ValueError(f"Bad cell ref: {cell}")
    col_abs, col_letters, row_abs, row_digits = m.groups()
    col_idx = column_index_from_string(col_letters) + col_offset
    return f"{col_abs}{get_column_letter(col_idx)}{row_abs}{row_digits}"


def shift_a1_range_cols(a1_range: str, col_offset: int) -> str:
    parts = str(a1_range).replace(",", " ").split()
    out = []
    for part in parts:
        if ":" in part:
            min_col, min_row, max_col, max_row = range_boundaries(part)
            min_col += col_offset
            max_col += col_offset
            out.append(
                f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            )
        else:
            out.append(_shift_cell_a1(part, col_offset))
    return " ".join(out)


def copy_conditional_formatting(src_ws: Worksheet, dst_ws: Worksheet, col_offset: int):
    cfs = src_ws.conditional_formatting
    rules_map = getattr(cfs, "_cf_rules", None)
    if not rules_map:
        return

    for cf_obj, rules in rules_map.items():
        sqref = getattr(cf_obj, "sqref", None)
        if sqref is None:
            continue

        dst_sqref = shift_a1_range_cols(str(sqref), col_offset)
        for rule in rules:
            dst_ws.conditional_formatting.add(dst_sqref, copy(rule))


# ---------------- block copy ----------------

def copy_sheet_block(src_ws_values: Worksheet, src_ws_style: Worksheet, dst_ws: Worksheet, col_offset: int):
    min_row, min_col, max_row, max_col = used_range(src_ws_values)

    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            v_cell = src_ws_values.cell(row=r, column=c)   # data_only=True (значение)
            s_cell = src_ws_style.cell(row=r, column=c)    # data_only=False (стиль/формула)
            dst_cell = dst_ws.cell(row=r, column=c + col_offset)

            copy_cell_value_only(v_cell, dst_cell)
            copy_cell_style_only(s_cell, dst_cell)

    copy_merged_cells(src_ws_style, dst_ws, col_offset)
    copy_dimensions(src_ws_style, dst_ws, col_offset)
    copy_conditional_formatting(src_ws_style, dst_ws, col_offset)


def main(src_dir: str, out_path: str):
    src_dir = Path(src_dir)
    out_path = Path(out_path)

    out_wb = Workbook()
    if out_wb.active.title == "Sheet":
        out_wb.remove(out_wb.active)
    dst_ws = out_wb.create_sheet(DST_SHEET)

    col_offset = 0
    block_width = None

    for name in ORDER:
        src_file = src_dir / f"{name}.xlsx"
        if not src_file.exists():
            raise FileNotFoundError(f"Не найден файл: {src_file}")

        # 1) значения (Excel должен быть уже пересчитан и сохранён!)
        wb_values = load_workbook(src_file, data_only=True)

        # 2) стили/CF/формулы (для оформления и CF)
        wb_style = load_workbook(src_file, data_only=False)

        if SRC_SHEET not in wb_values.sheetnames or SRC_SHEET not in wb_style.sheetnames:
            raise ValueError(f"В файле {src_file.name} нет листа {SRC_SHEET}")

        ws_values = wb_values[SRC_SHEET]
        ws_style = wb_style[SRC_SHEET]

        min_row, min_col, max_row, max_col = used_range(ws_values)
        current_width = (max_col - min_col + 1)

        if block_width is None:
            block_width = current_width
        elif current_width != block_width:
            raise ValueError(
                f"Размер блока отличается: {src_file.name} имеет ширину {current_width}, ожидалось {block_width}"
            )

        copy_sheet_block(ws_values, ws_style, dst_ws, col_offset)

        col_offset += block_width + GAP_COLS

        wb_values.close()
        wb_style.close()

    out_wb.save(out_path)
    print(f"OK: сохранено {out_path}")


if __name__ == "__main__":
    main(DEFAULT_SRC_DIR, DEFAULT_OUT_PATH)
