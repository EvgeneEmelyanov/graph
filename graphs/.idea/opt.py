import os
from dataclasses import dataclass
from typing import List, Tuple, Optional, Dict

import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from matplotlib.patches import Rectangle
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter


# =========================
# НАСТРОЙКИ
# =========================

EXCEL_PATH = r"D:\D.xlsx"
SHEET_NAME = "SWEEP_2"
OUTPUT_DIR = r"D:\results"

BASE_RANGE = "A2:IE6"

X_AXIS_LABEL = "Мощность ДГУ, кВт"
Y_AXIS_LABEL = "Количество ДГУ"

DRAW_CELL_TEXT = False

SHOW_CELL_GRID = True
CELL_GRID_LW = 0.35
CELL_GRID_ALPHA = 0.35

X_LABEL_ROTATION = 90

SCENARIO_AGGREGATION = "mean"   # "mean" / "geometric"

WHITE_ORANGE = LinearSegmentedColormap.from_list(
    "white_orange",
    [(1.0, 1.0, 1.0), (1.0, 0.85, 0.6), (1.0, 0.55, 0.0)],
)

# =========================
# ШТРИХОВКА ОПТИМАЛЬНЫХ ЗОН
# =========================

# LCOE / Экономика
HATCH1 = "---"
HATCH1_COLOR = "black"

# Надежность
HATCH2 = "|||"
HATCH2_COLOR = "black"

# Прозрачность фона hatch-прямоугольника
HATCH_FACE_ALPHA = 0.0

# Толщина видимой границы прямоугольника.
# Оставляем очень маленькой, чтобы hatch рисовался корректно,
# но визуально рамка почти не читалась.
HATCH_EDGE_LINEWIDTH = 0.01

# =========================
# КРИТЕРИИ
# =========================

CRITERIA = [
    {
        "name": "LCOE",
        "row_offset": 0,
        "objective_mode": "min",
        "target_value": None,
        "weight": 1.0,
    },
    {
        "name": "ENS",
        "row_offset": 24,
        "objective_mode": "min",
        "target_value": None,
        "weight": 1.0,
    },
    {
        "name": "LOLH",
        "row_offset": 32,
        "objective_mode": "min",
        "target_value": None,
        "weight": 1.0,
    },
    {
        "name": "LOLP",
        "row_offset": 56,
        "objective_mode": "min",
        "target_value": None,
        "weight": 1.0,
    },
    {
        "name": "LPSP",
        "row_offset": 64,
        "objective_mode": "min",
        "target_value": None,
        "weight": 1.0,
    },
]

GROUPS = [
    {
        "name": "Надежность",
        "criteria": ["ENS", "LOLH", "LOLP", "LPSP"],
        "weight": 1.0,
    },
    {
        "name": "Экономика",
        "criteria": ["LCOE"],
        "weight": 1.0,
    },
]

# =========================
# РЕЖИМ ВИЗУАЛИЗАЦИИ
# =========================
# "single" -> одна оценка
# "dual"   -> две оценки одновременно
PLOT_MODE_STYLE = "dual"

# -------------------------
# SINGLE MODE
# -------------------------
# "criterion" / "group" / "multi"
PLOT_MODE = "group"
THRESHOLD_SHARE = 0.999

SELECTED_CRITERION = "LCOE"
SELECTED_GROUP = "Надежность"

MULTI_SELECTION = [
    {"type": "group", "name": "Экономика", "weight": 1.0},
    {"type": "group", "name": "Надежность", "weight": 1.0},
]

# -------------------------
# DUAL MODE
# -------------------------
DUAL_1 = {
    "scope_type": "criterion",   # "criterion" / "group" / "multi"
    "criterion_name": "LCOE",
    "group_name": "Экономика",
    "multi_selection": [
        {"type": "group", "name": "Экономика", "weight": 1.0},
    ],
    "threshold_share": 0.99,
    "hatch": HATCH1,
    "hatch_color": HATCH1_COLOR,
    "label": "Экономика / LCOE",
}

DUAL_2 = {
    "scope_type": "group",
    "criterion_name": "ENS",
    "group_name": "Надежность",
    "multi_selection": [
        {"type": "group", "name": "Надежность", "weight": 1.0},
    ],
    "threshold_share": 0.9999,
    "hatch": HATCH2,
    "hatch_color": HATCH2_COLOR,
    "label": "Надежность",
}

OUTPUT_BASENAME = "heat_map"


# =========================
# DATA MODEL
# =========================

@dataclass
class Table2D:
    x_labels: List[str]
    y_labels: List[str]
    values: np.ndarray


@dataclass
class CriterionSpec:
    name: str
    row_offset: int
    block_range: str
    objective_mode: str
    target_value: Optional[float]
    weight: float = 1.0


@dataclass
class CriterionData:
    name: str
    block_range: str
    scenarios: List[Table2D]


@dataclass
class GroupSpec:
    name: str
    criteria: List[str]
    weight: float = 1.0


@dataclass
class ScoreResult:
    table: Table2D
    best_percent: float
    threshold_percent: float
    near_opt_mask: np.ndarray
    title_suffix: str = ""


# =========================
# HELPERS
# =========================

def _to_float_ru_or_nan(v) -> float:
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).replace("\xa0", "").replace(" ", "").strip()
    if s == "":
        return np.nan
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return np.nan


def build_shifted_range(base_range: str, row_offset: int) -> str:
    min_col, min_row, max_col, max_row = range_boundaries(base_range)
    new_min_row = min_row + row_offset
    new_max_row = max_row + row_offset
    c1 = get_column_letter(min_col)
    c2 = get_column_letter(max_col)
    return f"{c1}{new_min_row}:{c2}{new_max_row}"


def _is_col_empty(ws, col: int, min_row: int, max_row: int) -> bool:
    for r in range(min_row, max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip() != "":
            return False
    return True


def _find_nonempty_col_groups(
    ws,
    min_col: int,
    max_col: int,
    min_row: int,
    max_row: int
) -> List[Tuple[int, int]]:
    groups = []
    in_group = False
    start = None

    for c in range(min_col, max_col + 1):
        empty = _is_col_empty(ws, c, min_row, max_row)

        if not empty and not in_group:
            start = c
            in_group = True
        elif empty and in_group:
            groups.append((start, c - 1))
            start = None
            in_group = False

    if in_group:
        groups.append((start, max_col))

    return groups


def _read_table2d_from_group(ws, col_start: int, col_end: int, row_start: int, row_end: int) -> Table2D:
    x_labels = []
    for c in range(col_start + 1, col_end + 1):
        v = ws.cell(row=row_start, column=c).value
        x_labels.append("" if v is None else str(v).strip())

    y_labels = []
    rows = []

    for r in range(row_start + 1, row_end + 1):
        yv = ws.cell(row=r, column=col_start).value
        y_label = "" if yv is None else str(yv).strip()

        vals = []
        has_any_value = False
        for c in range(col_start + 1, col_end + 1):
            num = _to_float_ru_or_nan(ws.cell(row=r, column=c).value)
            vals.append(num)
            if np.isfinite(num):
                has_any_value = True

        if y_label == "" and not has_any_value:
            continue

        y_labels.append(y_label)
        rows.append(vals)

    values = np.array(rows, dtype=float)

    if values.ndim != 2:
        raise ValueError("Не удалось прочитать матрицу как 2D-массив.")

    return Table2D(
        x_labels=x_labels,
        y_labels=y_labels,
        values=values
    )


def parse_criterion_block(ws, block_range: str) -> List[Table2D]:
    min_col, min_row, max_col, max_row = range_boundaries(block_range)

    groups = _find_nonempty_col_groups(ws, min_col, max_col, min_row, max_row)
    if not groups:
        raise ValueError(f"В диапазоне {block_range} не найдено ни одной матрицы-сценария.")

    scenarios = []
    for c1, c2 in groups:
        scenarios.append(_read_table2d_from_group(ws, c1, c2, min_row, max_row))

    return scenarios


def validate_same_geometry_in_scenarios(criterion_name: str, scenarios: List[Table2D]) -> None:
    if not scenarios:
        raise ValueError(f"У критерия '{criterion_name}' нет сценариев.")

    ref = scenarios[0]
    ref_shape = ref.values.shape
    ref_x = ref.x_labels
    ref_y = ref.y_labels

    for i, t in enumerate(scenarios[1:], start=2):
        if t.values.shape != ref_shape:
            raise ValueError(
                f"Критерий '{criterion_name}', сценарий {i}: shape={t.values.shape}, ожидалось {ref_shape}"
            )
        if t.x_labels != ref_x:
            raise ValueError(f"Критерий '{criterion_name}', сценарий {i}: другие x_labels.")
        if t.y_labels != ref_y:
            raise ValueError(f"Критерий '{criterion_name}', сценарий {i}: другие y_labels.")


def validate_same_geometry_between_criteria(criteria_data: Dict[str, CriterionData]) -> None:
    if not criteria_data:
        raise ValueError("Нет данных критериев для проверки.")

    names = list(criteria_data.keys())
    ref = criteria_data[names[0]].scenarios[0]

    for name in names[1:]:
        current = criteria_data[name].scenarios[0]
        if current.values.shape != ref.values.shape:
            raise ValueError(
                f"Критерий '{name}' имеет shape={current.values.shape}, ожидалось {ref.values.shape}"
            )
        if current.x_labels != ref.x_labels:
            raise ValueError(f"Критерий '{name}' имеет другие x_labels.")
        if current.y_labels != ref.y_labels:
            raise ValueError(f"Критерий '{name}' имеет другие y_labels.")


def normalize_matrix(values: np.ndarray, objective_mode: str, target_value: float = None) -> np.ndarray:
    arr = np.array(values, dtype=float)
    finite = np.isfinite(arr)

    out = np.full_like(arr, np.nan, dtype=float)

    if not np.any(finite):
        return out

    if objective_mode == "min":
        vmin = np.nanmin(arr)
        vmax = np.nanmax(arr)
        if np.isclose(vmin, vmax):
            out[finite] = 1.0
            return out
        out[finite] = (vmax - arr[finite]) / (vmax - vmin)
        return out

    elif objective_mode == "max":
        vmin = np.nanmin(arr)
        vmax = np.nanmax(arr)
        if np.isclose(vmin, vmax):
            out[finite] = 1.0
            return out
        out[finite] = (arr[finite] - vmin) / (vmax - vmin)
        return out

    elif objective_mode == "target":
        if target_value is None:
            raise ValueError("Для objective_mode='target' нужно задать target_value.")

        dist = np.abs(arr - target_value)
        dmin = np.nanmin(dist)
        dmax = np.nanmax(dist)

        if np.isclose(dmin, dmax):
            out[finite] = 1.0
            return out

        out[finite] = (dmax - dist[finite]) / (dmax - dmin)
        return out

    else:
        raise ValueError(f"Неизвестный objective_mode: {objective_mode}")


def build_criteria_specs(base_range: str, criteria_config: List[Dict]) -> Dict[str, CriterionSpec]:
    out = {}
    for c in criteria_config:
        block_range = build_shifted_range(base_range, int(c["row_offset"]))
        spec = CriterionSpec(
            name=c["name"],
            row_offset=int(c["row_offset"]),
            block_range=block_range,
            objective_mode=c["objective_mode"],
            target_value=c.get("target_value"),
            weight=float(c.get("weight", 1.0)),
        )
        if spec.name in out:
            raise ValueError(f"Критерий '{spec.name}' задан несколько раз.")
        out[spec.name] = spec
    return out


def build_group_specs(group_config: List[Dict], criteria_specs: Dict[str, CriterionSpec]) -> Dict[str, GroupSpec]:
    out = {}
    for g in group_config:
        name = g["name"]
        criteria = list(g["criteria"])
        for c_name in criteria:
            if c_name not in criteria_specs:
                raise ValueError(f"Группа '{name}' ссылается на неизвестный критерий '{c_name}'.")

        spec = GroupSpec(
            name=name,
            criteria=criteria,
            weight=float(g.get("weight", 1.0)),
        )
        if spec.name in out:
            raise ValueError(f"Группа '{spec.name}' задана несколько раз.")
        out[spec.name] = spec
    return out


def parse_all_criteria(
    xlsx_path: str,
    sheet_name: str,
    criteria_specs: Dict[str, CriterionSpec]
) -> Dict[str, CriterionData]:
    wb = load_workbook(xlsx_path, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found.")
        ws = wb[sheet_name]

        result = {}
        for name, spec in criteria_specs.items():
            scenarios = parse_criterion_block(ws, spec.block_range)
            validate_same_geometry_in_scenarios(name, scenarios)

            result[name] = CriterionData(
                name=name,
                block_range=spec.block_range,
                scenarios=scenarios
            )

        return result
    finally:
        wb.close()


def aggregate_scenario_scores(score_matrices: List[np.ndarray], method: str) -> np.ndarray:
    if not score_matrices:
        raise ValueError("Пустой набор score-матриц сценариев.")

    stack = np.stack(score_matrices, axis=0)

    if method == "mean":
        return np.nanmean(stack, axis=0)

    elif method == "geometric":
        eps = 1e-12
        safe = np.where(np.isfinite(stack), np.clip(stack, eps, None), np.nan)
        logv = np.log(safe)
        return np.exp(np.nanmean(logv, axis=0))

    else:
        raise ValueError(f"Неизвестный метод агрегации сценариев: {method}")


def compute_weighted_mean_score(score_matrices: List[np.ndarray], weights: List[float]) -> np.ndarray:
    if not score_matrices:
        raise ValueError("Пустой набор score-матриц.")

    stack = np.stack(score_matrices, axis=0)
    w = np.array(weights, dtype=float)

    if stack.shape[0] != w.shape[0]:
        raise ValueError("Число матриц не совпадает с числом весов.")

    out = np.full(stack.shape[1:], np.nan, dtype=float)

    finite_mask = np.isfinite(stack)
    weighted_stack = np.where(finite_mask, stack * w[:, None, None], 0.0)
    weight_stack = np.where(finite_mask, w[:, None, None], 0.0)

    numerator = np.sum(weighted_stack, axis=0)
    denominator = np.sum(weight_stack, axis=0)

    valid = denominator > 0
    out[valid] = numerator[valid] / denominator[valid]

    return out


def get_criterion_result(
    criteria_data: Dict[str, CriterionData],
    criteria_specs: Dict[str, CriterionSpec],
    criterion_name: str,
    threshold_share: float,
    scenario_aggregation: str
) -> ScoreResult:
    if criterion_name not in criteria_data:
        raise ValueError(f"Критерий '{criterion_name}' не найден.")

    spec = criteria_specs[criterion_name]
    cdata = criteria_data[criterion_name]

    score_matrices = []
    ref = cdata.scenarios[0]

    for scenario_table in cdata.scenarios:
        score = normalize_matrix(
            scenario_table.values,
            objective_mode=spec.objective_mode,
            target_value=spec.target_value
        ) * 100.0
        score_matrices.append(score)

    final_score = aggregate_scenario_scores(score_matrices, scenario_aggregation)

    best_percent = np.nanmax(final_score)
    threshold_percent = threshold_share * best_percent
    near_opt_mask = np.isfinite(final_score) & (final_score >= threshold_percent)

    mode_suffix = f"target={spec.target_value}" if spec.objective_mode == "target" else spec.objective_mode

    result_table = Table2D(
        x_labels=ref.x_labels,
        y_labels=ref.y_labels,
        values=final_score
    )

    return ScoreResult(
        table=result_table,
        best_percent=best_percent,
        threshold_percent=threshold_percent,
        near_opt_mask=near_opt_mask,
        title_suffix=(
            f"критерий: {criterion_name}, режим: {mode_suffix}, "
            f"сценариев: {len(cdata.scenarios)}"
        )
    )


def compute_group_score(
    criteria_data: Dict[str, CriterionData],
    criteria_specs: Dict[str, CriterionSpec],
    group_specs: Dict[str, GroupSpec],
    group_name: str,
    threshold_share: float,
    scenario_aggregation: str
) -> ScoreResult:
    if group_name not in group_specs:
        raise ValueError(f"Группа '{group_name}' не найдена.")

    gspec = group_specs[group_name]

    criterion_score_matrices = []
    criterion_weights = []
    ref_table = None

    for criterion_name in gspec.criteria:
        cres = get_criterion_result(
            criteria_data=criteria_data,
            criteria_specs=criteria_specs,
            criterion_name=criterion_name,
            threshold_share=threshold_share,
            scenario_aggregation=scenario_aggregation
        )

        if ref_table is None:
            ref_table = cres.table

        criterion_score_matrices.append(cres.table.values)
        criterion_weights.append(criteria_specs[criterion_name].weight)

    group_score = compute_weighted_mean_score(criterion_score_matrices, criterion_weights)

    best_percent = np.nanmax(group_score)
    threshold_percent = threshold_share * best_percent
    near_opt_mask = np.isfinite(group_score) & (group_score >= threshold_percent)

    result_table = Table2D(
        x_labels=ref_table.x_labels,
        y_labels=ref_table.y_labels,
        values=group_score
    )

    return ScoreResult(
        table=result_table,
        best_percent=best_percent,
        threshold_percent=threshold_percent,
        near_opt_mask=near_opt_mask,
        title_suffix=f"группа критериев: {group_name}"
    )


def compute_multi_score(
    criteria_data: Dict[str, CriterionData],
    criteria_specs: Dict[str, CriterionSpec],
    group_specs: Dict[str, GroupSpec],
    multi_selection: List[Dict],
    threshold_share: float,
    scenario_aggregation: str
) -> ScoreResult:
    if not multi_selection:
        raise ValueError("MULTI_SELECTION пуст.")

    component_scores = []
    component_weights = []
    ref_table = None
    title_parts = []

    for item in multi_selection:
        item_type = item["type"]
        name = item["name"]
        outer_weight = float(item.get("weight", 1.0))

        if item_type == "criterion":
            res = get_criterion_result(
                criteria_data=criteria_data,
                criteria_specs=criteria_specs,
                criterion_name=name,
                threshold_share=threshold_share,
                scenario_aggregation=scenario_aggregation
            )
            component_scores.append(res.table.values)
            component_weights.append(outer_weight)
            title_parts.append(f"criterion:{name}")

            if ref_table is None:
                ref_table = res.table

        elif item_type == "group":
            res = compute_group_score(
                criteria_data=criteria_data,
                criteria_specs=criteria_specs,
                group_specs=group_specs,
                group_name=name,
                threshold_share=threshold_share,
                scenario_aggregation=scenario_aggregation
            )
            component_scores.append(res.table.values)
            component_weights.append(outer_weight)
            title_parts.append(f"group:{name}")

            if ref_table is None:
                ref_table = res.table

        else:
            raise ValueError(f"Неизвестный item['type']: {item_type}")

    multi_score = compute_weighted_mean_score(component_scores, component_weights)

    best_percent = np.nanmax(multi_score)
    threshold_percent = threshold_share * best_percent
    near_opt_mask = np.isfinite(multi_score) & (multi_score >= threshold_percent)

    result_table = Table2D(
        x_labels=ref_table.x_labels,
        y_labels=ref_table.y_labels,
        values=multi_score
    )

    return ScoreResult(
        table=result_table,
        best_percent=best_percent,
        threshold_percent=threshold_percent,
        near_opt_mask=near_opt_mask,
        title_suffix="multi: " + ", ".join(title_parts)
    )


def evaluate_scope(
    scope_type: str,
    threshold_share: float,
    criteria_data: Dict[str, CriterionData],
    criteria_specs: Dict[str, CriterionSpec],
    group_specs: Dict[str, GroupSpec],
    scenario_aggregation: str,
    criterion_name: Optional[str] = None,
    group_name: Optional[str] = None,
    multi_selection: Optional[List[Dict]] = None
) -> ScoreResult:
    if scope_type == "criterion":
        if not criterion_name:
            raise ValueError("Для scope_type='criterion' нужен criterion_name.")
        return get_criterion_result(
            criteria_data=criteria_data,
            criteria_specs=criteria_specs,
            criterion_name=criterion_name,
            threshold_share=threshold_share,
            scenario_aggregation=scenario_aggregation
        )

    elif scope_type == "group":
        if not group_name:
            raise ValueError("Для scope_type='group' нужен group_name.")
        return compute_group_score(
            criteria_data=criteria_data,
            criteria_specs=criteria_specs,
            group_specs=group_specs,
            group_name=group_name,
            threshold_share=threshold_share,
            scenario_aggregation=scenario_aggregation
        )

    elif scope_type == "multi":
        if not multi_selection:
            raise ValueError("Для scope_type='multi' нужен multi_selection.")
        return compute_multi_score(
            criteria_data=criteria_data,
            criteria_specs=criteria_specs,
            group_specs=group_specs,
            multi_selection=multi_selection,
            threshold_share=threshold_share,
            scenario_aggregation=scenario_aggregation
        )

    else:
        raise ValueError(f"Неизвестный scope_type: {scope_type}")


def _ideal_text_color(val: float) -> str:
    return "white" if val >= 60.0 else "black"


def compute_layout(nx: int, ny: int, draw_cell_text: bool):
    if draw_cell_text:
        cell_w = 0.42
        cell_h = 0.72
        fig_w = max(14.0, nx * cell_w + 3.2)
        fig_h = max(6.0, ny * cell_h + 2.6)
        x_font = 8
        y_font = 9
        axis_font = 11
        cell_font = 7
        bottom_margin = 0.28
    else:
        cell_w = 0.22
        cell_h = 0.52
        fig_w = max(12.0, nx * cell_w + 2.4)
        fig_h = max(4.2, ny * cell_h + 2.0)
        x_font = 7
        y_font = 8
        axis_font = 10
        cell_font = 0
        bottom_margin = 0.22

    return {
        "figsize": (fig_w, fig_h),
        "x_font": x_font,
        "y_font": y_font,
        "axis_font": axis_font,
        "cell_font": cell_font,
        "bottom_margin": bottom_margin,
    }


def add_hatched_cell(ax, x: int, y: int, hatch: str, edgecolor: str):
    rect = Rectangle(
        (x - 0.5, y - 0.5),
        1,
        1,
        facecolor=(1, 1, 1, HATCH_FACE_ALPHA),
        edgecolor=edgecolor,
        linewidth=HATCH_EDGE_LINEWIDTH,
        hatch=hatch
    )
    ax.add_patch(rect)


def resolve_single_hatch(scope_type: str, criterion_name: str, group_name: str):
    if scope_type == "criterion":
        if criterion_name == "LCOE":
            return HATCH1, HATCH1_COLOR, "LCOE / Экономика"
        return HATCH2, HATCH2_COLOR, f"{criterion_name}"

    if scope_type == "group":
        if group_name == "Экономика":
            return HATCH1, HATCH1_COLOR, "Экономика"
        return HATCH2, HATCH2_COLOR, group_name

    # multi
    return HATCH1, HATCH1_COLOR, "Multi"


def plot_single_score_heatmap(
    score_result: ScoreResult,
    threshold_share: float,
    out_path: str,
    draw_cell_text: bool,
    hatch: str,
    hatch_color: str,
    hatch_label: str
) -> None:
    vals = score_result.table.values
    ny, nx = vals.shape

    layout = compute_layout(nx, ny, draw_cell_text)
    fig, ax = plt.subplots(figsize=layout["figsize"])

    im = ax.imshow(
        vals,
        origin="lower",
        aspect="auto",
        interpolation="nearest",
        cmap=WHITE_ORANGE,
        vmin=0.0,
        vmax=100.0
    )

    cbar = plt.colorbar(im, ax=ax)
    cbar.set_label("Итоговая оценка, %", fontsize=layout["axis_font"])

    if SHOW_CELL_GRID:
        ax.set_xticks(np.arange(-0.5, nx, 1), minor=True)
        ax.set_yticks(np.arange(-0.5, ny, 1), minor=True)
        ax.grid(which="minor", color="black", linewidth=CELL_GRID_LW, alpha=CELL_GRID_ALPHA)

    ax.set_xticks(np.arange(nx))
    ax.set_yticks(np.arange(ny))
    ax.set_xticklabels(
        score_result.table.x_labels,
        rotation=X_LABEL_ROTATION,
        ha="center",
        va="top",
        fontsize=layout["x_font"]
    )
    ax.set_yticklabels(score_result.table.y_labels, fontsize=layout["y_font"])
    ax.set_xlabel(X_AXIS_LABEL, fontsize=layout["axis_font"])
    ax.set_ylabel(Y_AXIS_LABEL, fontsize=layout["axis_font"])

    ax.set_title(
        f"Оптимальная зона\n"
        f"{score_result.title_suffix}\n"
        f"Штриховка: {hatch_label}\n"
        f"near-optimal ≥ {threshold_share:.5f} от лучшего "
        f"({score_result.threshold_percent:.1f}% из {score_result.best_percent:.1f}%)",
        fontsize=layout["axis_font"] + 1
    )

    for y in range(ny):
        for x in range(nx):
            v = vals[y, x]

            if draw_cell_text and np.isfinite(v):
                ax.text(
                    x, y, f"{v:.1f}",
                    ha="center",
                    va="center",
                    fontsize=layout["cell_font"],
                    color=_ideal_text_color(v)
                )

            if score_result.near_opt_mask[y, x]:
                add_hatched_cell(ax, x, y, hatch=hatch, edgecolor=hatch_color)

    plt.tight_layout()
    plt.subplots_adjust(bottom=layout["bottom_margin"])
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def plot_dual_score_heatmap(
    base_result: ScoreResult,
    second_result: ScoreResult,
    dual_1: Dict,
    dual_2: Dict,
    out_path: str,
    draw_cell_text: bool
) -> None:
    vals = base_result.table.values
    ny, nx = vals.shape

    mask1 = base_result.near_opt_mask
    mask2 = second_result.near_opt_mask
    both_mask = mask1 & mask2
    only1_mask = mask1 & (~mask2)
    only2_mask = mask2 & (~mask1)

    layout = compute_layout(nx, ny, draw_cell_text)
    fig, ax = plt.subplots(figsize=layout["figsize"])

    im = ax.imshow(
        vals,
        origin="lower",
        aspect="auto",
        interpolation="nearest",
        cmap=WHITE_ORANGE,
        vmin=0.0,
        vmax=100.0
    )

    cbar = plt.colorbar(im, ax=ax)
    cbar.set_label("Итоговая оценка базовой оценки, %", fontsize=layout["axis_font"])

    if SHOW_CELL_GRID:
        ax.set_xticks(np.arange(-0.5, nx, 1), minor=True)
        ax.set_yticks(np.arange(-0.5, ny, 1), minor=True)
        ax.grid(which="minor", color="black", linewidth=CELL_GRID_LW, alpha=CELL_GRID_ALPHA)

    ax.set_xticks(np.arange(nx))
    ax.set_yticks(np.arange(ny))
    ax.set_xticklabels(
        base_result.table.x_labels,
        rotation=X_LABEL_ROTATION,
        ha="center",
        va="top",
        fontsize=layout["x_font"]
    )
    ax.set_yticklabels(base_result.table.y_labels, fontsize=layout["y_font"])
    ax.set_xlabel(X_AXIS_LABEL, fontsize=layout["axis_font"])
    ax.set_ylabel(Y_AXIS_LABEL, fontsize=layout["axis_font"])

    ax.set_title(
        "Оптимальная зона: 2 точки зрения\n"
        f"1) {dual_1['label']} | threshold={dual_1['threshold_share']}\n"
        f"2) {dual_2['label']} | threshold={dual_2['threshold_share']}\n"
        "Штриховка: 1-я / 2-я / обе",
        fontsize=layout["axis_font"] + 1
    )

    for y in range(ny):
        for x in range(nx):
            v = vals[y, x]

            if draw_cell_text and np.isfinite(v):
                ax.text(
                    x, y, f"{v:.1f}",
                    ha="center",
                    va="center",
                    fontsize=layout["cell_font"],
                    color=_ideal_text_color(v)
                )

            if only1_mask[y, x]:
                add_hatched_cell(
                    ax,
                    x, y,
                    hatch=dual_1["hatch"],
                    edgecolor=dual_1["hatch_color"]
                )

            if only2_mask[y, x]:
                add_hatched_cell(
                    ax,
                    x, y,
                    hatch=dual_2["hatch"],
                    edgecolor=dual_2["hatch_color"]
                )

            if both_mask[y, x]:
                add_hatched_cell(
                    ax,
                    x, y,
                    hatch=dual_1["hatch"],
                    edgecolor=dual_1["hatch_color"]
                )
                add_hatched_cell(
                    ax,
                    x, y,
                    hatch=dual_2["hatch"],
                    edgecolor=dual_2["hatch_color"]
                )

    plt.tight_layout()
    plt.subplots_adjust(bottom=layout["bottom_margin"])
    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def build_output_file_name(output_basename: str, style_mode: str, suffix: str) -> str:
    return f"{output_basename}_{style_mode}_{suffix}.png"


# =========================
# MAIN
# =========================

if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    criteria_specs = build_criteria_specs(BASE_RANGE, CRITERIA)
    group_specs = build_group_specs(GROUPS, criteria_specs)

    criteria_data = parse_all_criteria(
        xlsx_path=EXCEL_PATH,
        sheet_name=SHEET_NAME,
        criteria_specs=criteria_specs
    )

    print("Прочитаны критерии:")
    for name, cdata in criteria_data.items():
        spec = criteria_specs[name]
        ref = cdata.scenarios[0]
        print(
            f"  {name}: range={spec.block_range}, "
            f"scenarios={len(cdata.scenarios)}, "
            f"shape={ref.values.shape}"
        )

    validate_same_geometry_between_criteria(criteria_data)

    if PLOT_MODE_STYLE == "single":
        score_result = evaluate_scope(
            scope_type=PLOT_MODE,
            threshold_share=THRESHOLD_SHARE,
            criteria_data=criteria_data,
            criteria_specs=criteria_specs,
            group_specs=group_specs,
            scenario_aggregation=SCENARIO_AGGREGATION,
            criterion_name=SELECTED_CRITERION,
            group_name=SELECTED_GROUP,
            multi_selection=MULTI_SELECTION
        )

        single_hatch, single_hatch_color, single_label = resolve_single_hatch(
            scope_type=PLOT_MODE,
            criterion_name=SELECTED_CRITERION,
            group_name=SELECTED_GROUP
        )

        file_name = build_output_file_name(
            output_basename=OUTPUT_BASENAME,
            style_mode="single",
            suffix=PLOT_MODE
        )
        out_path = os.path.join(OUTPUT_DIR, file_name)

        plot_single_score_heatmap(
            score_result=score_result,
            threshold_share=THRESHOLD_SHARE,
            out_path=out_path,
            draw_cell_text=DRAW_CELL_TEXT,
            hatch=single_hatch,
            hatch_color=single_hatch_color,
            hatch_label=single_label
        )

        print()
        print("=== RESULT SINGLE ===")
        print(f"PLOT_MODE_STYLE         = {PLOT_MODE_STYLE}")
        print(f"PLOT_MODE               = {PLOT_MODE}")
        print(f"SCENARIO_AGGREGATION    = {SCENARIO_AGGREGATION}")
        print(f"best_percent            = {score_result.best_percent:.4f}%")
        print(f"threshold_percent       = {score_result.threshold_percent:.4f}%")
        print(f"Heatmap saved to        = {out_path}")

    elif PLOT_MODE_STYLE == "dual":
        result1 = evaluate_scope(
            scope_type=DUAL_1["scope_type"],
            threshold_share=DUAL_1["threshold_share"],
            criteria_data=criteria_data,
            criteria_specs=criteria_specs,
            group_specs=group_specs,
            scenario_aggregation=SCENARIO_AGGREGATION,
            criterion_name=DUAL_1.get("criterion_name"),
            group_name=DUAL_1.get("group_name"),
            multi_selection=DUAL_1.get("multi_selection")
        )

        result2 = evaluate_scope(
            scope_type=DUAL_2["scope_type"],
            threshold_share=DUAL_2["threshold_share"],
            criteria_data=criteria_data,
            criteria_specs=criteria_specs,
            group_specs=group_specs,
            scenario_aggregation=SCENARIO_AGGREGATION,
            criterion_name=DUAL_2.get("criterion_name"),
            group_name=DUAL_2.get("group_name"),
            multi_selection=DUAL_2.get("multi_selection")
        )

        if result1.table.x_labels != result2.table.x_labels or result1.table.y_labels != result2.table.y_labels:
            raise ValueError("Результаты dual-mode имеют разную геометрию.")

        file_name = build_output_file_name(
            output_basename=OUTPUT_BASENAME,
            style_mode="dual",
            suffix="overlay"
        )
        out_path = os.path.join(OUTPUT_DIR, file_name)

        plot_dual_score_heatmap(
            base_result=result1,
            second_result=result2,
            dual_1=DUAL_1,
            dual_2=DUAL_2,
            out_path=out_path,
            draw_cell_text=DRAW_CELL_TEXT
        )

        both_mask = result1.near_opt_mask & result2.near_opt_mask

        print()
        print("=== RESULT DUAL ===")
        print(f"PLOT_MODE_STYLE         = {PLOT_MODE_STYLE}")
        print(f"SCENARIO_AGGREGATION    = {SCENARIO_AGGREGATION}")
        print(f"Dual #1                 = {DUAL_1['label']}")
        print(f"Dual #1 threshold       = {DUAL_1['threshold_share']}")
        print(f"Dual #2                 = {DUAL_2['label']}")
        print(f"Dual #2 threshold       = {DUAL_2['threshold_share']}")
        print(f"Cells in mask #1        = {int(np.sum(result1.near_opt_mask))}")
        print(f"Cells in mask #2        = {int(np.sum(result2.near_opt_mask))}")
        print(f"Cells in intersection   = {int(np.sum(both_mask))}")
        print(f"Heatmap saved to        = {out_path}")

    else:
        raise ValueError(f"Неизвестный PLOT_MODE_STYLE={PLOT_MODE_STYLE}")