import numpy as np
from openpyxl import load_workbook
import plotly.graph_objects as go
from plotly.subplots import make_subplots

# ============================================================
# 0) НАСТРОЙКИ EXCEL (ОДИН ФАЙЛ)
# ============================================================
# FILE  = r"D:\10_results\02_Scopus\5.1.xlsx"
# FILE  = r"D:\10_results\def.xlsx"
FILE  = r"D:\3.xlsx"
SHEET = "SWEEP_2"

OUT_HTML = r"C:\Users\Balt_\Desktop\4_surfaces.html"
AUTO_OPEN = True  # открыть в браузере после сохранения

# ============================================================
# 1) ВЫБОР НАБОРА ДАННЫХ (1/2/3)
# ============================================================
DATASET = 3  # <-- меняй: 1, 2 или 3

X_ROW = 2
TRI_LIMIT = 1.0000001  # для triangle

if DATASET == 1:
    Y_START_ROW, Y_END_ROW = 3, 9
    X_START_COL, X_END_COL = 2, 12
    # 4 блока Z (строки в Excel): [Fuel, Hours, ENS, Costs]
    Z_BLOCKS = [(3, 9), (14, 20), (25, 31), (36, 42)]
    SHAPE_MODE = "rect"
elif DATASET == 2:
    Y_START_ROW, Y_END_ROW = 3, 12
    X_START_COL, X_END_COL = 2, 12
    Z_BLOCKS = [(3, 12), (17, 26), (31, 40), (45, 54)]
    SHAPE_MODE = "rect"
elif DATASET == 3:
    Y_START_ROW, Y_END_ROW = 3, 13
    X_START_COL, X_END_COL = 2, 12
    Z_BLOCKS = [(3, 13), (33, 43), (48, 58), (63, 73)]
    SHAPE_MODE = "triangle"
else:
    raise ValueError("DATASET должен быть 1, 2 или 3")

# ============================================================
# 2) "УГЛЫ ОБЗОРА" — для Plotly это камера. Сохраняем смысл:
#    elev/azim -> camera eye (приближенно).
# ============================================================
VIEW_ANGLES = {
    "fuel":  dict(elev=22, azim=65),
    "hours": dict(elev=22, azim=35),
    "ens":   dict(elev=22, azim=35),
    "costs": dict(elev=22, azim=35),
}

def view_to_camera(elev_deg, azim_deg, r=1.9):
    # Приближенная конверсия (matplotlib elev/azim -> plotly camera eye)
    elev = np.deg2rad(elev_deg)
    azim = np.deg2rad(azim_deg)
    x = r * np.cos(elev) * np.cos(azim)
    y = r * np.cos(elev) * np.sin(azim)
    z = r * np.sin(elev)
    return dict(eye=dict(x=float(x), y=float(y), z=float(z)))

# ============================================================
# 3) ЦВЕТОВАЯ ШКАЛА (аналог твоего custom cmap)
# ============================================================
colors = ["limegreen", "yellowgreen", "yellow", "orange", "red", "maroon"]
# Plotly colorscale: список [позиция 0..1, цвет]
PLOTLY_COLORSCALE = [[i / (len(colors) - 1), c] for i, c in enumerate(colors)]

# ============================================================
# 4) ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ЧТЕНИЯ ИЗ EXCEL
# ============================================================
def to_float(v):
    """Аккуратный парсер чисел из Excel (учёт строк, пробелов, запятых, ###)."""
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        s = v.strip()
        if not s or s == "###":
            return np.nan
        s = s.replace(" ", "").replace("\u00A0", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return np.nan
    return np.nan

def _cell_to_label(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return str(v).strip()

def read_axis_labels(ws):
    """
    X-лейбл из (row=1, col=2), Y-лейбл из (row=1, col=3).
    """
    xlab = _cell_to_label(ws.cell(row=1, column=2).value)
    ylab = _cell_to_label(ws.cell(row=1, column=3).value)
    return (xlab or "Емкость СНЭ,\nкВт*ч"), (ylab or "Ток разряда, С")

def get_z_label(ws, z_start_row):
    """
    - если блок начинается с Y_START_ROW -> Z из A1
    - иначе Z из (row=z_start_row-2, col=X_START_COL-1)
    """
    if z_start_row == Y_START_ROW:
        return _cell_to_label(ws.cell(row=1, column=1).value)
    return _cell_to_label(ws.cell(row=z_start_row - 2, column=X_START_COL - 1).value)

def read_block(ws, z_start_row, z_end_row):
    """
    Возвращает: x_vals, y_vals, Z (2D), z_label
    """
    zlab = get_z_label(ws, z_start_row)

    x = np.array(
        [to_float(ws.cell(row=X_ROW, column=col).value) for col in range(X_START_COL, X_END_COL + 1)],
        dtype=float
    )
    y = np.array(
        [to_float(ws.cell(row=row, column=1).value) for row in range(Y_START_ROW, Y_END_ROW + 1)],
        dtype=float
    )

    expected_rows = len(y)
    actual_rows = z_end_row - z_start_row + 1
    if actual_rows != expected_rows:
        raise RuntimeError(
            f"Диапазон Z по строкам не совпадает с Y: Y={expected_rows} строк, Z={actual_rows} строк.\n"
            f"Проверь Z_BLOCKS и Y_START_ROW/Y_END_ROW для DATASET={DATASET}."
        )

    Z = np.full((len(y), len(x)), np.nan, dtype=float)
    for i, row in enumerate(range(z_start_row, z_end_row + 1)):
        for j, col in enumerate(range(X_START_COL, X_END_COL + 1)):
            Z[i, j] = to_float(ws.cell(row=row, column=col).value)

    if SHAPE_MODE == "triangle":
        Xg, Yg = np.meshgrid(x, y)
        mask = np.isnan(Z) | ((Xg + Yg) > TRI_LIMIT)
        Z = Z.copy()
        Z[mask] = np.nan

    if np.all(np.isnan(x)) or np.all(np.isnan(y)):
        raise RuntimeError("X или Y не считались (всё NaN). Проверь лист/диапазоны.")
    if np.all(np.isnan(Z)):
        raise RuntimeError(
            "Блок Z полностью NaN. Частая причина: в xlsx нет сохранённого кэша формул "
            "(openpyxl data_only=True читает кэш)."
        )

    return x, y, Z, zlab

# ============================================================
# 5) ЧТЕНИЕ ИЗ EXCEL (ОДИН РАЗ)
# ============================================================
wb = load_workbook(FILE, data_only=True)
ws = wb[SHEET]

X_AXIS_TITLE, Y_AXIS_TITLE = read_axis_labels(ws)

(x_vals, y_vals, Z_fuel,  zlab_fuel)  = read_block(ws, *Z_BLOCKS[0])
(_,      _,      Z_hours, zlab_hours) = read_block(ws, *Z_BLOCKS[1])
(_,      _,      Z_ens,   zlab_ens)   = read_block(ws, *Z_BLOCKS[2])
(_,      _,      Z_costs, zlab_costs) = read_block(ws, *Z_BLOCKS[3])

X, Y = np.meshgrid(x_vals, y_vals)

# ============================================================
# 6) СБОРКА 2×2 В HTML (Plotly)
# ============================================================
fig = make_subplots(
    rows=2, cols=2,
    specs=[[{"type": "surface"}, {"type": "surface"}],
           [{"type": "surface"}, {"type": "surface"}]],
    horizontal_spacing=0.02,
    vertical_spacing=0.05,
    subplot_titles=(
        "Экономические затраты",
        "Недоотпуск энергии",
        "Расход топлива",
        "Моточасы ДГУ",
    )
)

def add_surface(row, col, Z, zlabel, show_colorbar=False):
    Z = np.asarray(Z, dtype=float)
    zmin = float(np.nanmin(Z))
    zmax = float(np.nanmax(Z))

    surface = go.Surface(
        x=X, y=Y, z=Z,
        colorscale=PLOTLY_COLORSCALE,
        cmin=zmin, cmax=zmax,
        showscale=show_colorbar,
        colorbar=dict(
            title=zlabel or "Z",
            len=0.45
        ) if show_colorbar else None,
        hovertemplate=(
            f"{X_AXIS_TITLE}: %{{x}}<br>"
            f"{Y_AXIS_TITLE}: %{{y}}<br>"
            f"{(zlabel or 'Z')}: %{{z}}<extra></extra>"
        )
    )
    fig.add_trace(surface, row=row, col=col)

# Если хочешь одну общую шкалу — можно поставить show_colorbar=True только у первого
add_surface(1, 1, Z_fuel,  zlab_fuel  or "Z", show_colorbar=True)
add_surface(1, 2, Z_hours, zlab_hours or "Z", show_colorbar=False)
add_surface(2, 1, Z_ens,   zlab_ens   or "Z", show_colorbar=False)
add_surface(2, 2, Z_costs, zlab_costs or "Z", show_colorbar=False)

# ============================================================
# 7) ОСИ + КАМЕРА ДЛЯ КАЖДОГО SUBPLOT (scene/scene2/scene3/scene4)
# ============================================================
def scene_layout(x_title, y_title, z_title, view_key):
    cam = view_to_camera(VIEW_ANGLES[view_key]["elev"], VIEW_ANGLES[view_key]["azim"])
    return dict(
        xaxis=dict(title=x_title, backgroundcolor="white", gridcolor="lightgray", zerolinecolor="lightgray"),
        yaxis=dict(title=y_title, backgroundcolor="white", gridcolor="lightgray", zerolinecolor="lightgray"),
        zaxis=dict(title=z_title, backgroundcolor="white", gridcolor="lightgray", zerolinecolor="lightgray"),
        camera=cam
    )

fig.update_layout(
    template="plotly_white",
    width=1300,
    height=1100,
    margin=dict(l=10, r=10, t=70, b=10),
    scene = scene_layout(X_AXIS_TITLE, Y_AXIS_TITLE, (zlab_fuel  or "Z"), "fuel"),
    scene2= scene_layout(X_AXIS_TITLE, Y_AXIS_TITLE, (zlab_hours or "Z"), "hours"),
    scene3= scene_layout(X_AXIS_TITLE, Y_AXIS_TITLE, (zlab_ens   or "Z"), "ens"),
    scene4= scene_layout(X_AXIS_TITLE, Y_AXIS_TITLE, (zlab_costs or "Z"), "costs"),
)

# ============================================================
# 8) СОХРАНЕНИЕ HTML + ОТКРЫТИЕ В БРАУЗЕРЕ
# ============================================================
fig.write_html(OUT_HTML, auto_open=AUTO_OPEN, include_plotlyjs="cdn")

print("Готово:", OUT_HTML)
print(f"FILE={FILE} | SHEET={SHEET} | DATASET={DATASET} | SHAPE_MODE={SHAPE_MODE}")
