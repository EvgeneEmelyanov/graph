import re
from dataclasses import dataclass
from typing import List, Optional, Tuple, Dict, Any

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# =========================
# DEFAULTS (EDIT THESE ONCE)
# =========================

DEFAULT_EXCEL_PATH = r"D:\10_results\combined.2.xlsx"
DEFAULT_SHEET = "SWEEP_2"

DEFAULT_BLOCK_RANGE = "A2:BC13"

# Разрыв между блоками (пример: A2:BC13 -> следующий A17:BC28 => 3 строки пустые)
DEFAULT_GAP_ROWS = 3

# Подписи сценариев по умолчанию (или None чтобы вообще не использовать)
# Должно быть РОВНО столько, сколько таблиц в блоке.
DEFAULT_SCENARIO_LABELS: Optional[List[str]] = [
    "50% SS", "50% D", "100% SS", "100% D",
    "200% SS", "200% D", "400% SS", "400% D",
]

USE_DEFAULT_LABELS = True

# Подписи двух нижних рядов (слева от подписей значений)
TOP_AXIS_NAME = "Мощность ВЭУ"
BOTTOM_AXIS_NAME = "Схема"

# ========== НАСТРОЙКИ СТРЕЛОК (РАЗНЫЕ ДЛЯ ДВУХ ГРАФИКОВ) ==========
# График A минимум для каждого по горизонтали — подписываются (x_labels первой таблицы)
ARROWS_A_FIRST = dict(angle_deg=180, length_y=0.3, tip_gap_y=0.00, text_gap_y=0.06, lw=1.0, fontsize=9, box=True)
ARROWS_A_LAST  = dict(angle_deg=0, length_y=0.3, tip_gap_y=0.00, text_gap_y=0.06, lw=1.0, fontsize=9, box=True)

# График B минимум по вертикали — подписываются (y_labels первой таблицы)
ARROWS_B_FIRST = dict(angle_deg=180, length_y=0.3, tip_gap_y=0.00, text_gap_y=0.05, lw=1.0, fontsize=9, box=True)
ARROWS_B_LAST  = dict(angle_deg=0, length_y=0.35, tip_gap_y=0.00, text_gap_y=0.05, lw=1.0, fontsize=9, box=True)


# =========================
# Data model
# =========================

@dataclass
class Table2D:
    x_labels: List[str]
    y_labels: List[str]
    values: np.ndarray


# =========================
# Parsing helpers (tabs)
# =========================

def _to_float_ru(s: str) -> float:
    s = str(s).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    return float(s)

def _split_blocks_by_tabs(line: str) -> List[str]:
    return [b for b in re.split(r"\t{2,}", line.rstrip("\n")) if b.strip()]

def _split_cells_by_tabs(block: str) -> List[str]:
    return [c for c in re.split(r"\t+", block.strip()) if c.strip()]

def parse_tables_from_paste(paste_text: str) -> List[Table2D]:
    lines = [ln for ln in paste_text.splitlines() if ln.strip()]
    if not lines:
        raise ValueError("Empty input.")

    header_blocks = _split_blocks_by_tabs(lines[0])
    if not header_blocks:
        raise ValueError("Header line could not be split into table blocks. Keep tabs in the input.")

    x_labels_per_table: List[List[str]] = []
    for hb in header_blocks:
        cells = _split_cells_by_tabs(hb)
        x_labels_per_table.append([c.strip() for c in cells])

    n_tables = len(x_labels_per_table)
    n_cols = len(x_labels_per_table[0])
    if any(len(x) != n_cols for x in x_labels_per_table):
        raise ValueError("Not all tables have the same number of X columns.")

    y_labels: List[str] = []
    values_per_table: List[List[List[float]]] = [[] for _ in range(n_tables)]

    for ln in lines[1:]:
        blocks = _split_blocks_by_tabs(ln)
        if len(blocks) != n_tables:
            raise ValueError(
                f"Line has {len(blocks)} blocks, expected {n_tables}. "
                f"Make sure table gaps are preserved as multiple TABs."
            )

        row_y: Optional[str] = None
        for ti, blk in enumerate(blocks):
            cells = _split_cells_by_tabs(blk)
            if len(cells) != (1 + n_cols):
                raise ValueError(
                    f"Table {ti+1}: row has {len(cells)} cells, expected {1+n_cols} (y + {n_cols} values)."
                )

            yv = cells[0].strip()
            vals = [_to_float_ru(c) for c in cells[1:]]

            if row_y is None:
                row_y = yv
            elif row_y != yv:
                raise ValueError("Y labels differ across tables on the same row; input misaligned.")

            values_per_table[ti].append(vals)

        y_labels.append(str(row_y))

    tables: List[Table2D] = []
    for ti in range(n_tables):
        arr = np.array(values_per_table[ti], dtype=float)
        tables.append(Table2D(x_labels=x_labels_per_table[ti], y_labels=y_labels, values=arr))

    return tables


# =========================
# Excel reading
# =========================

def _block_range_for_graph_index(base_range: str, graph_index: int, gap_rows: int) -> Tuple[str, int, int, int, int]:
    if graph_index < 1:
        raise ValueError("graph_index must be >= 1")

    min_col, min_row, max_col, max_row = range_boundaries(base_range)
    height = (max_row - min_row + 1)

    shift = (graph_index - 1) * (height + gap_rows)
    new_min_row = min_row + shift
    new_max_row = max_row + shift

    a1 = f"{get_column_letter(min_col)}{new_min_row}:{get_column_letter(max_col)}{new_max_row}"
    return a1, min_col, new_min_row, max_col, new_max_row


def _cells_to_tabbed_text(ws: Worksheet, min_col: int, min_row: int, max_col: int, max_row: int) -> str:
    lines = []
    for r in range(min_row, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append("" if v is None else str(v))
        lines.append("\t".join(row_vals))
    return "\n".join(lines)


def parse_tables_from_excel(xlsx_path: str, sheet_name: str, a1_range: str) -> Tuple[List[Table2D], str]:
    wb = load_workbook(xlsx_path, data_only=True)  # значения формул
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

        ws = wb[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(a1_range)

        # y-axis label from A(row_start-1)
        y_title_row = min_row - 1
        y_title = ""
        if y_title_row >= 1:
            v = ws.cell(row=y_title_row, column=1).value  # A
            y_title = "" if v is None else str(v).strip()

        block_text = _cells_to_tabbed_text(ws, min_col, min_row, max_col, max_row)
        tables = parse_tables_from_paste(block_text)
        return tables, y_title
    finally:
        wb.close()


# =========================
# Reductions
# =========================

def min_over_capacity_per_current(table: Table2D) -> pd.Series:
    mins = np.min(table.values, axis=0)
    return pd.Series(mins, index=table.x_labels, name="min_over_capacity")

def min_over_current_per_capacity(table: Table2D) -> pd.Series:
    mins = np.min(table.values, axis=1)
    return pd.Series(mins, index=table.y_labels, name="min_over_current")


# =========================
# Annotation (polar)
# =========================

def _annotate_point_polar(
    ax,
    x_point: float,
    y_point: float,
    text: str,
    angle_deg: float,
    length_y: float,
    tip_gap_y: float,
    text_gap_y: float,
    lw: float = 1.0,
    fontsize: int = 9,
    box: bool = True,
):
    xmin, xmax = ax.get_xlim()
    ymin, ymax = ax.get_ylim()
    xr = (xmax - xmin) if xmax > xmin else 1.0
    yr = (ymax - ymin) if ymax > ymin else 1.0

    # поправка, чтобы угол был "геометрический" на экране
    k = xr / yr
    theta = np.deg2rad(angle_deg)

    ux = np.cos(theta) * k
    uy = np.sin(theta)

    norm = np.hypot(ux, uy)
    ux /= norm
    uy /= norm

    # наконечник рядом с точкой (с небольшим отступом от центра точки)
    x_tip = x_point - ux * tip_gap_y
    y_tip = y_point - uy * tip_gap_y

    # хвост стрелки
    x_tail = x_tip + ux * length_y
    y_tail = y_tip + uy * length_y

    # текст дальше хвоста
    x_text = x_tail + ux * text_gap_y
    y_text = y_tail + uy * text_gap_y

    bbox = dict(boxstyle="round,pad=0.15", facecolor="white", edgecolor="none", alpha=0.85) if box else None

    ax.annotate(
        text,
        xy=(x_tip, y_tip),
        xytext=(x_text, y_text),
        textcoords="data",
        ha="center",
        va="center",
        fontsize=fontsize,
        color="black",
        bbox=bbox,
        arrowprops=dict(arrowstyle="->", color="black", lw=lw),
        clip_on=True,
    )


# =========================
# Plotting
# =========================

def _infer_double_labels(labels: Optional[List[str]], n: int) -> Tuple[List[str], List[str]]:
    if labels is None:
        base = [f"T{i+1}" for i in range(n)]
        return base, base

    top, bottom = [], []
    for lab in labels:
        parts = str(lab).strip().split()
        if len(parts) >= 2:
            top.append(parts[0])
            bottom.append(parts[1])
        else:
            top.append(lab)
            bottom.append(lab)
    return top, bottom


def _plot_groups_points_double_x(
    ax,
    groups,
    top_labels: List[str],
    bottom_labels: List[str],
    title: str,
    y_axis_title: str,
    top_axis_name: str = "Мощность ВЭУ",
    bottom_axis_name: str = "Схема",
    jitter: float = 0.08,
    show_median: bool = True,
    show_minmax: bool = True,
    group_mode: str = "fixed",   # fixed | none
    group_size: Optional[int] = None,
    group_separators: bool = True,
    annotate_first_group: bool = False,
    level_labels_for_annotation: Optional[List[str]] = None,
    arrow_first: Optional[Dict[str, Any]] = None,
    arrow_last: Optional[Dict[str, Any]] = None,
):
    n = len(groups)
    x = np.arange(n, dtype=float)

    # =========================
    # COLORS: number of colors = graphs per group
    # =========================
    if group_mode == "fixed":
        if not group_size or group_size < 1:
            raise ValueError("group_size must be >= 1 when group_mode='fixed'")
        colors_count = group_size
    elif group_mode == "none":
        colors_count = n
    else:
        raise ValueError("group_mode must be one of: 'fixed', 'none'")

    base_colors = plt.rcParams["axes.prop_cycle"].by_key().get("color", ["C0", "C1", "C2", "C3"])
    colors = [base_colors[i % len(base_colors)] for i in range(colors_count)]

    # =========================
    # points (COLOR) + minmax + median (BLACK)
    # =========================
    for i, vals in enumerate(groups):
        vals = np.asarray(vals, dtype=float)
        if vals.size == 0:
            continue

        point_color = colors[i % colors_count]

        offs = np.linspace(-jitter, jitter, num=len(vals)) if len(vals) > 1 else np.array([0.0])

        # --- points: COLORED ---
        ax.scatter(
            np.full_like(vals, x[i]) + offs,
            vals,
            color=point_color
        )

        # --- min–max: ALWAYS BLACK ---
        if show_minmax and len(vals) > 1:
            ax.vlines(
                x[i],
                float(np.min(vals)),
                float(np.max(vals)),
                color="black"
            )

        # --- median: ALWAYS BLACK ---
        if show_median:
            med = float(np.median(vals))
            ax.hlines(
                med,
                x[i] - 0.18,
                x[i] + 0.18,
                color="black"
            )

    # =========================
    # X labels (two rows)
    # =========================
    ax.set_xticks(x)
    ax.set_xticklabels(bottom_labels)

    ax_bottom2 = ax.secondary_xaxis("bottom")
    ax_bottom2.spines["bottom"].set_position(("outward", 20))

    centers, uniq_top = [], []
    i = 0
    while i < len(top_labels):
        label = top_labels[i]
        j = i
        while j < len(top_labels) and top_labels[j] == label:
            j += 1
        centers.append((i + (j - 1)) / 2.0)
        uniq_top.append(label)
        i = j

    ax_bottom2.set_xticks(centers)
    ax_bottom2.set_xticklabels(uniq_top)

    # row labels aligned with tick LABEL rows
    x_left = -0.04
    y_row1 = -0.10
    y_row2 = -0.22
    ax.text(x_left, y_row1, bottom_axis_name, transform=ax.transAxes,
            ha="right", va="center", color="black", clip_on=False)
    ax.text(x_left, y_row2, top_axis_name, transform=ax.transAxes,
            ha="right", va="center", color="black", clip_on=False)

    # =========================
    # group separators (BLACK)
    # =========================
    if group_separators and group_mode != "none":
        lw = float(ax.spines["left"].get_linewidth() or 1.0)
        for k in range(group_size, n, group_size):
            ax.axvline(k - 0.5, color="black", linewidth=lw)

    ax.set_title(title)
    ax.set_ylabel(y_axis_title if y_axis_title else "")
    ax.grid(True, axis="y")

    # =========================
    # annotations (arrows)
    # =========================
    if annotate_first_group and level_labels_for_annotation is not None:
        vals0 = np.asarray(groups[0], dtype=float)
        if vals0.size >= 2 and len(level_labels_for_annotation) >= 2 and arrow_first and arrow_last:
            offs0 = np.linspace(-jitter, jitter, num=len(vals0)) if len(vals0) > 1 else np.array([0.0])

            x_first = 0.0 + float(offs0[0])
            y_first = float(vals0[0])
            x_last = 0.0 + float(offs0[-1])
            y_last = float(vals0[-1])

            t_first = str(level_labels_for_annotation[0])
            t_last = str(level_labels_for_annotation[-1])

            ax.figure.canvas.draw()

            _annotate_point_polar(ax, x_first, y_first, t_first, **arrow_first)
            _annotate_point_polar(ax, x_last, y_last, t_last, **arrow_last)


def build_two_figures_from_tables(
    tables: List[Table2D],
    y_axis_title: str,
    scenario_labels: Optional[List[str]] = None,
    top_axis_name: str = "Мощность ВЭУ",
    bottom_axis_name: str = "Схема",
    group_mode: str = "fixed",
    group_size: Optional[int] = None,
    group_separators: bool = True,
    arrows_A_first: Optional[Dict[str, Any]] = None,
    arrows_A_last: Optional[Dict[str, Any]] = None,
    arrows_B_first: Optional[Dict[str, Any]] = None,
    arrows_B_last: Optional[Dict[str, Any]] = None,
):
    n = len(tables)

    groups_A = [min_over_capacity_per_current(t).values for t in tables]
    groups_B = [min_over_current_per_capacity(t).values for t in tables]

    top_labels, bottom_labels = _infer_double_labels(scenario_labels, n)

    first_currents = list(tables[0].x_labels)
    first_capacities = list(tables[0].y_labels)

    figA, axA = plt.subplots(figsize=(max(10, 1.2*n), 4))
    _plot_groups_points_double_x(
        axA,
        groups_A,
        top_labels=top_labels,
        bottom_labels=bottom_labels,
        title="",
        y_axis_title=y_axis_title,
        top_axis_name=top_axis_name,
        bottom_axis_name=bottom_axis_name,
        group_mode=group_mode,
        group_size=group_size,
        group_separators=group_separators,
        annotate_first_group=True,
        level_labels_for_annotation=first_currents,
        arrow_first=arrows_A_first,
        arrow_last=arrows_A_last,
    )
    plt.tight_layout(rect=(0, 0.18, 1, 1))

    figB, axB = plt.subplots(figsize=(max(10, 1.2*n), 4))
    _plot_groups_points_double_x(
        axB,
        groups_B,
        top_labels=top_labels,
        bottom_labels=bottom_labels,
        title="",
        y_axis_title=y_axis_title,
        top_axis_name=top_axis_name,
        bottom_axis_name=bottom_axis_name,
        group_mode=group_mode,
        group_size=group_size,
        group_separators=group_separators,
        annotate_first_group=True,
        level_labels_for_annotation=first_capacities,
        arrow_first=arrows_B_first,
        arrow_last=arrows_B_last,
    )
    plt.tight_layout(rect=(0, 0.18, 1, 1))

    return figA, figB


# =========================
# Console input (ONLY what you want)
# =========================

def read_grouping_from_console() -> tuple[str, Optional[int], bool]:
    s = input("Grouping mode [fixed/none] (default fixed): ").strip().lower()
    if not s:
        s = "fixed"
    if s not in {"fixed", "none"}:
        raise ValueError("Grouping mode must be: fixed, none")

    group_size = None
    if s == "fixed":
        s_gs = input("Enter group size (default 2): ").strip()
        group_size = int(s_gs) if s_gs else 2
        if group_size < 1:
            raise ValueError("group size must be >= 1")

    sep_in = input("Draw group separators? [y/n] (default y): ").strip().lower()
    group_separators = (sep_in != "n")

    return s, group_size, group_separators


# =========================
# Main
# =========================

if __name__ == "__main__":
    xlsx_path = DEFAULT_EXCEL_PATH
    sheet_name = DEFAULT_SHEET

    graph_idx_s = input("Graph index (1..N) (default 1): ").strip()
    graph_idx = int(graph_idx_s) if graph_idx_s else 1
    if graph_idx < 1:
        raise ValueError("Graph index must be >= 1")

    block_range, _, start_row, _, _ = _block_range_for_graph_index(
        DEFAULT_BLOCK_RANGE, graph_idx, DEFAULT_GAP_ROWS
    )

    print(f"Using block range: {block_range}")
    print(f"Y-axis title cell: A{start_row - 1}")

    tables, y_axis_title = parse_tables_from_excel(xlsx_path, sheet_name, block_range)

    labels = DEFAULT_SCENARIO_LABELS if USE_DEFAULT_LABELS else None
    if labels is not None and len(labels) != len(tables):
        raise ValueError(
            f"DEFAULT_SCENARIO_LABELS has {len(labels)} items, but block contains {len(tables)} tables."
        )

    group_mode, group_size, group_separators = read_grouping_from_console()

    build_two_figures_from_tables(
        tables,
        y_axis_title=y_axis_title,
        scenario_labels=labels,
        top_axis_name=TOP_AXIS_NAME,
        bottom_axis_name=BOTTOM_AXIS_NAME,
        group_mode=group_mode,
        group_size=group_size,
        group_separators=group_separators,
        arrows_A_first=ARROWS_A_FIRST,
        arrows_A_last=ARROWS_A_LAST,
        arrows_B_first=ARROWS_B_FIRST,
        arrows_B_last=ARROWS_B_LAST,
    )
    plt.show()
