import os
import re
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from openpyxl import load_workbook

# =========================
# НАСТРОЙКИ
# =========================
EXCEL_PATH = r"D:\comb_results.xlsx"
SHEET_NAME = "SWEEP_2"
OUTPUT_DIR = r"D:\results"

SCHEME_1_NAME = "Секционированная система"
SCHEME_2_NAME = "Двойная система"

PEAK_LOAD_KW = 1346.0
WT_COUNT = 2
SIM_YEARS = 20

# Обрезка данных для построения графиков:
# (сколько убрать сначала, сколько убрать с конца)
# Примеры:
# X_TRIM = (0, 0) -> ничего не обрезать
# X_TRIM = (1, 2) -> убрать 1 первое значение по X и 2 последних
# Y_TRIM = (0, 1) -> убрать 1 последнее значение по Y
X_TRIM = (0, 0)
Y_TRIM = (0, 0)

METRICS = [
    ("LCOE", "LCOE, руб/кВт∙ч"),
    ("LPSP", "LPSP"),
    ("LOLP", "LOLP"),
    ("ENS_evtN", "Кол-во событий ENS"),
    ("ENS_evtMaxH", "Макс. длит. ENS, ч"),
]

HEADER_ALIASES = {
    "LCOE": [
        "LCOE",
        "LCOE,руб/кВт∙ч",
        "LCOE, руб/кВт∙ч",
        "LCOE,руб/кВт·ч",
        "LCOE, руб/кВт·ч",
        "LCOE,руб/кВт*ч",
        "LCOE, руб/кВт*ч",
    ],
    "ENS": [
        "ENS",
        "ENS,кВт∙ч",
        "ENS, кВт∙ч",
        "ENS,кВт·ч",
        "ENS, кВт·ч",
        "ENS,кВт*ч",
        "ENS, кВт*ч",
    ],
    "LOLH": [
        "LOLH",
    ],
    "LOLP": [
        "LOLP",
    ],
    "LPSP": [
        "LPSP",
    ],
    "ENS_evtN": [
        "ENS_evtN",
    ],
    "ENS_evtMaxH": [
        "ENS_evtMaxH",
        "ENS_evt_MaxH",
        "ENS_evtMaxH ",
        "ENS_evt_MaxH ",
    ],
}

# =========================
# СПОКОЙНАЯ МАТОВАЯ ПАЛИТРА
# =========================
MATTE_METRIC_CMAP = LinearSegmentedColormap.from_list(
    "matte_metric",
    [
        "#f2f2f2",
        "#cfd8dc",
        "#8ea9b8",
        "#5f86a2",
        "#426f91",
    ]
)

MATTE_DIFF_CMAP = LinearSegmentedColormap.from_list(
    "matte_diff",
    [
        "#4d7a99",
        "#8fa8b7",
        "#e9e9e9",
        "#ddb48c",
        "#c98d5b",
    ]
)

GRID_COLOR = "#ffffff"
GRID_ALPHA = 0.35
TITLE_SIZE = 12
LABEL_SIZE = 12
TICK_SIZE = 11


# =========================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================
def normalize_text(x):
    if x is None:
        return ""
    return (
        str(x)
        .strip()
        .replace(" ", "")
        .replace("·", "∙")
        .replace("*", "∙")
    )


def is_empty(v):
    return v is None or str(v).strip() == ""


def to_float(v):
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if s == "":
        return np.nan
    try:
        return float(s)
    except ValueError:
        return np.nan


def safe_filename(name):
    """
    Делает безопасное имя файла:
    убирает недопустимые символы \\ / : * ? " < > |
    """
    s = str(name).strip()
    s = re.sub(r'[\\/:*?"<>|]+', "_", s)
    s = s.replace("\n", "_").replace("\r", "_")
    return s


def validate_trim(trim, axis_name):
    if not isinstance(trim, (tuple, list)) or len(trim) != 2:
        raise ValueError(f"{axis_name}_TRIM должен быть парой из 2 чисел, например (1, 2)")
    start_trim, end_trim = trim
    if not isinstance(start_trim, int) or not isinstance(end_trim, int):
        raise ValueError(f"{axis_name}_TRIM должен содержать целые числа")
    if start_trim < 0 or end_trim < 0:
        raise ValueError(f"{axis_name}_TRIM не может содержать отрицательные значения")
    return start_trim, end_trim


def apply_axis_trim(x_raw, y_raw, z, x_trim=(0, 0), y_trim=(0, 0)):
    """
    Обрезает данные для построения графика:
    x_trim = (убрать_слева, убрать_справа)
    y_trim = (убрать_сверху, убрать_снизу)
    """
    x_left, x_right = validate_trim(x_trim, "X")
    y_top, y_bottom = validate_trim(y_trim, "Y")

    x_len = len(x_raw)
    y_len = len(y_raw)

    if x_left + x_right >= x_len:
        raise ValueError(
            f"Слишком большая обрезка по X: {x_trim}, доступно значений: {x_len}"
        )

    if y_top + y_bottom >= y_len:
        raise ValueError(
            f"Слишком большая обрезка по Y: {y_trim}, доступно значений: {y_len}"
        )

    x_end = x_len - x_right if x_right > 0 else x_len
    y_end = y_len - y_bottom if y_bottom > 0 else y_len

    x_raw_cut = x_raw[x_left:x_end]
    y_raw_cut = y_raw[y_top:y_end]
    z_cut = z[y_top:y_end, x_left:x_end]

    return x_raw_cut, y_raw_cut, z_cut


def find_metric_cells(ws, metric_key):
    """
    Ищет все ячейки с подписью метрики на листе.
    Возвращает список (row, col).
    """
    aliases_norm = {normalize_text(a) for a in HEADER_ALIASES[metric_key]}
    found = []

    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = normalize_text(ws.cell(r, c).value)
            if v in aliases_norm:
                found.append((r, c))

    if not found:
        raise ValueError(f"Не найдена метрика: {metric_key}")

    return found


def extract_block_from_label(ws, label_row, label_col):
    """
    Структура блока:
      (label_row, label_col)       : подпись критерия
      (label_row + 1, label_col+1) : начало X-подписей
      (label_row + 2, label_col)   : начало Y-подписей
      (label_row + 2, label_col+1) : начало матрицы

    Конец X — первая пустая ячейка вправо в строке X.
    Конец Y — первая пустая ячейка вниз в колонке Y.
    """
    x_row = label_row + 1
    y_start_row = label_row + 2
    x_start_col = label_col + 1
    y_col = label_col

    x_cols = []
    c = x_start_col
    while c <= ws.max_column:
        v = ws.cell(x_row, c).value
        if is_empty(v):
            break
        x_cols.append(c)
        c += 1

    if not x_cols:
        raise ValueError(
            f"У блока {ws.cell(label_row, label_col).coordinate} не найдены подписи X"
        )

    y_rows = []
    r = y_start_row
    while r <= ws.max_row:
        v = ws.cell(r, y_col).value
        if is_empty(v):
            break
        y_rows.append(r)
        r += 1

    if not y_rows:
        raise ValueError(
            f"У блока {ws.cell(label_row, label_col).coordinate} не найдены подписи Y"
        )

    x_raw = [to_float(ws.cell(x_row, c).value) for c in x_cols]
    y_raw = [to_float(ws.cell(r, y_col).value) for r in y_rows]

    z = np.array(
        [[to_float(ws.cell(r, c).value) for c in x_cols] for r in y_rows],
        dtype=float
    )

    return x_raw, y_raw, z


def extract_two_scheme_blocks(ws, metric_key):
    """
    Для одной метрики ищет все блоки по подписи метрики.
    Берет первые два блока слева направо / сверху вниз.
    """
    metric_cells = find_metric_cells(ws, metric_key)
    metric_cells = sorted(metric_cells, key=lambda rc: (rc[0], rc[1]))

    if len(metric_cells) < 2:
        raise ValueError(
            f"Для метрики {metric_key} найдено меньше двух блоков: {metric_cells}"
        )

    (r1, c1), (r2, c2) = metric_cells[:2]

    x1_raw, y1_raw, z1 = extract_block_from_label(ws, r1, c1)
    x2_raw, y2_raw, z2 = extract_block_from_label(ws, r2, c2)

    return x1_raw, y1_raw, z1, x2_raw, y2_raw, z2


def transform_metric(metric_key, arr):
    arr = arr.astype(float).copy()

    if metric_key == "LOLH":
        arr = arr / SIM_YEARS

    return arr


def battery_capacity_labels(x_raw):
    return [f"{int(round(v * 1))}" if np.isfinite(v) else "" for v in x_raw] #*100


def wt_power_percent_labels(y_raw, wt_count=WT_COUNT, peak_load_kw=PEAK_LOAD_KW):
    labels = []
    for v in y_raw:
        if np.isfinite(v):
            labels.append(f"{int(round(v * wt_count / peak_load_kw * 100))}")
        else:
            labels.append("")
    return labels


def safe_rel_diff(a, b):
    """
    Δ% = (b - a) / a * 100
    Если a == 0 -> nan
    """
    out = np.full_like(a, np.nan, dtype=float)
    mask = np.abs(a) > 1e-12
    out[mask] = (a[mask] - b[mask]) / a[mask] * 100.0
    return out


def setup_heatmap_axis(ax, xlabels, ylabels, title, show_xlabel=False, show_ylabel=False):
    ax.set_title(title, fontsize=TITLE_SIZE, fontweight="bold")

    ax.set_xticks(np.arange(len(xlabels)))
    ax.set_xticklabels(
        xlabels,
        fontsize=TICK_SIZE,
        rotation=90,
        ha="center",
        va="top"
    )

    ax.set_yticks(np.arange(len(ylabels)))
    ax.set_yticklabels(ylabels, fontsize=TICK_SIZE)

    if show_xlabel:
        ax.set_xlabel("Мощность ДГУ, кВт", fontsize=LABEL_SIZE)
    if show_ylabel:
        ax.set_ylabel("Мощность ВЭУ, %", fontsize=LABEL_SIZE)

    ax.set_xticks(np.arange(-0.5, len(xlabels), 1), minor=True)
    ax.set_yticks(np.arange(-0.5, len(ylabels), 1), minor=True)
    ax.grid(which="minor", color=GRID_COLOR, linestyle="-", linewidth=0.7, alpha=GRID_ALPHA)
    ax.tick_params(which="minor", bottom=False, left=False)
    ax.tick_params(axis="x", pad=2)


def save_individual_figure(metric_label, z1, z2, diff, xlabels, ylabels, out_path):
    fig, axes = plt.subplots(1, 3, figsize=(16, 4.8), constrained_layout=True)

    common_vmin = np.nanmin([np.nanmin(z1), np.nanmin(z2)])
    common_vmax = np.nanmax([np.nanmax(z1), np.nanmax(z2)])

    axes[0].imshow(
        z1,
        aspect="auto",
        origin="upper",
        vmin=common_vmin,
        vmax=common_vmax,
        cmap=MATTE_METRIC_CMAP
    )
    setup_heatmap_axis(
        axes[0], xlabels, ylabels,
        f"{metric_label} — {SCHEME_1_NAME}",
        show_xlabel=True, show_ylabel=True
    )

    im2 = axes[1].imshow(
        z2,
        aspect="auto",
        origin="upper",
        vmin=common_vmin,
        vmax=common_vmax,
        cmap=MATTE_METRIC_CMAP
    )
    setup_heatmap_axis(
        axes[1], xlabels, ylabels,
        f"{metric_label} — {SCHEME_2_NAME}",
        show_xlabel=True, show_ylabel=False
    )

    lim = np.nanmax(np.abs(diff))
    if not np.isfinite(lim) or lim == 0:
        lim = 1.0

    im3 = axes[2].imshow(
        diff,
        aspect="auto",
        origin="upper",
        vmin=-lim,
        vmax=lim,
        cmap=MATTE_DIFF_CMAP
    )
    setup_heatmap_axis(
        axes[2], xlabels, ylabels,
        f"{metric_label} — Δ, %",
        show_xlabel=True, show_ylabel=False
    )

    cbar1 = fig.colorbar(im2, ax=axes[:2], fraction=0.025, pad=0.02)
    cbar1.ax.set_ylabel(metric_label, rotation=90, fontsize=LABEL_SIZE)
    cbar1.ax.tick_params(labelsize=TICK_SIZE)

    cbar2 = fig.colorbar(im3, ax=axes[2], fraction=0.046, pad=0.04)
    cbar2.ax.set_ylabel("Δ, %", rotation=90, fontsize=LABEL_SIZE)
    cbar2.ax.tick_params(labelsize=TICK_SIZE)

    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    prepared = []

    for metric_key, metric_label in METRICS:
        x1_raw, y1_raw, z1, x2_raw, y2_raw, z2 = extract_two_scheme_blocks(ws, metric_key)

        z1 = transform_metric(metric_key, z1)
        z2 = transform_metric(metric_key, z2)

        if len(x1_raw) != len(x2_raw) or len(y1_raw) != len(y2_raw):
            raise ValueError(f"Размер сеток у схем не совпадает для метрики {metric_key}")

        if not np.allclose(np.array(x1_raw), np.array(x2_raw), equal_nan=True):
            raise ValueError(f"Ось X у схем не совпадает для метрики {metric_key}")

        if not np.allclose(np.array(y1_raw), np.array(y2_raw), equal_nan=True):
            raise ValueError(f"Ось Y у схем не совпадает для метрики {metric_key}")

        x1_raw, y1_raw, z1 = apply_axis_trim(x1_raw, y1_raw, z1, x_trim=X_TRIM, y_trim=Y_TRIM)
        x2_raw, y2_raw, z2 = apply_axis_trim(x2_raw, y2_raw, z2, x_trim=X_TRIM, y_trim=Y_TRIM)

        xlabels = battery_capacity_labels(x1_raw)
        ylabels = wt_power_percent_labels(y1_raw)

        diff = safe_rel_diff(z1, z2)

        prepared.append({
            "metric_key": metric_key,
            "metric_label": metric_label,
            "xlabels": xlabels,
            "ylabels": ylabels,
            "z1": z1,
            "z2": z2,
            "diff": diff,
        })

        safe_name = safe_filename(metric_key)

        save_individual_figure(
            metric_label=metric_label,
            z1=z1,
            z2=z2,
            diff=diff,
            xlabels=xlabels,
            ylabels=ylabels,
            out_path=os.path.join(OUTPUT_DIR, f"{safe_name}_compare.png")
        )

    fig, axes = plt.subplots(
        len(prepared), 3,
        figsize=(17, 4.2 * len(prepared)),
        constrained_layout=True
    )

    if len(prepared) == 1:
        axes = np.array([axes])

    for i, item in enumerate(prepared):
        z1 = item["z1"]
        z2 = item["z2"]
        diff = item["diff"]
        xlabels = item["xlabels"]
        ylabels = item["ylabels"]
        metric_label = item["metric_label"]

        common_vmin = np.nanmin([np.nanmin(z1), np.nanmin(z2)])
        common_vmax = np.nanmax([np.nanmax(z1), np.nanmax(z2)])

        ax = axes[i, 0]
        ax.imshow(
            z1,
            aspect="auto",
            origin="upper",
            vmin=common_vmin,
            vmax=common_vmax,
            cmap=MATTE_METRIC_CMAP
        )
        setup_heatmap_axis(
            ax, xlabels, ylabels,
            f"{metric_label} — {SCHEME_1_NAME}",
            show_xlabel=(i == len(prepared) - 1),
            show_ylabel=True
        )

        ax = axes[i, 1]
        im_mid = ax.imshow(
            z2,
            aspect="auto",
            origin="upper",
            vmin=common_vmin,
            vmax=common_vmax,
            cmap=MATTE_METRIC_CMAP
        )
        setup_heatmap_axis(
            ax, xlabels, ylabels,
            f"{metric_label} — {SCHEME_2_NAME}",
            show_xlabel=(i == len(prepared) - 1),
            show_ylabel=False
        )

        ax = axes[i, 2]
        lim = np.nanmax(np.abs(diff))
        if not np.isfinite(lim) or lim == 0:
            lim = 1.0

        im_right = ax.imshow(
            diff,
            aspect="auto",
            origin="upper",
            vmin=-lim,
            vmax=lim,
            cmap=MATTE_DIFF_CMAP
        )
        setup_heatmap_axis(
            ax, xlabels, ylabels,
            f"{metric_label} — Δ, %",
            show_xlabel=(i == len(prepared) - 1),
            show_ylabel=False
        )

        cbar_left = fig.colorbar(im_mid, ax=[axes[i, 0], axes[i, 1]], fraction=0.018, pad=0.015)
        cbar_left.ax.set_ylabel(metric_label, rotation=90, fontsize=LABEL_SIZE)
        cbar_left.ax.tick_params(labelsize=TICK_SIZE)

        cbar_right = fig.colorbar(im_right, ax=axes[i, 2], fraction=0.046, pad=0.02)
        cbar_right.ax.set_ylabel("Δ, %", rotation=90, fontsize=LABEL_SIZE)
        cbar_right.ax.tick_params(labelsize=TICK_SIZE)

    fig.suptitle(
        ""
        "",
        fontsize=16,
        fontweight="bold"
    )

    summary_path = os.path.join(OUTPUT_DIR, "reliability_compare_summary.png")
    fig.savefig(summary_path, dpi=300, bbox_inches="tight")
    plt.close(fig)

    print("Готово.")
    print("Размеры матриц определялись автоматически по Excel.")
    print(f"Сводная фигура: {summary_path}")
    print(f"Отдельные фигуры сохранены в папку: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()