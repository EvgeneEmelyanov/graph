import os
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap
from openpyxl import load_workbook

# =========================
# НАСТРОЙКИ
# =========================
EXCEL_PATH = r"D:\comb_results.xlsx"
SHEET_NAME = "SWEEP_2"
OUTPUT_DIR = r"D:\results"

SCHEME_1_NAME = "Статический уровень"
SCHEME_2_NAME = "Адаптивный уровень"

MATRIX_ROWS = 11
MATRIX_COLS = 9

LEFT_Y_COL = 1
LEFT_DATA_START_COL = 2

RIGHT_Y_COL = 12
RIGHT_DATA_START_COL = 13

ECON_METRICS = [
    ("LCOE", "LCOE, руб/кВт∙ч"),
    ("Fuel", "Расход топлива, тыс.тонн"),
    ("Hours", "Моточасы, тыс.мч"),
]

FAIL_METRICS = [
    ("FailDg", "FailDg"),
    ("FailBt", "FailBt"),
]

HEADER_ALIASES = {
    "LCOE": [
        "LCOE",
        "LCOE, руб/кВт∙ч",
        "LCOE, руб/кВт·ч",
    ],
    "Fuel": [
        "Расход топлива, тыс.тонн",
    ],
    "Hours": [
        "Моточасы, тыс.мч",
    ],
    "FailDg": [
        "FailDg",
    ],
    "FailBt": [
        "FailBt",
    ],
}

# =========================
# ПАЛИТРЫ
# =========================
MATTE_METRIC_CMAP = LinearSegmentedColormap.from_list(
    "matte_metric",
    ["#f2f2f2", "#cfd8dc", "#8ea9b8", "#5f86a2", "#426f91"]
)

MATTE_DIFF_CMAP = LinearSegmentedColormap.from_list(
    "matte_diff",
    ["#4d7a99", "#8fa8b7", "#e9e9e9", "#ddb48c", "#c98d5b"]
)

GRID_COLOR = "#ffffff"
GRID_ALPHA = 0.35
TITLE_SIZE = 12
LABEL_SIZE = 10
TICK_SIZE = 9

# =========================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# =========================
def normalize_text(x):
    if x is None:
        return ""
    return str(x).strip().replace(" ", "").replace("·", "∙")


def to_float(v):
    if v is None:
        return np.nan
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(",", ".")
    if s == "":
        return np.nan
    return float(s)


def format_axis_value(v):
    if v is None:
        return ""
    if isinstance(v, (float, np.floating)) and np.isnan(v):
        return ""
    if isinstance(v, (int, np.integer)):
        return str(v)
    if isinstance(v, (float, np.floating)):
        if abs(v - round(v)) < 1e-9:
            return str(int(round(v)))
        return f"{v:g}"
    return str(v).strip()


def axis_labels_from_excel(values):
    return [format_axis_value(v) for v in values]


def col_range(start_col, count):
    return range(start_col, start_col + count)


def row_range(start_row, count):
    return range(start_row, start_row + count)


def find_metric_start_row(ws, metric_key):
    aliases_norm = {normalize_text(a) for a in HEADER_ALIASES[metric_key]}

    for r in range(1, ws.max_row + 1):
        v_left = normalize_text(ws.cell(r, 1).value)
        v_right = normalize_text(ws.cell(r, RIGHT_Y_COL).value)
        if v_left in aliases_norm or v_right in aliases_norm:
            return r

    raise ValueError(f"Не найден блок для метрики: {metric_key}")


def extract_block(ws, start_row, y_col, data_start_col, matrix_rows, matrix_cols):
    x_row = start_row + 1
    data_row_start = start_row + 2

    x_cols = list(col_range(data_start_col, matrix_cols))
    y_rows = list(row_range(data_row_start, matrix_rows))

    x_raw = [to_float(ws.cell(x_row, c).value) for c in x_cols]
    y_raw = [to_float(ws.cell(r, y_col).value) for r in y_rows]

    z = np.array(
        [[to_float(ws.cell(r, c).value) for c in x_cols] for r in y_rows],
        dtype=float
    )

    return x_raw, y_raw, z


def extract_two_scheme_blocks(ws, metric_key):
    start_row = find_metric_start_row(ws, metric_key)

    x1_raw, y1_raw, z1 = extract_block(
        ws=ws,
        start_row=start_row,
        y_col=LEFT_Y_COL,
        data_start_col=LEFT_DATA_START_COL,
        matrix_rows=MATRIX_ROWS,
        matrix_cols=MATRIX_COLS
    )

    x2_raw, y2_raw, z2 = extract_block(
        ws=ws,
        start_row=start_row,
        y_col=RIGHT_Y_COL,
        data_start_col=RIGHT_DATA_START_COL,
        matrix_rows=MATRIX_ROWS,
        matrix_cols=MATRIX_COLS
    )

    return x1_raw, y1_raw, z1, x2_raw, y2_raw, z2


def safe_rel_diff(a, b):
    out = np.full_like(a, np.nan, dtype=float)
    mask = np.abs(a) > 1e-12
    out[mask] = (b[mask] - a[mask]) / a[mask] * 100.0
    return out


def setup_heatmap_axis(ax, xlabels, ylabels, title, show_xlabel=False, show_ylabel=False):
    ax.set_title(title, fontsize=TITLE_SIZE, fontweight="bold")

    ax.set_xticks(np.arange(len(xlabels)))
    ax.set_xticklabels(xlabels, fontsize=TICK_SIZE)

    ax.set_yticks(np.arange(len(ylabels)))
    ax.set_yticklabels(ylabels, fontsize=TICK_SIZE)

    if show_xlabel:
        ax.set_xlabel("Уровень разряда", fontsize=LABEL_SIZE)
    if show_ylabel:
        ax.set_ylabel("Емкость СНЭ", fontsize=LABEL_SIZE)

    ax.set_xticks(np.arange(-0.5, len(xlabels), 1), minor=True)
    ax.set_yticks(np.arange(-0.5, len(ylabels), 1), minor=True)
    ax.grid(which="minor", color=GRID_COLOR, linestyle="-", linewidth=0.7, alpha=GRID_ALPHA)
    ax.tick_params(which="minor", bottom=False, left=False)


def save_group_figure(prepared, title, out_path):
    fig, axes = plt.subplots(
        len(prepared), 3,
        figsize=(17, 4.2 * len(prepared)),
        constrained_layout=True
    )

    if len(prepared) == 1:
        axes = np.array([axes])

    for i, item in enumerate(prepared):
        z1 = item["z1"]
        z2 = item["z2"]
        diff = item["diff"]
        xlabels = item["xlabels"]
        ylabels = item["ylabels"]
        metric_label = item["metric_label"]

        common_vmin = np.nanmin([np.nanmin(z1), np.nanmin(z2)])
        common_vmax = np.nanmax([np.nanmax(z1), np.nanmax(z2)])

        ax = axes[i, 0]
        ax.imshow(
            z1,
            aspect="auto",
            origin="upper",
            vmin=common_vmin,
            vmax=common_vmax,
            cmap=MATTE_METRIC_CMAP
        )
        setup_heatmap_axis(
            ax, xlabels, ylabels,
            f"{metric_label} — {SCHEME_1_NAME}",
            show_xlabel=(i == len(prepared) - 1),
            show_ylabel=True
        )

        ax = axes[i, 1]
        im_mid = ax.imshow(
            z2,
            aspect="auto",
            origin="upper",
            vmin=common_vmin,
            vmax=common_vmax,
            cmap=MATTE_METRIC_CMAP
        )
        setup_heatmap_axis(
            ax, xlabels, ylabels,
            f"{metric_label} — {SCHEME_2_NAME}",
            show_xlabel=(i == len(prepared) - 1),
            show_ylabel=False
        )

        ax = axes[i, 2]
        lim = np.nanmax(np.abs(diff))
        if not np.isfinite(lim) or lim == 0:
            lim = 1.0

        im_right = ax.imshow(
            diff,
            aspect="auto",
            origin="upper",
            vmin=-lim,
            vmax=lim,
            cmap=MATTE_DIFF_CMAP
        )
        setup_heatmap_axis(
            ax, xlabels, ylabels,
            f"{metric_label} — Δ, %",
            show_xlabel=(i == len(prepared) - 1),
            show_ylabel=False
        )

        cbar_left = fig.colorbar(im_mid, ax=[axes[i, 0], axes[i, 1]], fraction=0.018, pad=0.015)
        cbar_left.ax.set_ylabel(metric_label, rotation=90, fontsize=LABEL_SIZE)
        cbar_left.ax.tick_params(labelsize=TICK_SIZE)

        cbar_right = fig.colorbar(im_right, ax=axes[i, 2], fraction=0.046, pad=0.02)
        cbar_right.ax.set_ylabel("Δ, %", rotation=90, fontsize=LABEL_SIZE)
        cbar_right.ax.tick_params(labelsize=TICK_SIZE)

    fig.suptitle(
        title + "\nПодписи осей взяты напрямую из Excel",
        fontsize=16,
        fontweight="bold"
    )

    fig.savefig(out_path, dpi=300, bbox_inches="tight")
    plt.close(fig)


def prepare_metrics(ws, metric_defs):
    prepared = []

    for metric_key, metric_label in metric_defs:
        x1_raw, y1_raw, z1, x2_raw, y2_raw, z2 = extract_two_scheme_blocks(ws, metric_key)

        if len(x1_raw) != len(x2_raw) or len(y1_raw) != len(y2_raw):
            raise ValueError(f"Размер сеток у схем не совпадает для метрики {metric_key}")

        if not np.allclose(np.array(x1_raw), np.array(x2_raw), equal_nan=True):
            raise ValueError(f"Ось X у схем не совпадает для метрики {metric_key}")

        if not np.allclose(np.array(y1_raw), np.array(y2_raw), equal_nan=True):
            raise ValueError(f"Ось Y у схем не совпадает для метрики {metric_key}")

        xlabels = axis_labels_from_excel(x1_raw)
        ylabels = axis_labels_from_excel(y1_raw)
        diff = safe_rel_diff(z1, z2)

        prepared.append({
            "metric_key": metric_key,
            "metric_label": metric_label,
            "xlabels": xlabels,
            "ylabels": ylabels,
            "z1": z1,
            "z2": z2,
            "diff": diff,
        })

    return prepared


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[SHEET_NAME]

    econ_prepared = prepare_metrics(ws, ECON_METRICS)
    fail_prepared = prepare_metrics(ws, FAIL_METRICS)

    save_group_figure(
        prepared=econ_prepared,
        title="Сравнение схем по экономическим показателям",
        out_path=os.path.join(OUTPUT_DIR, "economics_compare_summary.png")
    )

    save_group_figure(
        prepared=fail_prepared,
        title="Сравнение схем по показателям отказов",
        out_path=os.path.join(OUTPUT_DIR, "fail_compare_summary.png")
    )

    print("Готово.")
    print(os.path.join(OUTPUT_DIR, "economics_compare_summary.png"))
    print(os.path.join(OUTPUT_DIR, "economics_compare_summary.png"))
    print(os.path.join(OUTPUT_DIR, "fail_compare_summary.png"))


if __name__ == "__main__":
    main()